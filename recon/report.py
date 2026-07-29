# -*- coding: utf-8 -*-
"""對帳結果輸出：終端機文字報表 + Excel 匯出"""
from __future__ import annotations

import io


def to_text(res) -> str:
    L = []
    a = L.append
    s = res.summary()
    a("=" * 62)
    if res.period:
        a(f"對帳區間：{res.period[0]} ~ {res.period[1]}")
    a("=" * 62)
    a(f"帳單 {s['帳單筆數']} 筆 / NT${s['帳單總額']:,}")
    a(f"  非匯款（現金/後付/管確認）{s['非匯款筆數']} 筆 / NT${s['非匯款金額']:,}")
    a(f"  應匯款 NT${s['應匯款金額']:,}　已收 NT${s['已收金額']:,}　"
      f"差額 NT${s['差額']:,}")
    a("")
    a(f"✅ 完全相符 {len(res.matched)} 組 / "
      f"NT${sum(g['bank_amount'] for g in res.matched):,}")

    if res.wrong_last5:
        a("")
        a(f"⚠️  末五碼登錄錯誤（款項已收）{len(res.wrong_last5)} 筆")
        for w in res.wrong_last5:
            a(f"   {w['customer']}　NT${w['amount']:,}　"
              f"帳單 {w['bill_last5']} → 實際 {w['bank_last5']}　"
              f"{w['bank_date']}　{w['payer_name']}　［{w['reason']}］")

    if res.short_paid:
        a("")
        a(f"🔻 短收 {len(res.short_paid)} 組")
        for g in res.short_paid:
            a(f"   末五碼 {g['last5']}　{'、'.join(g['customers'])}　"
              f"應收 NT${g['bill_amount']:,} / 實收 NT${g['bank_amount']:,}　"
              f"短 NT${-g['diff']:,}")

    if res.over_paid:
        a("")
        a(f"🔺 溢收 {len(res.over_paid)} 組")
        for g in res.over_paid:
            a(f"   末五碼 {g['last5']}　{'、'.join(g['customers'])}　"
              f"應收 NT${g['bill_amount']:,} / 實收 NT${g['bank_amount']:,}　"
              f"多 NT${g['diff']:,}")

    if res.missing:
        a("")
        a(f"❌ 查無入帳 {len(res.missing)} 筆 / "
          f"NT${sum(m['amount'] for m in res.missing):,}")
        for m in res.missing:
            a(f"   {m['customer']}　NT${m['amount']:,}　末五碼 {m['last5']}　"
              f"出貨 {m['ship_date']}　銷帳註記 {m['paid_at']}")

    if res.unmatched_bank:
        a("")
        a(f"📥 銀行有、帳單無 {len(res.unmatched_bank)} 筆 / "
          f"NT${sum(u['amount'] for u in res.unmatched_bank):,}")
        for u in res.unmatched_bank:
            a(f"   {u['date']}　NT${u['amount']:>8,}　{u['category']}　"
              f"末五碼 {u['last5'] or '-'}　{u['payer_name']}{u['memo']}")

    if res.parse_errors:
        a("")
        a(f"🐞 無法解析 {len(res.parse_errors)} 行")
        for e in res.parse_errors[:10]:
            a(f"   第 {e['lineno']} 行：{e['line']}")
    a("=" * 62)
    return "\n".join(L)


def to_excel(res) -> bytes:
    """需要 openpyxl。回傳 xlsx bytes。"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    head = Font(bold=True, color="FFFFFF")
    fill = PatternFill("solid", fgColor="374151")

    def sheet(title, cols, rows):
        ws = wb.create_sheet(title)
        ws.append(cols)
        for c in ws[1]:
            c.font, c.fill = head, fill
        for r in rows:
            ws.append(r)
        for i, col in enumerate(cols, 1):
            ws.column_dimensions[ws.cell(1, i).column_letter].width = max(12, len(col) * 2.2)
        ws.freeze_panes = "A2"
        return ws

    wb.remove(wb.active)
    sheet("摘要", ["項目", "數值"], list(res.summary().items()))

    sheet("末五碼錯誤", ["客戶", "金額", "帳單末五碼", "實際末五碼", "入帳日", "匯款人", "判定"],
          [[w["customer"], w["amount"], w["bill_last5"], w["bank_last5"],
            w["bank_date"], w["payer_name"], w["reason"]] for w in res.wrong_last5])

    sheet("短收溢收", ["末五碼", "客戶", "應收", "實收", "差額"],
          [[g["last5"], "、".join(g["customers"]), g["bill_amount"],
            g["bank_amount"], g["diff"]] for g in res.short_paid + res.over_paid])

    sheet("查無入帳", ["客戶", "金額", "末五碼", "出貨日", "銷帳註記日"],
          [[m["customer"], m["amount"], m["last5"], m["ship_date"], m["paid_at"]]
           for m in res.missing])

    sheet("銀行有帳單無", ["日期", "金額", "分類", "末五碼", "匯款人", "摘要", "交易代碼"],
          [[u["date"], u["amount"], u["category"], u["last5"],
            u["payer_name"], u["memo"], u["tx_code"]] for u in res.unmatched_bank])

    sheet("相符明細", ["末五碼", "客戶", "金額", "帳單筆數", "入帳筆數"],
          [[g["last5"], "、".join(g["customers"]), g["bank_amount"],
            len(g["bills"]), len(g["banks"])] for g in res.matched])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
