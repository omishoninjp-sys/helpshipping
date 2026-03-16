import os
import asyncio
import re
import json
import queue
import threading
import io
from datetime import datetime, timedelta
from pathlib import Path

import requests as http_requests
from flask import Flask, request, jsonify, send_file, Response, render_template_string
from playwright.async_api import async_playwright
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

app = Flask(__name__)

# ─── 全域 job 管理 ─────────────────────────────────────────────────────────────
jobs = {}   # job_id -> { queue, stop_event, result }


# ─── 爬蟲核心 ──────────────────────────────────────────────────────────────────

def order_in_range(order, date_from, date_to):
    if not date_from and not date_to:
        return True
    date_str = ""
    for key in order:
        if any(k in key.lower() for k in ["date", "日期", "time", "建立", "created"]):
            date_str = order[key]
            break
    if not date_str:
        return True
    for fmt in ["%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%Y/%m/%d"]:
        try:
            order_date = datetime.strptime(date_str[:19], fmt[:len(date_str[:19])])
            if date_from and order_date < datetime.strptime(date_from, "%Y-%m-%d"):
                return False
            if date_to and order_date > datetime.strptime(date_to, "%Y-%m-%d") + timedelta(days=1):
                return False
            return True
        except ValueError:
            continue
    return True


async def parse_orders_on_page(page):
    await page.wait_for_load_state("networkidle")
    orders = []
    rows = await page.query_selector_all("table tbody tr")
    if rows:
        headers = []
        for th in await page.query_selector_all("table thead th"):
            headers.append((await th.inner_text()).strip())
        for row in rows:
            cells = await row.query_selector_all("td")
            order = {}
            img_urls = []
            for i, cell in enumerate(cells):
                key = headers[i] if i < len(headers) else f"col_{i}"
                img = await cell.query_selector("img")
                if img:
                    src = await img.get_attribute("src") or ""
                    if src:
                        order[f"{key}_img_url"] = src
                        img_urls.append(src)
                order[key] = (await cell.inner_text()).strip()
            if img_urls:
                order["_img_urls"] = img_urls
            if order:
                orders.append(order)
        return orders
    return orders


async def run_scraper_async(cfg, date_from, date_to, log_q, stop_event):
    base_url  = cfg["base_url"].rstrip("/")
    login_url = base_url + "/login"
    order_url = cfg["order_url"] or base_url + "/order"
    username  = cfg["username"]
    password  = cfg["password"]
    embed_img = cfg.get("embed_images", True)

    def log(msg):
        log_q.put({"type": "log", "msg": msg})

    all_orders = []
    try:
        async with async_playwright() as pw:
            browser = await pw.chromium.launch(headless=True)
            context = await browser.new_context(
                user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
                viewport={"width": 1440, "height": 900},
            )
            page = await context.new_page()

            log("🔐 登入中...")
            await page.goto(login_url, wait_until="networkidle")
            await page.fill('input[name="account"]', username)
            await page.fill('input[name="password"]', password)
            await page.click('button[type="submit"]')
            await page.wait_for_load_state("networkidle")
            log("✅ 登入成功！")

            url = order_url
            params = []
            if date_from: params.append(f"start_date={date_from}")
            if date_to:   params.append(f"end_date={date_to}")
            if params:    url += "?" + "&".join(params)
            log(f"🌐 前往：{url}")
            await page.goto(url, wait_until="networkidle")

            page_num, zero_count = 1, 0
            while not stop_event.is_set():
                log(f"📄 爬取第 {page_num} 頁...")
                try:
                    orders = await parse_orders_on_page(page)
                except Exception as e:
                    log(f"⚠️ 第 {page_num} 頁失敗：{e}")
                    break

                filtered = [o for o in orders if order_in_range(o, date_from, date_to)]
                log(f"   取得 {len(orders)} 筆，符合 {len(filtered)} 筆")
                all_orders.extend(filtered)

                if len(filtered) == 0:
                    zero_count += 1
                    if zero_count >= 2:
                        log("🛑 連續兩頁無符合資料，自動停止")
                        break
                else:
                    zero_count = 0

                try:
                    next_btn = await page.query_selector(
                        "a[rel='next'], .pagination .next:not(.disabled), "
                        "li.next:not(.disabled) a, [aria-label='Next page']:not([disabled])"
                    )
                    if not next_btn:
                        log("✅ 已到最後一頁")
                        break
                    await next_btn.click()
                    await page.wait_for_load_state("networkidle", timeout=15000)
                    page_num += 1
                except Exception:
                    log("✅ 分頁結束")
                    break

            await page.close()
            await context.close()
            await browser.close()

    except Exception as e:
        log(f"❌ 爬蟲錯誤：{e}")
        log_q.put({"type": "error", "msg": str(e)})
        return None

    if not all_orders:
        log("⚠️ 無訂單資料")
        log_q.put({"type": "done", "count": 0, "file": None})
        return None

    # 產生 Excel
    log(f"📊 產生 Excel，共 {len(all_orders)} 筆...")
    excel_buf = build_excel(all_orders, base_url, order_url, embed_img, log)
    log(f"✅ 完成！共 {len(all_orders)} 筆訂單")
    log_q.put({"type": "done", "count": len(all_orders), "file": excel_buf})
    return excel_buf


def build_excel(orders, base_url, order_url, embed_images, log):
    wb = Workbook()
    ws = wb.active
    ws.title = "訂單列表"

    all_keys = []
    for o in orders:
        for k in o:
            if k not in all_keys and not k.startswith("_"):
                all_keys.append(k)

    hfill = PatternFill("solid", fgColor="1F4E79")
    hfont = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    thin = Border(left=Side(style="thin"), right=Side(style="thin"),
                  top=Side(style="thin"), bottom=Side(style="thin"))

    for col_idx, key in enumerate(all_keys, 1):
        cell = ws.cell(row=1, column=col_idx, value=key)
        cell.fill = hfill
        cell.font = hfont
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin
    ws.row_dimensions[1].height = 30

    for row_idx, order in enumerate(orders, 2):
        ws.row_dimensions[row_idx].height = 60
        for col_idx, key in enumerate(all_keys, 1):
            img_url_key = f"{key}_img_url"
            img_url = order.get(img_url_key, "") or ""
            if not img_url and "_img_urls" in order and key.lower() in ["圖片", "image", "img", "photo"]:
                img_url = order["_img_urls"][0] if order["_img_urls"] else ""

            cell = ws.cell(row=row_idx, column=col_idx)
            if img_url:
                if not img_url.startswith("http"):
                    img_url = base_url + img_url
                if embed_images:
                    cell.value = f'=IMAGE("{img_url}")'
                else:
                    cell.value = img_url
                    cell.hyperlink = img_url
                    cell.font = Font(name="Arial", size=9, color="0563C1", underline="single")
                cell.alignment = Alignment(horizontal="center", vertical="center")
                ws.column_dimensions[get_column_letter(col_idx)].width = 18
            else:
                cell.value = order.get(key, "")
                cell.alignment = Alignment(vertical="center", wrap_text=True)
                cell.font = Font(name="Arial", size=10)
            cell.border = thin
            if row_idx % 2 == 0 and not img_url:
                cell.fill = PatternFill("solid", fgColor="EBF3FB")

    for col_idx, key in enumerate(all_keys, 1):
        img_url_key = f"{key}_img_url"
        if any(img_url_key in o for o in orders) or key.lower() in ["圖片", "image", "img", "photo"]:
            continue
        max_len = max((len(str(order.get(key, ""))) for order in orders), default=10)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 4, len(key) + 4), 40)

    ws.freeze_panes = "A2"

    ws_sum = wb.create_sheet("摘要")
    for r, (k, v) in enumerate([
        ("爬取時間", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("訂單總數", len(orders)),
        ("資料來源", order_url)
    ], 1):
        ws_sum.cell(row=r, column=1, value=k).font = Font(bold=True, name="Arial")
        ws_sum.cell(row=r, column=2, value=v)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def run_job(job_id, cfg, date_from, date_to):
    job = jobs[job_id]
    loop = asyncio.new_event_loop()
    asyncio.set_event_loop(loop)
    try:
        loop.run_until_complete(
            run_scraper_async(cfg, date_from, date_to, job["queue"], job["stop_event"])
        )
    finally:
        loop.close()


# ─── 前端 HTML ─────────────────────────────────────────────────────────────────

HTML = '''<!DOCTYPE html>
<html lang="zh-TW">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>ERP 訂單爬蟲</title>
<style>
  *{box-sizing:border-box;margin:0;padding:0}
  body{font-family:"Segoe UI",Arial,sans-serif;background:#f0f4f8;min-height:100vh}
  header{background:#1F4E79;color:#fff;padding:18px 32px;display:flex;align-items:center;gap:12px;box-shadow:0 2px 8px rgba(0,0,0,.2)}
  header h1{font-size:1.4rem;font-weight:700}
  header span{font-size:.9rem;color:#a8d0ea;margin-left:auto}
  .container{max-width:780px;margin:32px auto;padding:0 16px}
  .card{background:#fff;border-radius:12px;padding:24px;margin-bottom:20px;box-shadow:0 2px 8px rgba(0,0,0,.08)}
  .card h2{font-size:1rem;font-weight:700;color:#1F4E79;margin-bottom:16px;padding-bottom:8px;border-bottom:2px solid #D6E4F0}
  .grid{display:grid;grid-template-columns:1fr 1fr;gap:12px}
  label{display:block;font-size:.85rem;font-weight:600;color:#555;margin-bottom:4px}
  input{width:100%;padding:9px 12px;border:1.5px solid #d0dce8;border-radius:8px;font-size:.95rem;outline:none;transition:border .2s}
  input:focus{border-color:#2E75B6}
  .full{grid-column:1/-1}
  .pass-wrap{position:relative}
  .pass-wrap input{padding-right:80px}
  .toggle-pass{position:absolute;right:10px;top:50%;transform:translateY(-50%);font-size:.8rem;color:#2E75B6;cursor:pointer;user-select:none}
  .quick-btns{display:flex;gap:8px;flex-wrap:wrap;margin-top:8px}
  .quick-btn{padding:5px 14px;border:1.5px solid #2E75B6;border-radius:20px;background:#fff;color:#2E75B6;font-size:.82rem;cursor:pointer;transition:all .15s}
  .quick-btn:hover{background:#2E75B6;color:#fff}
  .opt-row{display:flex;align-items:center;gap:8px;margin-top:12px}
  .opt-row input[type=checkbox]{width:16px;height:16px;accent-color:#1F4E79}
  .opt-row label{margin:0;font-size:.88rem;color:#444;font-weight:400}
  .btn-row{display:flex;gap:12px;margin-top:4px}
  .btn{padding:12px 32px;border:none;border-radius:8px;font-size:1rem;font-weight:700;cursor:pointer;transition:all .15s}
  .btn-start{background:#1F4E79;color:#fff}
  .btn-start:hover{background:#163a5c}
  .btn-start:disabled{background:#b0c4d8;cursor:not-allowed}
  .btn-stop{background:#c0392b;color:#fff}
  .btn-stop:hover{background:#922b21}
  .btn-stop:disabled{background:#e0b0aa;cursor:not-allowed}
  .progress{height:6px;background:#D6E4F0;border-radius:4px;overflow:hidden;margin-bottom:16px}
  .progress-bar{height:100%;width:0%;background:#2E75B6;border-radius:4px;transition:width .3s;animation:none}
  .progress-bar.running{width:100%;animation:indeterminate 1.5s infinite linear;background:linear-gradient(90deg,#D6E4F0 0%,#2E75B6 50%,#D6E4F0 100%);background-size:200% 100%}
  @keyframes indeterminate{0%{background-position:200% 0}100%{background-position:-200% 0}}
  .log-box{background:#0d1117;border-radius:8px;padding:16px;height:220px;overflow-y:auto;font-family:"Courier New",monospace;font-size:.82rem;color:#79c0ff}
  .log-box .log-line{padding:1px 0;border-bottom:none}
  .log-box .err{color:#f85149}
  .badge{display:inline-block;padding:3px 10px;border-radius:12px;font-size:.8rem;font-weight:700}
  .badge-success{background:#d1f2eb;color:#1e8449}
  .badge-error{background:#fadbd8;color:#c0392b}
  .download-area{display:none;margin-top:16px;padding:16px;background:#eaf7ee;border-radius:8px;border:1.5px solid #82e0aa;text-align:center}
  .download-area.show{display:block}
  .download-btn{display:inline-block;padding:10px 28px;background:#1e8449;color:#fff;border-radius:8px;text-decoration:none;font-weight:700;font-size:.95rem}
  .download-btn:hover{background:#196f3d}
  @media(max-width:540px){.grid{grid-template-columns:1fr}}
</style>
</head>
<body>
<header>
  <span>📦</span>
  <h1>ERP 訂單爬蟲</h1>
  <span>自動登入 · 爬取訂單 · 匯出 Excel</span>
</header>

<div class="container">

  <!-- 連線設定 -->
  <div class="card">
    <h2>🔗 連線設定</h2>
    <div class="grid">
      <div class="full">
        <label>網站網址</label>
        <input id="base_url" type="text" placeholder="https://erp.example.com" autocomplete="off">
      </div>
      <div class="full">
        <label>訂單頁網址</label>
        <input id="order_url" type="text" placeholder="https://erp.example.com/order" autocomplete="off">
      </div>
      <div>
        <label>帳號</label>
        <input id="username" type="text" placeholder="輸入帳號" autocomplete="off">
      </div>
      <div>
        <label>密碼</label>
        <div class="pass-wrap">
          <input id="password" type="password" placeholder="輸入密碼" autocomplete="off">
          <span class="toggle-pass" onclick="togglePass()">顯示</span>
        </div>
      </div>
    </div>
  </div>

  <!-- 日期範圍 -->
  <div class="card">
    <h2>📅 日期範圍</h2>
    <div class="grid">
      <div>
        <label>開始日期</label>
        <input id="date_from" type="date">
      </div>
      <div>
        <label>結束日期</label>
        <input id="date_to" type="date">
      </div>
    </div>
    <div class="quick-btns">
      <button class="quick-btn" onclick="quickDate(0)">今天</button>
      <button class="quick-btn" onclick="quickDate(7)">近 7 天</button>
      <button class="quick-btn" onclick="quickDate(30)">近 30 天</button>
      <button class="quick-btn" onclick="quickDate(-1)">本月</button>
      <button class="quick-btn" onclick="quickDate(-2)">全部</button>
    </div>
    <div class="opt-row">
      <input type="checkbox" id="embed_images" checked>
      <label for="embed_images">圖片欄使用 =IMAGE() 公式（Excel / Google Sheets 均支援）</label>
    </div>
  </div>

  <!-- 操作 -->
  <div class="card">
    <h2>🚀 執行</h2>
    <div class="btn-row">
      <button class="btn btn-start" id="btn-start" onclick="startJob()">▶ 開始爬取</button>
      <button class="btn btn-stop" id="btn-stop" onclick="stopJob()" disabled>⏹ 停止</button>
    </div>
    <br>
    <div class="progress"><div class="progress-bar" id="progress-bar"></div></div>
    <div class="log-box" id="log-box"><div class="log-line" style="color:#555">等待開始...</div></div>

    <div class="download-area" id="download-area">
      <p style="font-size:1rem;font-weight:700;color:#1e8449;margin-bottom:10px" id="done-msg"></p>
      <a class="download-btn" id="download-btn" href="#">⬇ 下載 Excel</a>
    </div>
  </div>

</div>

<script>
let currentJobId = null;

// 預設日期
const today = new Date();
const fmt = d => d.toISOString().split('T')[0];
document.getElementById('date_from').value = fmt(new Date(today.getFullYear(), today.getMonth(), 1));
document.getElementById('date_to').value = fmt(today);

// 從 localStorage 還原設定
const saved = JSON.parse(localStorage.getItem('erp_cfg') || '{}');
if (saved.base_url)  document.getElementById('base_url').value  = saved.base_url;
if (saved.order_url) document.getElementById('order_url').value = saved.order_url;
if (saved.username)  document.getElementById('username').value  = saved.username;

function togglePass() {
  const inp = document.getElementById('password');
  const btn = document.querySelector('.toggle-pass');
  inp.type = inp.type === 'password' ? 'text' : 'password';
  btn.textContent = inp.type === 'password' ? '顯示' : '隱藏';
}

function quickDate(days) {
  const t = new Date();
  if (days === 0) {
    document.getElementById('date_from').value = fmt(t);
    document.getElementById('date_to').value   = fmt(t);
  } else if (days > 0) {
    const f = new Date(t); f.setDate(f.getDate() - days);
    document.getElementById('date_from').value = fmt(f);
    document.getElementById('date_to').value   = fmt(t);
  } else if (days === -1) {
    document.getElementById('date_from').value = fmt(new Date(t.getFullYear(), t.getMonth(), 1));
    document.getElementById('date_to').value   = fmt(t);
  } else {
    document.getElementById('date_from').value = '';
    document.getElementById('date_to').value   = '';
  }
}

function log(msg, isErr=false) {
  const box = document.getElementById('log-box');
  const line = document.createElement('div');
  line.className = 'log-line' + (isErr ? ' err' : '');
  line.textContent = msg;
  box.appendChild(line);
  box.scrollTop = box.scrollHeight;
}

function clearLog() {
  document.getElementById('log-box').innerHTML = '';
}

async function startJob() {
  const base_url   = document.getElementById('base_url').value.trim();
  const order_url  = document.getElementById('order_url').value.trim();
  const username   = document.getElementById('username').value.trim();
  const password   = document.getElementById('password').value.trim();
  const date_from  = document.getElementById('date_from').value;
  const date_to    = document.getElementById('date_to').value;
  const embed      = document.getElementById('embed_images').checked;

  if (!base_url || !username || !password) {
    alert('請填寫網站網址、帳號和密碼');
    return;
  }

  // 儲存設定
  localStorage.setItem('erp_cfg', JSON.stringify({ base_url, order_url, username }));

  clearLog();
  document.getElementById('download-area').classList.remove('show');
  document.getElementById('btn-start').disabled = true;
  document.getElementById('btn-stop').disabled  = false;
  document.getElementById('progress-bar').classList.add('running');

  log(`🚀 開始  日期：${date_from || '不限'} ～ ${date_to || '不限'}`);

  const res = await fetch('/api/start', {
    method: 'POST',
    headers: {'Content-Type':'application/json'},
    body: JSON.stringify({ base_url, order_url, username, password, date_from, date_to, embed_images: embed })
  });
  const data = await res.json();
  if (!data.job_id) { log('❌ 建立任務失敗', true); return; }

  currentJobId = data.job_id;

  // SSE 監聽日誌
  // 輪詢方式，每 2 秒問一次，不受 proxy 60 秒逾時影響
  function pollStatus() {
    fetch(`/api/status/${currentJobId}`)
      .then(r => r.json())
      .then(data => {
        (data.messages || []).forEach(m => {
          if (m.type === 'log')        log(m.msg);
          else if (m.type === 'error') log('❌ ' + m.msg, true);
        });
        if (data.done) {
          document.getElementById('progress-bar').classList.remove('running');
          document.getElementById('btn-start').disabled = false;
          document.getElementById('btn-stop').disabled  = true;
          if (data.has_file && data.count > 0) {
            document.getElementById('done-msg').textContent = `✅ 完成！共 ${data.count} 筆訂單`;
            document.getElementById('download-btn').href = `/api/download/${currentJobId}`;
            document.getElementById('download-area').classList.add('show');
          } else if (!data.has_file) {
            log('⚠️ 無資料可下載');
          }
        } else {
          setTimeout(pollStatus, 2000);
        }
      })
      .catch(() => setTimeout(pollStatus, 3000));
  }
  setTimeout(pollStatus, 2000);
}

async function stopJob() {
  if (!currentJobId) return;
  await fetch(`/api/stop/${currentJobId}`, { method: 'POST' });
  log('⏹ 停止請求已送出...');
  document.getElementById('btn-stop').disabled = true;
}
</script>
</body>
</html>
'''


# ─── API Routes ────────────────────────────────────────────────────────────────

@app.route("/")
def index():
    return render_template_string(HTML)


@app.route("/api/start", methods=["POST"])
def start():
    data = request.json
    cfg = {
        "base_url":     data.get("base_url", ""),
        "order_url":    data.get("order_url", ""),
        "username":     data.get("username", ""),
        "password":     data.get("password", ""),
        "embed_images": data.get("embed_images", True),
    }
    date_from = data.get("date_from", "")
    date_to   = data.get("date_to", "")

    job_id = datetime.now().strftime("%Y%m%d%H%M%S%f")
    jobs[job_id] = {
        "queue":      queue.Queue(),
        "stop_event": threading.Event(),
        "result":     None,
    }

    threading.Thread(
        target=run_job,
        args=(job_id, cfg, date_from, date_to),
        daemon=True
    ).start()

    return jsonify({"job_id": job_id})


@app.route("/api/status/<job_id>")
def status(job_id):
    job = jobs.get(job_id)
    if not job:
        return jsonify({"error": "job not found"}), 404

    messages = []
    # 把 queue 裡所有訊息一次取出
    while True:
        try:
            msg = job["queue"].get_nowait()
            # 把 file buffer 存起來，不要放進 JSON
            if msg.get("type") == "done" and msg.get("file"):
                jobs[job_id]["result"] = msg.pop("file")
            messages.append(msg)
        except queue.Empty:
            break

    done = any(m.get("type") in ("done", "error") for m in messages)
    count = next((m.get("count", 0) for m in messages if m.get("type") == "done"), 0)

    return jsonify({
        "messages": messages,
        "done": done,
        "count": count,
        "has_file": jobs[job_id].get("result") is not None
    })


@app.route("/api/stop/<job_id>", methods=["POST"])
def stop(job_id):
    job = jobs.get(job_id)
    if job:
        job["stop_event"].set()
    return jsonify({"ok": True})


@app.route("/api/download/<job_id>")
def download(job_id):
    job = jobs.get(job_id)
    if not job or not job.get("result"):
        return "file not ready", 404
    buf = job["result"]
    buf.seek(0)
    fname = f"orders_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=False)
