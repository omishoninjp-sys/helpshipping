#!/usr/bin/env python3
"""
Shopify × JPD 雲倉 串接工具
御用達-光頭哥 專用
"""

from flask import Flask, render_template, request, jsonify, Response
import requests
import json
import os
import sqlite3
from datetime import datetime, timedelta
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

app = Flask(__name__)

# ============ 從環境變數讀取設定 ============
SHOPIFY_STORE = os.environ.get("SHOPIFY_STORE", "")
SHOPIFY_ACCESS_TOKEN = os.environ.get("SHOPIFY_ACCESS_TOKEN", "")
JPD_EMAIL = os.environ.get("JPD_EMAIL", "")
JPD_PASSWORD = os.environ.get("JPD_PASSWORD", "")
JPD_BASE_URL = os.environ.get("JPD_BASE_URL", "https://biz.cloudwh.jp")
JPD_WAREHOUSE_ID = int(os.environ.get("JPD_WAREHOUSE_ID", "1"))
JPD_DELIV_ID = int(os.environ.get("JPD_DELIV_ID", "40"))
DB_PATH = os.environ.get("DB_PATH", "orders.db")
# =============================================


# ============ SQLite 初始化 ============
def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    conn = get_db()
    conn.execute("""
        CREATE TABLE IF NOT EXISTS order_history (
            id              INTEGER PRIMARY KEY AUTOINCREMENT,
            order_id        TEXT,
            logis_num       TEXT,
            customer_order_id TEXT,
            shopify_order_id TEXT,
            shopify_order_name TEXT,
            recipient       TEXT,
            phone           TEXT,
            address         TEXT,
            items_json      TEXT,
            package_ids     TEXT,
            mode            TEXT DEFAULT 'self',
            status          TEXT DEFAULT '待发货',
            memo            TEXT DEFAULT '',
            created_at      TEXT NOT NULL
        )
    """)
    # ── 出檔案記錄表 ──
    conn.execute("""
        CREATE TABLE IF NOT EXISTS export_batches (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            vendor      TEXT NOT NULL,
            order_count INTEGER DEFAULT 0,
            exported_at TEXT NOT NULL
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS export_items (
            id               INTEGER PRIMARY KEY AUTOINCREMENT,
            batch_id         INTEGER NOT NULL,
            shopify_order_id TEXT NOT NULL,
            customer_order_id TEXT DEFAULT '',
            exported_at      TEXT NOT NULL
        )
    """)
    conn.commit()
    conn.close()

init_db()

def save_order_history(data):
    """建立運單成功後存入本地歷史"""
    try:
        conn = get_db()
        conn.execute("""
            INSERT INTO order_history 
            (order_id, logis_num, customer_order_id, shopify_order_id, shopify_order_name,
             recipient, phone, address, items_json, package_ids, mode, status, memo, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            data.get("order_id", ""),
            data.get("logis_num", ""),
            data.get("customer_order_id", ""),
            data.get("shopify_order_id", ""),
            data.get("shopify_order_name", ""),
            data.get("recipient", ""),
            data.get("phone", ""),
            data.get("address", ""),
            json.dumps(data.get("items", []), ensure_ascii=False),
            data.get("package_ids", ""),
            data.get("mode", "self"),
            data.get("status", "待发货"),
            data.get("memo", ""),
            datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ))
        conn.commit()
        conn.close()
        print(f"💾 運單歷史已存檔: {data.get('customer_order_id')} / {data.get('logis_num')}")
    except Exception as e:
        print(f"⚠️ 存檔失敗: {e}")


def shopify_request(endpoint: str, method: str = "GET", data: dict = None) -> dict:
    """Shopify API 請求"""
    url = f"https://{SHOPIFY_STORE}.myshopify.com/admin/api/2026-01/{endpoint}"
    headers = {
        "X-Shopify-Access-Token": SHOPIFY_ACCESS_TOKEN,
        "Content-Type": "application/json"
    }
    try:
        if method == "GET":
            response = requests.get(url, headers=headers, timeout=30, verify=True)
        elif method == "POST":
            response = requests.post(url, headers=headers, json=data, timeout=30, verify=True)
        elif method == "PUT":
            response = requests.put(url, headers=headers, json=data, timeout=30, verify=True)
        return response.json()
    except requests.exceptions.SSLError:
        import urllib3
        urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
        try:
            if method == "GET":
                response = requests.get(url, headers=headers, timeout=30, verify=False)
            elif method == "POST":
                response = requests.post(url, headers=headers, json=data, timeout=30, verify=False)
            elif method == "PUT":
                response = requests.put(url, headers=headers, json=data, timeout=30, verify=False)
            return response.json()
        except Exception as e:
            return {"error": f"SSL 錯誤: {str(e)}"}
    except Exception as e:
        return {"error": str(e)}


def jpd_request(operation: str, data: dict) -> dict:
    """JPD 雲倉 API 請求"""
    url = f"{JPD_BASE_URL}/api/json.php?Service=SDC&Operation={operation}"
    payload = {
        "login_email": JPD_EMAIL,
        "login_password": JPD_PASSWORD,
        "data": data
    }
    print(f"\n{'='*50}")
    print(f"📤 JPD API 請求: {operation}")
    print(f"Data: {json.dumps(data, ensure_ascii=False, indent=2)}")
    try:
        response = requests.post(url, json=payload, timeout=30)
        result = response.json()
        print(f"📥 回應: {json.dumps(result, ensure_ascii=False, indent=2)}")
        return result
    except Exception as e:
        print(f"❌ 錯誤: {e}")
        return {"error": str(e)}


# ── 共用：解析收件人名字 ──
_INVALID_NAMES = {"本人", "本人本人", "本人 本人", "同上", "同收件人", "test", "測試", ".", "-", ""}

def _parse_recipient(order: dict) -> str:
    shipping  = order.get("shipping_address", {}) or {}
    cust      = order.get("customer", {}) or {}
    billing   = order.get("billing_address", {}) or {}

    def is_valid(name):
        return bool(name and name.strip() not in _INVALID_NAMES)

    sc = f"{(shipping.get('last_name') or '').strip()}{(shipping.get('first_name') or '').strip()}".strip()
    cc = f"{(cust.get('last_name') or '').strip()}{(cust.get('first_name') or '').strip()}".strip()
    bc = f"{(billing.get('last_name') or '').strip()}{(billing.get('first_name') or '').strip()}".strip()

    if is_valid(sc): return sc
    if is_valid(cc): return cc
    if is_valid(bc): return bc
    return (shipping.get("name") or "").strip()


def _parse_address(order: dict) -> str:
    shipping = order.get("shipping_address", {}) or {}
    return " ".join(filter(None, [
        shipping.get("province", ""), shipping.get("city", ""),
        shipping.get("address1", ""), shipping.get("address2", "")
    ])).strip()


# ============================================================
# 基本路由
# ============================================================

@app.route("/")
def index():
    return render_template("index.html")


@app.route("/health")
def health():
    db_info = {}
    try:
        conn = get_db()
        count = conn.execute("SELECT COUNT(*) FROM order_history").fetchone()[0]
        latest = conn.execute("SELECT created_at FROM order_history ORDER BY id DESC LIMIT 1").fetchone()
        conn.close()
        db_info = {"db_path": DB_PATH, "order_count": count, "latest_order": latest[0] if latest else None}
    except Exception as e:
        db_info = {"db_path": DB_PATH, "error": str(e)}
    return jsonify({"status": "ok", "timestamp": datetime.now().isoformat(),
                    "shopify_store": SHOPIFY_STORE, "jpd_configured": bool(JPD_EMAIL), "db": db_info})


# ============================================================
# Shopify
# ============================================================

@app.route("/api/shopify/orders")
def get_shopify_orders():
    status = request.args.get("status", "unfulfilled")
    limit  = request.args.get("limit", 250)
    result = shopify_request(f"orders.json?status=any&fulfillment_status={status}&limit={limit}")

    print(f"\n{'='*50}")
    print(f"📦 Shopify API 回應:")
    print(json.dumps(result, ensure_ascii=False, indent=2)[:1000])

    if "orders" not in result:
        error_msg = result.get("error") or result.get("errors") or str(result)
        return jsonify({"success": False, "error": error_msg})

    orders = []
    for order in result["orders"]:
        shipping = order.get("shipping_address", {}) or {}
        customer_name = _parse_recipient(order)
        active_items = [
            {"title": item["title"], "variant_title": item.get("variant_title", ""),
             "quantity": item.get("fulfillable_quantity", item["quantity"]),
             "price": item["price"], "sku": item.get("sku", "")}
            for item in order["line_items"]
            if item.get("fulfillable_quantity", item["quantity"]) > 0
        ]
        if not active_items:
            continue
        orders.append({
            "id": order["id"],
            "order_number": order["order_number"],
            "name": order["name"],
            "created_at": order["created_at"],
            "total_price": order.get("current_total_price", order["total_price"]),
            "currency": order["currency"],
            "fulfillment_status": order["fulfillment_status"] or "unfulfilled",
            "customer_name": customer_name,
            "phone": shipping.get("phone", ""),
            "address": _parse_address(order),
            "line_items": active_items
        })
    return jsonify({"success": True, "orders": orders})


@app.route("/api/shopify/order/<order_id>")
def get_shopify_order(order_id):
    result = shopify_request(f"orders/{order_id}.json")
    if "order" in result:
        return jsonify({"success": True, "order": result["order"]})
    return jsonify({"success": False, "error": result.get("error", "Order not found")})


# ============================================================
# JPD
# ============================================================

@app.route("/api/jpd/packages")
def get_jpd_packages():
    result = jpd_request("TSearchPackages", {
        "stock_date_from": (datetime.now().replace(day=1)).strftime("%Y-%m-%d 00:00:00")
    })
    if "OperationResult" not in result:
        return jsonify({"success": False, "error": "Unknown error"})
    op_result = result["OperationResult"]
    if op_result["Request"]["IsValid"] != "True":
        return jsonify({"success": False, "error": op_result["Request"].get("Errors", {})})
    packages = op_result["Result"].get("Data", [])

    pkg_to_order = {}
    try:
        conn = get_db()
        rows = conn.execute(
            "SELECT customer_order_id, logis_num, recipient, phone, address, package_ids FROM order_history WHERE package_ids IS NOT NULL AND package_ids != ''"
        ).fetchall()
        conn.close()
        for row in rows:
            for pid in str(row["package_ids"] or "").split(","):
                pid = pid.strip()
                if pid:
                    pkg_to_order[pid] = {
                        "customer_order_id": row["customer_order_id"] or "",
                        "logis_num": row["logis_num"] or "",
                        "recipient": row["recipient"] or "",
                        "tel": row["phone"] or "",
                        "addr1": row["address"] or "",
                    }
    except Exception as e:
        print(f"⚠️ 讀取本地歷史失敗: {e}")

    order_map = {}
    try:
        orders_result = jpd_request("TSearchOrders", {})
        if "OperationResult" in orders_result:
            orders_op = orders_result["OperationResult"]
            if orders_op["Request"]["IsValid"] == "True":
                for o in orders_op.get("Result", {}).get("Data", []):
                    oid = str(o.get("order_id", ""))
                    if oid:
                        order_map[oid] = {
                            "recipient": o.get("recipient", ""), "tel": o.get("tel", ""),
                            "addr1": o.get("addr1", ""), "customer_order_id": o.get("customer_order_id", ""),
                            "logis_num": o.get("logis_num", ""),
                        }
    except Exception as e:
        print(f"⚠️ 查詢 JPD 運單失敗: {e}")

    for pkg in packages:
        pkg_id = str(pkg.get("package_id", ""))
        oid    = str(pkg.get("order_id", ""))
        info = pkg_to_order.get(pkg_id)
        if not info and oid and oid != "0":
            info = order_map.get(oid)
        if info:
            pkg["recipient"] = info.get("recipient", "")
            pkg["tel"]       = info.get("tel", "")
            pkg["addr1"]     = info.get("addr1", "")
            pkg["customer_order_id"] = info.get("customer_order_id", "")
            pkg["logis_num"] = info.get("logis_num", "")
        else:
            pkg["recipient"] = pkg["tel"] = pkg["addr1"] = pkg["customer_order_id"] = pkg["logis_num"] = ""

    return jsonify({"success": True, "packages": packages})


@app.route("/api/jpd/orders")
def get_jpd_orders():
    result = jpd_request("TSearchOrders", {})
    if "OperationResult" in result:
        op_result = result["OperationResult"]
        if op_result["Request"]["IsValid"] == "True":
            return jsonify({"success": True, "orders": op_result["Result"].get("Data", [])})
    result = jpd_request("TSearchOrders", {"create_date": datetime.now().strftime("%Y-%m-%d")})
    if "OperationResult" in result:
        op_result = result["OperationResult"]
        if op_result["Request"]["IsValid"] == "True":
            return jsonify({"success": True, "orders": op_result["Result"].get("Data", [])})
    return jsonify({"success": False, "error": "Failed to fetch JPD orders"})


@app.route("/api/jpd/order_history")
def get_order_history():
    conn = get_db()
    rows = conn.execute("SELECT * FROM order_history ORDER BY id DESC").fetchall()
    conn.close()
    orders = []
    for r in rows:
        row = dict(r)
        try:
            row["items"] = json.loads(row.get("items_json") or "[]")
        except:
            row["items"] = []
        orders.append(row)
    return jsonify({"success": True, "orders": orders})


@app.route("/api/jpd/create_order", methods=["POST"])
def create_jpd_order():
    data = request.json
    mode = data.get("mode", "self")
    declare_list = []
    for item in data.get("declare_list", []):
        declare_list.append({
            "product_name": item.get("product_name", "商品"),
            "product_name_local": item.get("product_name_local", item.get("product_name", "商品")),
            "product_num": int(item.get("product_num", 1)),
            "product_price": int(item.get("product_price", 100))
        })
    total_num   = sum(int(i.get("product_num", 1)) for i in data.get("declare_list", []))
    total_price = sum(int(i.get("product_price", 0)) * int(i.get("product_num", 1)) for i in data.get("declare_list", []))
    package_ids = []

    if mode == "warehouse":
        if not data.get("package_ids"):
            return jsonify({"success": False, "error": "倉庫代發模式需要選擇已入庫的包裹"})
        package_ids = data["package_ids"]
    else:
        forecast_data = {"packages": [{
            "local_logis_num": data["customer_order_id"],
            "client_cid": data["customer_order_id"],
            "client_pid": data["customer_order_id"],
            "client_date": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "warehouse_id": JPD_WAREHOUSE_ID,
            "product_name": declare_list[0]["product_name"] if declare_list else "商品",
            "product_num": total_num,
            "product_price": total_price,
            "declare_list": declare_list
        }]}
        forecast_result = jpd_request("TForecastPackage", forecast_data)
        if "OperationResult" in forecast_result:
            op_result = forecast_result["OperationResult"]
            if op_result["Request"]["IsValid"] == "True":
                result_data = op_result.get("Result", {})
                if result_data.get("Result") == "SUCCESS":
                    for pkg in result_data.get("Data", []):
                        if pkg.get("package_id"):
                            package_ids.append(pkg["package_id"])
                else:
                    return jsonify({"success": False, "error": f"預報包裹失敗: {result_data.get('Data', {}).get('msg', '未知錯誤')}"})
            else:
                return jsonify({"success": False, "error": f"預報包裹失敗: {op_result['Request'].get('Errors', {})}"})
        else:
            return jsonify({"success": False, "error": "預報包裹 API 回應異常"})
        if not package_ids:
            return jsonify({"success": False, "error": "預報包裹失敗：未取得 package_id"})

    recipient      = data["recipient"]
    shopify_order_id = data.get("shopify_order_id")
    if shopify_order_id:
        order_detail = shopify_request(f"orders/{shopify_order_id}.json")
        if "order" in order_detail:
            recipient = _parse_recipient(order_detail["order"]) or recipient
            print(f"📝 JPD 收件人: '{recipient}'")

    order_data = {
        "customer_order_id": data["customer_order_id"],
        "deliv_id": JPD_DELIV_ID,
        "recipient": recipient,
        "id_issure": "",
        "area": 3,
        "addr1": data["address"],
        "addr2": "", "addr3": "", "addr4": "",
        "tel": data["phone"],
        "memo": data.get("memo", ""),
        "create_order_pdf": "y",
        "warehouse_id": JPD_WAREHOUSE_ID,
        "create_package": "n",
        "create_sender": "y",
        "packages": [{"package_id": int(pid), "declare_list": declare_list} for pid in package_ids]
    }
    result = jpd_request("TCreateOrder", order_data)

    if "OperationResult" in result:
        op_result = result["OperationResult"]
        if op_result["Request"]["IsValid"] == "True":
            result_data = op_result["Result"]
            if result_data.get("Result") == "SUCCESS":
                jpd_data = result_data.get("Data", {})
                save_order_history({
                    "order_id": jpd_data.get("order_id", ""),
                    "logis_num": jpd_data.get("logis_num", ""),
                    "customer_order_id": data["customer_order_id"],
                    "shopify_order_id": data.get("shopify_order_id", ""),
                    "shopify_order_name": data.get("customer_order_id", ""),
                    "recipient": recipient, "phone": data.get("phone", ""),
                    "address": data.get("address", ""),
                    "items": data.get("declare_list", []),
                    "package_ids": ",".join(str(p) for p in package_ids),
                    "mode": mode, "memo": data.get("memo", ""),
                })
                return jsonify({"success": True, "order_id": jpd_data.get("order_id"),
                                "logis_num": jpd_data.get("logis_num"), "message": "運單創建成功"})
            else:
                return jsonify({"success": False, "error": result_data.get("Data", {}).get("msg", "創建失敗")})
        else:
            errors = op_result["Request"].get("Errors", {})
            error_list = errors.get("Error", [])
            if isinstance(error_list, dict):
                error_list = [error_list]
            is_duplicate = any("已存在" in str(e.get("Message", "")) for e in error_list)
            if is_duplicate:
                search_result = jpd_request("TSearchOrders", {"customer_order_id": data["customer_order_id"]})
                order_info = {}
                if "OperationResult" in search_result:
                    search_op = search_result["OperationResult"]
                    if search_op["Request"]["IsValid"] == "True":
                        search_data = search_op.get("Result", {}).get("Data", [])
                        if search_data:
                            order_info = search_data[0]
                save_order_history({
                    "order_id": order_info.get("order_id", ""),
                    "logis_num": order_info.get("logis_num", ""),
                    "customer_order_id": data["customer_order_id"],
                    "shopify_order_id": data.get("shopify_order_id", ""),
                    "shopify_order_name": data.get("customer_order_id", ""),
                    "recipient": recipient, "phone": data.get("phone", ""),
                    "address": data.get("address", ""),
                    "items": data.get("declare_list", []),
                    "package_ids": ",".join(str(p) for p in package_ids),
                    "mode": mode, "memo": data.get("memo", ""),
                })
                return jsonify({"success": True, "order_id": order_info.get("order_id", ""),
                                "logis_num": order_info.get("logis_num", ""),
                                "message": "此運單已存在，無需重複建立"})
            return jsonify({"success": False, "error": str(errors)})
    return jsonify({"success": False, "error": "API 回應異常"})


@app.route("/api/jpd/confirm_order", methods=["POST"])
def confirm_jpd_order():
    data = request.json
    result = jpd_request("TConfirmOrder", {"customer_order_id": data.get("customer_order_id")})
    if "OperationResult" in result:
        op_result = result["OperationResult"]
        if op_result["Request"]["IsValid"] == "True" and op_result["Result"].get("Result") == "SUCCESS":
            return jsonify({"success": True, "message": "確定發貨成功"})
    return jsonify({"success": False, "error": "確定發貨失敗"})


@app.route("/api/jpd/cancel_order", methods=["POST"])
def cancel_jpd_order():
    data = request.json
    result = jpd_request("TDeleteOrder", {"customer_order_id": data.get("customer_order_id")})
    if "OperationResult" in result:
        op_result = result["OperationResult"]
        if op_result["Request"]["IsValid"] == "True" and op_result["Result"].get("Result") == "SUCCESS":
            return jsonify({"success": True, "message": "訂單取消成功"})
    return jsonify({"success": False, "error": "取消訂單失敗"})


@app.route("/api/shopify/fulfill", methods=["POST"])
def fulfill_shopify_order():
    data = request.json
    order_id       = data.get("shopify_order_id")
    tracking_number = data.get("tracking_number")
    print(f"\n{'='*50}\n📝 回寫 Shopify 訂單: {order_id}\n📦 追蹤號: {tracking_number}")
    fo_result = shopify_request(f"orders/{order_id}/fulfillment_orders.json")
    if "fulfillment_orders" not in fo_result:
        return jsonify({"success": False, "error": "無法取得訂單資訊"})
    for fo in fo_result["fulfillment_orders"]:
        if fo["status"] in ["open", "in_progress"]:
            fulfill_data = {"fulfillment": {
                "line_items_by_fulfillment_order": [{"fulfillment_order_id": fo["id"]}],
                "tracking_info": {
                    "number": tracking_number,
                    "company": "SG 速貴專線",
                    "url": f"https://www.sgxpress.com/query/?logic_num={tracking_number}"
                },
                "notify_customer": True
            }}
            fulfill_result = shopify_request("fulfillments.json", "POST", fulfill_data)
            if "fulfillment" in fulfill_result:
                return jsonify({"success": True, "message": "出貨資訊已回寫 Shopify",
                                "fulfillment_id": fulfill_result["fulfillment"]["id"]})
            else:
                error_msg = fulfill_result.get("errors") or fulfill_result.get("error") or str(fulfill_result)
                return jsonify({"success": False, "error": f"回寫失敗: {error_msg}"})
    return jsonify({"success": False, "error": "找不到可出貨的訂單項目（可能已出貨）"})


# ============================================================
# 📤 出檔案 API
# ============================================================

@app.route("/api/exports/vendors")
def get_export_vendors():
    """廠商清單（固定兩個）"""
    return jsonify({"success": True, "vendors": [
        {"id": "nigel", "display_name": "Nigel"},
        {"id": "jpd",   "display_name": "小客戶大價值 (JpD)"},
    ]})


@app.route("/api/exports/pending")
def get_exports_pending():
    """
    待出檔案訂單：
    1. 讀 Shopify 所有 unfulfilled 訂單
    2. 交叉比對 order_history 補上 package_ids / logis_num
    3. 排除已在 export_items 中的訂單
    """
    try:
        conn = get_db()
        # 已出檔案的 shopify_order_id（字串集合）
        exported_ids = set(
            r[0] for r in conn.execute("SELECT shopify_order_id FROM export_items").fetchall()
        )
        # order_history 對照：customer_order_id → {package_ids, logis_num}
        history_map = {}
        for row in conn.execute(
            "SELECT customer_order_id, package_ids, logis_num FROM order_history WHERE customer_order_id IS NOT NULL AND customer_order_id != ''"
        ).fetchall():
            coid = (row["customer_order_id"] or "").strip()
            if coid and coid not in history_map:
                history_map[coid] = {
                    "package_ids": row["package_ids"] or "",
                    "logis_num":   row["logis_num"]   or "",
                }
        conn.close()
    except Exception as e:
        return jsonify({"success": False, "error": f"DB 錯誤: {e}"})

    # Shopify unfulfilled 訂單
    result = shopify_request("orders.json?status=any&fulfillment_status=unfulfilled&limit=250")
    if "orders" not in result:
        return jsonify({"success": False, "error": "Shopify API 失敗: " + str(result.get("error", result))})

    items = []
    for order in result["orders"]:
        shopify_id = str(order["id"])
        if shopify_id in exported_ids:
            continue  # 已出過檔案，跳過

        order_name = order["name"].lstrip("#")
        hist       = history_map.get(order_name, {})
        pkg_ids    = hist.get("package_ids", "")
        pkg_count  = len([p for p in pkg_ids.split(",") if p.strip()]) if pkg_ids else 0

        active_items = [
            {"title": it["title"], "variant_title": it.get("variant_title", ""),
             "quantity": it.get("fulfillable_quantity", it["quantity"]), "price": it["price"]}
            for it in order["line_items"]
            if it.get("fulfillable_quantity", it["quantity"]) > 0
        ]
        if not active_items:
            continue

        items.append({
            "id":               order["id"],          # Shopify 數字 ID（用於標記）
            "customer_order_id": order_name,
            "recipient":        _parse_recipient(order),
            "phone":            (order.get("shipping_address") or {}).get("phone", ""),
            "address":          _parse_address(order),
            "package_ids":      pkg_ids,
            "package_count":    pkg_count,
            "logis_num":        hist.get("logis_num", ""),
            "created_at":       order["created_at"],
            "line_items":       active_items,
        })

    return jsonify({"success": True, "items": items, "total": len(items)})


@app.route("/api/exports/history")
def get_exports_history():
    """最近 30 筆出檔案歷史"""
    conn = get_db()
    rows = conn.execute(
        "SELECT id, vendor, order_count, exported_at FROM export_batches ORDER BY id DESC LIMIT 30"
    ).fetchall()
    conn.close()
    _vendor_name = {"nigel": "Nigel", "jpd": "小客戶大價值 (JpD)"}
    history = []
    for r in rows:
        h = dict(r)
        h["batch_id"]    = f"B{str(h['id']).zfill(4)}"
        h["vendor_name"] = _vendor_name.get(h["vendor"], h["vendor"])
        history.append(h)
    return jsonify({"success": True, "history": history})


@app.route("/api/exports/generate", methods=["POST"])
def generate_export():
    """
    產生 Excel 並標記已出檔案。
    Body: { vendor: "nigel"|"jpd", ids: [shopify_order_id, ...] }
    """
    data   = request.json
    vendor = data.get("vendor", "nigel")
    ids    = [str(i) for i in data.get("ids", [])]

    if not ids:
        return jsonify({"success": False, "error": "請選擇訂單"}), 400

    # ── 取 Shopify 訂單資料 ──
    sf_result = shopify_request("orders.json?status=any&fulfillment_status=unfulfilled&limit=250")
    if "orders" not in sf_result:
        return jsonify({"success": False, "error": "Shopify API 失敗"}), 500

    # ── order_history 對照 ──
    conn = get_db()
    history_map = {}
    for row in conn.execute(
        "SELECT customer_order_id, package_ids, logis_num FROM order_history WHERE customer_order_id IS NOT NULL AND customer_order_id != ''"
    ).fetchall():
        coid = (row["customer_order_id"] or "").strip()
        if coid and coid not in history_map:
            history_map[coid] = {"package_ids": row["package_ids"] or "", "logis_num": row["logis_num"] or ""}

    # ── 建立要匯出的訂單清單 ──
    selected = []
    for order in sf_result["orders"]:
        if str(order["id"]) not in ids:
            continue
        order_name = order["name"].lstrip("#")
        hist       = history_map.get(order_name, {})
        active_items = [
            {"title": it["title"], "variant_title": it.get("variant_title", ""),
             "quantity": it.get("fulfillable_quantity", it["quantity"]), "price": it["price"]}
            for it in order["line_items"]
            if it.get("fulfillable_quantity", it["quantity"]) > 0
        ]
        dt   = datetime.fromisoformat(order["created_at"][:10])
        mmdd = dt.strftime("%m%d")
        selected.append({
            "shopify_id":       str(order["id"]),
            "customer_order_id": order_name,
            "recipient":        _parse_recipient(order),
            "phone":            (order.get("shipping_address") or {}).get("phone", ""),
            "address":          _parse_address(order),
            "package_ids":      hist.get("package_ids", ""),
            "logis_num":        hist.get("logis_num", ""),
            "created_at":       order["created_at"][:10],
            "preview_code":     f"{order_name}-{mmdd}",
            "line_items":       active_items,
        })

    if not selected:
        conn.close()
        return jsonify({"success": False, "error": "找不到選取的訂單（可能已被移除或已出貨）"}), 404

    # ── 產生 Excel ──
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "出貨清單"

    hdr_font  = Font(color="FFFFFF", bold=True, size=11)
    thin_side = Side(style='thin', color='555555')
    thin_bdr  = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    center_al = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_al   = Alignment(horizontal='left',   vertical='center', wrap_text=True)

    if vendor == "nigel":
        hdr_fill = PatternFill("solid", fgColor="2C3E50")
        headers  = ["#", "預覽編號", "客戶訂單號", "收件人", "電話", "地址",
                    "商品名稱", "數量", "金額(JPY)", "包裹ID", "運單號", "建立日期", "備註"]
        ws.append(headers)
        for i, cell in enumerate(ws[1]):
            cell.fill      = hdr_fill
            cell.font      = hdr_font
            cell.alignment = center_al
            cell.border    = thin_bdr
        ws.row_dimensions[1].height = 22

        for idx, o in enumerate(selected, 1):
            items_str   = " / ".join(
                f"{it['title']}{' ('+it['variant_title']+')' if it['variant_title'] else ''} ×{it['quantity']}"
                for it in o["line_items"]
            )
            total_price = sum(float(it["price"]) * int(it["quantity"]) for it in o["line_items"])
            total_qty   = sum(int(it["quantity"]) for it in o["line_items"])
            ws.append([idx, o["preview_code"], o["customer_order_id"], o["recipient"],
                       o["phone"], o["address"], items_str, total_qty, round(total_price),
                       o["package_ids"], o["logis_num"], o["created_at"], ""])

        col_widths = [4, 22, 20, 12, 15, 42, 50, 6, 10, 16, 18, 12, 10]

    else:  # jpd / 小客戶大價值：一行一品項
        hdr_fill = PatternFill("solid", fgColor="1A5276")
        headers  = ["預覽編號", "客戶訂單號", "收件人", "電話", "地址",
                    "商品名稱", "規格", "數量", "單價(JPY)", "小計(JPY)", "折合台幣",
                    "包裹ID", "運單號", "建立日期", "備註"]
        ws.append(headers)
        for cell in ws[1]:
            cell.fill      = hdr_fill
            cell.font      = hdr_font
            cell.alignment = center_al
            cell.border    = thin_bdr
        ws.row_dimensions[1].height = 22

        for o in selected:
            for it in o["line_items"]:
                qty      = int(it["quantity"])
                price    = round(float(it["price"]))
                subtotal = qty * price
                twd      = round(subtotal * 0.20)
                ws.append([o["preview_code"], o["customer_order_id"], o["recipient"],
                           o["phone"], o["address"], it["title"],
                           it.get("variant_title", ""), qty, price, subtotal, twd,
                           o["package_ids"], o["logis_num"], o["created_at"], ""])

        col_widths = [22, 20, 12, 15, 42, 40, 15, 6, 10, 10, 10, 16, 18, 12, 10]

    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border    = thin_bdr
            cell.alignment = left_al

    # ── 存匯出記錄到 DB ──
    now_str  = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    batch_id = conn.execute(
        "INSERT INTO export_batches (vendor, order_count, exported_at) VALUES (?, ?, ?)",
        (vendor, len(selected), now_str)
    ).lastrowid
    for o in selected:
        conn.execute(
            "INSERT INTO export_items (batch_id, shopify_order_id, customer_order_id, exported_at) VALUES (?, ?, ?, ?)",
            (batch_id, o["shopify_id"], o["customer_order_id"], now_str)
        )
    conn.commit()
    conn.close()

    # ── 回傳 Excel ──
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    today        = datetime.now().strftime("%Y%m%d")
    vendor_label = "Nigel" if vendor == "nigel" else "JpD"
    filename     = f"{vendor_label}_{today}_B{str(batch_id).zfill(4)}.xlsx"

    return Response(
        buf.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"}
    )


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    print(f"""
    ╔═══════════════════════════════════════════════════════════╗
    ║       Shopify × JPD 雲倉 串接工具                         ║
    ║       御用達-光頭哥 專用                                   ║
    ╚═══════════════════════════════════════════════════════════╝
    🌐 請打開瀏覽器訪問: http://localhost:{port}
    按 Ctrl+C 停止服務
    """)
    app.run(debug=True, host="0.0.0.0", port=port)
