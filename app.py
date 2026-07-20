"""
客人集運預報系統
GOYOUTATI x OMISHONIN 雲倉
"""

from flask import Flask, request, jsonify, render_template, make_response, send_file, session
from datetime import datetime, timedelta
import requests
import json
import os
import sqlite3
import csv
import tw_zip
import io
import time
import re
import secrets
import threading

# 廠商出貨檔案範本（Nigel / JpD…）
import vendors as vendor_templates

# PWA（manifest / service worker）
from pwa import register_pwa
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

app = Flask(__name__)
# Session 設定（環境變數 SESSION_SECRET 沒設就用隨機值，每次重啟會失效但不會暴露 fallback）
app.secret_key = os.environ.get("SESSION_SECRET") or secrets.token_hex(32)
app.config["SESSION_COOKIE_HTTPONLY"] = True
app.config["SESSION_COOKIE_SAMESITE"] = "Lax"
app.config["PERMANENT_SESSION_LIFETIME"] = timedelta(days=7)

# PWA：註冊 /sw.js + /manifest.webmanifest + /admin-manifest.webmanifest 三條路由
register_pwa(app)

# ============ 設定區（從環境變數讀取）============
JPD_BASE_URL = "https://biz.cloudwh.jp"
JPD_EMAIL = os.environ.get("JPD_EMAIL", "omishoninjp@gmail.com")
JPD_PASSWORD = os.environ.get("JPD_PASSWORD", "omi0131")
JPD_WAREHOUSE_ID = int(os.environ.get("JPD_WAREHOUSE_ID", "1"))
JPD_DELIV_ID = int(os.environ.get("JPD_DELIV_ID", "40"))  # 台灣空運線

SHOPIFY_STORE = os.environ.get("SHOPIFY_STORE", "")
SHOPIFY_ACCESS_TOKEN = os.environ.get("SHOPIFY_ACCESS_TOKEN", "")

# 預設運費（台幣/kg），0 表示未設定
DEFAULT_SHIPPING_RATE = int(os.environ.get("DEFAULT_SHIPPING_RATE", "0"))

# 台幣 → 日圓匯率（可透過環境變數調整）
TWD_TO_JPY_RATE = float(os.environ.get("TWD_TO_JPY_RATE", "5.0"))

DB_PATH = os.environ.get("DB_PATH", "packages.db")
# ================================


# ============ SQLite 初始化 ============

def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    # 等鎖最多 5 秒（預設 0 秒）→ 大幅減少「database is locked」錯誤、
    # 多 worker / LINE Bot / 客戶端同時操作時不互卡
    conn.execute("PRAGMA busy_timeout=5000")
    return conn


# ============ 加值服務預設目錄 ============
# 存進 admin_settings（key='extra_service_catalog'）後即可後台管理；此處僅為首次 seed。
# sel=True 才會出現在客戶端出貨申請的可勾選清單（變動價/特殊計費項目 sel=False，由管理員請款時確認）。
DEFAULT_EXTRA_SERVICES = [
    {"id": "es01", "name": "報關單修改",           "cat": "申報", "desc": "申報相關",                                   "price": 1573, "sel": True},
    {"id": "es02", "name": "出口子單申報",         "cat": "申報", "desc": "申報相關",                                   "price": 515,  "sel": True},
    {"id": "es03", "name": "特大型包裝箱加固",     "cat": "加固", "desc": "厚度 8mm，長寬高和 180cm（操作費＋資材費）", "price": 143,  "sel": True},
    {"id": "es04", "name": "套箱 — 長寬高和 160cm", "cat": "資材", "desc": "厚度 8mm，長寬高和 160cm",                    "price": 129,  "sel": True},
    {"id": "es05", "name": "套箱 — 長寬高和 140cm", "cat": "資材", "desc": "厚度 8mm，長寬高和 140cm",                    "price": 100,  "sel": True},
    {"id": "es06", "name": "特長件加值費",         "cat": "操作", "desc": "最長邊超過 1.5m 的包裹都需要添加此服務",       "price": 86,   "sel": True},
    {"id": "es07", "name": "普通包裝箱加固",       "cat": "加固", "desc": "厚度 8mm，長寬高和 160cm（操作費＋資材費）",   "price": 86,   "sel": True},
    {"id": "es08", "name": "外箱氣泡膜加固",       "cat": "加固", "desc": "依尺寸 NT$86～343/圈（變動）",                 "price": 86,   "sel": False},
    {"id": "es09", "name": "日本國內退換貨服務",   "cat": "操作", "desc": "服務費（1 個包裹手續費），運費以佐川官方為準", "price": 86,   "sel": True},
    {"id": "es10", "name": "精確分箱",             "cat": "操作", "desc": "根據客戶提供清單分箱",                         "price": 86,   "sel": True},
    {"id": "es11", "name": "退運",                 "cat": "操作", "desc": "退運服務",                                     "price": 86,   "sel": True},
    {"id": "es12", "name": "更改地址 — 跨省",       "cat": "操作", "desc": "適用服裝、雜貨管道（不成功不收費）",           "price": 86,   "sel": True},
    {"id": "es13", "name": "套箱 — 長寬高和 120cm", "cat": "資材", "desc": "厚度 8mm，長寬高和 120cm",                    "price": 72,   "sel": True},
    {"id": "es14", "name": "清點拍照",             "cat": "拍照", "desc": "取出所有商品排列拍照，每 10 個商品費用",       "price": 57,   "sel": False},
    {"id": "es15", "name": "隨機分箱",             "cat": "操作", "desc": "根據倉庫經驗隨機分箱",                         "price": 43,   "sel": True},
    {"id": "es16", "name": "套箱 — 長寬高和 80cm",  "cat": "資材", "desc": "厚度 8mm，長寬高和 80cm",                     "price": 43,   "sel": True},
    {"id": "es17", "name": "入庫清點",             "cat": "操作", "desc": "入庫清點商品數量（每箱）",                     "price": 37,   "sel": True},
    {"id": "es18", "name": "套隨機外箱",           "cat": "加固", "desc": "利用倉庫積存的廢舊箱子（規格不定）",           "price": 29,   "sel": True},
    {"id": "es19", "name": "商品氣柱加固",         "cat": "加固", "desc": "氣柱長度 20cm 單價，不足 20cm 以 20cm 計（變動）", "price": 29, "sel": False},
    {"id": "es20", "name": "商品氣泡膜加強",       "cat": "加固", "desc": "依尺寸 NT$29～86/圈（變動）",                  "price": 29,   "sel": False},
    {"id": "es21", "name": "拍照確認",             "cat": "拍照", "desc": "隨機角度，3 張",                               "price": 29,   "sel": True},
    {"id": "es22", "name": "面單拍照",             "cat": "拍照", "desc": "只拍照面單",                                   "price": 29,   "sel": True},
    {"id": "es23", "name": "開箱拍照",             "cat": "拍照", "desc": "外箱 1 張、開箱 1 張，共 2 張（不取出商品）",   "price": 29,   "sel": True},
    {"id": "es24", "name": "套箱 — 長寬高和 60cm",  "cat": "資材", "desc": "厚度 8mm，長寬高和 60cm",                     "price": 20,   "sel": True},
    {"id": "es25", "name": "去鞋盒",               "cat": "操作", "desc": "1 個鞋盒的操作費用",                           "price": 14,   "sel": True},
    {"id": "es26", "name": "合箱",                 "cat": "操作", "desc": "3 個以內免費，超過每個 +NT$14（由合箱費處理）", "price": 14,   "sel": False},
]


def get_extra_service_catalog(conn=None):
    """讀取加值服務目錄（admin_settings.extra_service_catalog）。缺則回預設。"""
    own = False
    if conn is None:
        conn = get_db(); own = True
    try:
        row = conn.execute("SELECT value FROM admin_settings WHERE key='extra_service_catalog'").fetchone()
    finally:
        if own:
            conn.close()
    if not row:
        return list(DEFAULT_EXTRA_SERVICES)
    try:
        data = json.loads(row["value"])
        return data if isinstance(data, list) else list(DEFAULT_EXTRA_SERVICES)
    except (ValueError, TypeError):
        return list(DEFAULT_EXTRA_SERVICES)


# ============ 台灣配送貨況（Google Sheet 同步）============
# 貨運行把台灣端派件資料填在這張試算表；系統定時抓 CSV 匯出、用「客戶編號」比對出貨單。
DEFAULT_TRACKING_SHEET_URL = (
    "https://docs.google.com/spreadsheets/d/"
    "1kLCVI56WuOuYwXQRM2dMKYz6B8f5ifTTU5I9bJ7z998/export?format=csv&gid=1114652436"
)
_sync_lock = threading.Lock()


def _get_setting(key, default=""):
    conn = get_db()
    try:
        row = conn.execute("SELECT value FROM admin_settings WHERE key=?", (key,)).fetchone()
    finally:
        conn.close()
    return row["value"] if row else default


def _set_setting(key, value):
    conn = get_db()
    conn.execute(
        "INSERT INTO admin_settings (key, value) VALUES (?, ?) "
        "ON CONFLICT(key) DO UPDATE SET value=excluded.value",
        (key, value)
    )
    conn.commit()
    conn.close()


def delivery_tracking_url(carrier, tracking):
    """依物流商組查詢網址。"""
    t = (tracking or "").strip()
    c = (carrier or "").strip()
    if not t:
        return ""
    if "新竹" in c or "HCT" in c.upper():
        return f"https://www.aftership.com/zh-hant/track/hct-logistics/{t}"
    # 預設黑貓
    return f"https://www.t-cat.com.tw/Inquire/TraceDetail.aspx?BillID={t}"


def sync_delivery_tracking():
    """抓取貨運行試算表 CSV，解析後 upsert 進 delivery_tracking。回傳寫入筆數。"""
    url = _get_setting("tracking_sheet_url", DEFAULT_TRACKING_SHEET_URL)
    resp = requests.get(url, timeout=20)
    resp.raise_for_status()
    text = resp.content.decode("utf-8-sig", errors="replace")
    reader = csv.reader(io.StringIO(text))
    rows = list(reader)
    if not rows:
        return 0

    # 依標題找欄位（容忍欄位順序變動）；找不到就用固定位置 C=2 / F=5 / G=6
    header = [h.strip() for h in rows[0]]
    def _col(names, fallback):
        for i, h in enumerate(header):
            if any(n in h for n in names):
                return i
        return fallback
    ci_code = _col(["客戶編號", "客編"], 2)
    ci_track = _col(["派件轉單號", "轉單號", "追蹤"], 5)
    ci_carrier = _col(["貨態查詢", "物流", "物流商"], 6)

    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    conn = get_db()
    count = 0
    for r in rows[1:]:
        if len(r) <= max(ci_code, ci_track, ci_carrier):
            continue
        code = (r[ci_code] or "").strip()
        tracking = (r[ci_track] or "").strip()
        carrier = (r[ci_carrier] or "").strip()
        if not code or not tracking:
            continue
        conn.execute(
            "INSERT INTO delivery_tracking (customer_code, carrier, tracking_num, synced_at) "
            "VALUES (?, ?, ?, ?) "
            "ON CONFLICT(customer_code) DO UPDATE SET "
            "carrier=excluded.carrier, tracking_num=excluded.tracking_num, synced_at=excluded.synced_at",
            (code, carrier, tracking, now)
        )
        count += 1
    conn.commit()
    conn.close()
    _set_setting("tracking_last_sync", now)
    return count


def maybe_auto_sync():
    """後台有人活動時，若距上次同步 > 24 小時就背景同步一次（不阻塞請求）。"""
    try:
        last = _get_setting("tracking_last_sync", "")
        if last:
            try:
                if (datetime.now() - datetime.strptime(last[:19], "%Y-%m-%d %H:%M:%S")).total_seconds() < 86400:
                    return
            except ValueError:
                pass
        if not _sync_lock.acquire(blocking=False):
            return
        # 先佔位，避免其他 worker/請求重複觸發
        _set_setting("tracking_last_sync", datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        def _run():
            try:
                n = sync_delivery_tracking()
                print(f"[tracking] 自動同步完成，{n} 筆", flush=True)
            except Exception as e:
                print(f"[tracking] 自動同步失敗: {e}", flush=True)
            finally:
                _sync_lock.release()
        threading.Thread(target=_run, daemon=True).start()
    except Exception as e:
        print(f"[tracking] maybe_auto_sync 例外: {e}", flush=True)


def init_db():
    conn = get_db()
    # ===== 啟用 WAL 模式（一次性設定，會持久化在 DB 檔案內）=====
    # WAL：讀寫不互鎖、併發效能大幅提升（讀者不擋寫者、寫者不擋讀者）
    # synchronous=NORMAL：搭配 WAL 安全，速度比 FULL 快很多
    try:
        conn.execute("PRAGMA journal_mode=WAL")
        conn.execute("PRAGMA synchronous=NORMAL")
        mode_row = conn.execute("PRAGMA journal_mode").fetchone()
        print(f"[DB] ✅ SQLite journal_mode = {mode_row[0]}", flush=True)
    except Exception as e:
        print(f"[DB] ⚠️ 啟用 WAL 失敗（將使用預設模式）: {e}", flush=True)

    conn.execute("""
        CREATE TABLE IF NOT EXISTS packages (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            g_code      TEXT    NOT NULL,
            logis_num   TEXT,
            product_name TEXT   DEFAULT '',
            weight      TEXT    DEFAULT '',
            status      TEXT    DEFAULT '已到貨',
            note        TEXT    DEFAULT '',
            in_date     TEXT,
            created_at  TEXT    NOT NULL
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS shipment_requests (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            g_code      TEXT    NOT NULL,
            customer_name TEXT  DEFAULT '',
            package_ids TEXT    NOT NULL,
            package_summary TEXT DEFAULT '',
            status      TEXT    DEFAULT '待處理',
            note        TEXT    DEFAULT '',
            admin_note  TEXT    DEFAULT '',
            created_at  TEXT    NOT NULL,
            updated_at  TEXT
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS forecasts (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            g_code      TEXT    NOT NULL,
            customer_name TEXT  DEFAULT '',
            items_json  TEXT    NOT NULL,
            status      TEXT    DEFAULT '待處理',
            note        TEXT    DEFAULT '',
            created_at  TEXT    NOT NULL
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS admin_settings (
            key   TEXT PRIMARY KEY,
            value TEXT NOT NULL
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS delivery_tracking (
            customer_code TEXT PRIMARY KEY,
            carrier       TEXT DEFAULT '',
            tracking_num  TEXT DEFAULT '',
            synced_at     TEXT DEFAULT ''
        )
    """)
    # ── 停用會員名單（集運系統層級，不動 Shopify；g_code 為鍵）──
    conn.execute("""
        CREATE TABLE IF NOT EXISTS disabled_members (
            g_code       TEXT PRIMARY KEY,
            reason       TEXT DEFAULT '',
            disabled_at  TEXT DEFAULT ''
        )
    """)
    # ── 代理每週分潤撥款記錄（agent_id + period_key 唯一）──
    conn.execute("""
        CREATE TABLE IF NOT EXISTS agent_payouts (
            id            INTEGER PRIMARY KEY AUTOINCREMENT,
            agent_id      INTEGER NOT NULL,
            period_key    TEXT NOT NULL,
            amount        REAL DEFAULT 0,
            payment_last5 TEXT DEFAULT '',
            paid_at       TEXT DEFAULT '',
            note          TEXT DEFAULT '',
            created_at    TEXT DEFAULT '',
            UNIQUE(agent_id, period_key)
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS addresses (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            g_code      TEXT    NOT NULL,
            label       TEXT    DEFAULT '',
            recipient   TEXT    NOT NULL,
            phone       TEXT    NOT NULL,
            zipcode     TEXT    DEFAULT '',
            address     TEXT    NOT NULL,
            is_default  INTEGER DEFAULT 0,
            created_at  TEXT    NOT NULL
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS announcements (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            title       TEXT    NOT NULL,
            content     TEXT    NOT NULL,
            is_active   INTEGER DEFAULT 1,
            created_at  TEXT    NOT NULL
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS admin_users (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            username    TEXT    UNIQUE NOT NULL,
            password    TEXT    NOT NULL,
            role        TEXT    DEFAULT 'admin',
            created_at  TEXT    NOT NULL
        )
    """)
    # ===== 代理帳號表（Phase 1）=====
    conn.execute("""
        CREATE TABLE IF NOT EXISTS agents (
            id              INTEGER PRIMARY KEY AUTOINCREMENT,
            username        TEXT UNIQUE NOT NULL,
            password        TEXT NOT NULL,
            prefix          TEXT UNIQUE NOT NULL,
            name            TEXT NOT NULL,
            min_rate        REAL DEFAULT 180,
            contact_phone   TEXT DEFAULT '',
            contact_email   TEXT DEFAULT '',
            status          TEXT DEFAULT 'active',
            note            TEXT DEFAULT '',
            created_at      TEXT NOT NULL
        )
    """)
    # ===== 會員表（代理建的客戶，存本地；你自己的客戶仍走 Shopify）=====
    conn.execute("""
        CREATE TABLE IF NOT EXISTS members (
            g_code          TEXT PRIMARY KEY,
            agent_id        INTEGER NOT NULL,
            name            TEXT NOT NULL,
            password        TEXT DEFAULT '',
            phone           TEXT DEFAULT '',
            address         TEXT DEFAULT '',
            line_id         TEXT DEFAULT '',
            email           TEXT DEFAULT '',
            note            TEXT DEFAULT '',
            status          TEXT DEFAULT 'active',
            created_at      TEXT NOT NULL
        )
    """)
    conn.execute("CREATE INDEX IF NOT EXISTS idx_members_agent ON members(agent_id)")
    # 帳單欄位遷移（已存在的表加欄位）
    for col, col_type, default in [
        ("admin_note", "TEXT", "''"),
        ("updated_at", "TEXT", "NULL"),
        ("billed_weight", "REAL", "0"),
        ("rate_per_kg", "REAL", "0"),
        ("shipping_fee", "REAL", "0"),
        ("handling_fee", "REAL", "0"),
        ("total_fee", "REAL", "0"),
        ("payment_last5", "TEXT", "''"),
        ("payment_at", "TEXT", "''"),
        ("tracking_num", "TEXT", "''"),
        ("extra_services", "TEXT", "''"),
        ("ship_recipient", "TEXT", "''"),
        ("ship_phone", "TEXT", "''"),
        ("ship_address", "TEXT", "''"),
        ("consolidation_fee", "REAL", "0"),
        ("letter_fee", "REAL", "0"),
        # 出檔案給廠商（Nigel / JpD）追蹤欄位
        ("exported_at", "TEXT", "''"),
        ("exported_vendor", "TEXT", "''"),
        ("exported_batch_id", "TEXT", "''"),
    ]:
        try:
            conn.execute(f"ALTER TABLE shipment_requests ADD COLUMN {col} {col_type} DEFAULT {default}")
        except:
            pass

    # ===== Phase 2: 加 agent_id 欄位（既有資料預設 0 = 主管理員的）=====
    for table in ["packages", "forecasts", "shipment_requests", "announcements"]:
        try:
            conn.execute(f"ALTER TABLE {table} ADD COLUMN agent_id INTEGER DEFAULT 0")
            print(f"[migrate] 已加 {table}.agent_id 欄位", flush=True)
        except:
            pass
    # 索引加速 agent 過濾
    for table in ["packages", "forecasts", "shipment_requests"]:
        try:
            conn.execute(f"CREATE INDEX IF NOT EXISTS idx_{table}_agent ON {table}(agent_id)")
        except:
            pass

    # ===== Phase 3+: 加 members.shipping_rate 欄位（代理為每個客戶設定獨立費率）=====
    # 預設 0 = 沿用該代理的 min_rate；>0 = 該會員的專屬費率
    try:
        conn.execute("ALTER TABLE members ADD COLUMN shipping_rate REAL DEFAULT 0")
        print("[migrate] 已加 members.shipping_rate 欄位", flush=True)
    except:
        pass

    # ===== 包裹類型欄位：區分「包裹」與「信件」（信件計費一件 +NT$20，見帳單邏輯）=====
    try:
        conn.execute("ALTER TABLE packages ADD COLUMN pkg_type TEXT DEFAULT '包裹'")
        print("[migrate] 已加 packages.pkg_type 欄位", flush=True)
    except:
        pass

    # ===== 出檔案客戶編號（{g_code}-{MMDD}）：存起來供台灣配送貨況比對 =====
    try:
        conn.execute("ALTER TABLE shipment_requests ADD COLUMN export_code TEXT DEFAULT ''")
        print("[migrate] 已加 shipment_requests.export_code 欄位", flush=True)
    except:
        pass

    # ===== 客戶 × 廠商編號對照（出檔案給 Nigel / JpD 等廠商時用） =====
    # 對 Shopify 主帳號客戶 + 代理客戶都通用
    conn.execute("""
        CREATE TABLE IF NOT EXISTS customer_vendor_codes (
            g_code TEXT NOT NULL,
            vendor TEXT NOT NULL,
            code TEXT NOT NULL,
            updated_at TEXT,
            PRIMARY KEY (g_code, vendor)
        )
    """)

    # ===== 代理品牌欄位（用於 referral URL + 登入後客製內容）=====
    for col, col_type, default in [
        ("contact_line", "TEXT", "''"),
        ("insurance_url", "TEXT", "''"),
        ("insurance_label", "TEXT", "''"),
        ("insurance_desc", "TEXT", "''"),
        ("signup_guide", "TEXT", "''"),
        ("promo_text", "TEXT", "''"),
        ("promo_price", "TEXT", "''"),
        ("owner_name", "TEXT", "''"),
        ("owner_address", "TEXT", "''"),
        ("bank_code", "TEXT", "''"),
        ("bank_name", "TEXT", "''"),
        ("bank_branch", "TEXT", "''"),
        ("bank_account", "TEXT", "''"),
        ("bank_account_name", "TEXT", "''"),
    ]:
        try:
            conn.execute(f"ALTER TABLE agents ADD COLUMN {col} {col_type} DEFAULT {default}")
            print(f"[migrate] 已加 agents.{col} 欄位", flush=True)
        except:
            pass

    # ===== 首次 seed 加值服務目錄（之後由後台管理，不覆寫既有值）=====
    try:
        has_cat = conn.execute("SELECT 1 FROM admin_settings WHERE key='extra_service_catalog'").fetchone()
        if not has_cat:
            conn.execute(
                "INSERT INTO admin_settings (key, value) VALUES ('extra_service_catalog', ?)",
                (json.dumps(DEFAULT_EXTRA_SERVICES, ensure_ascii=False),)
            )
            print("[migrate] 已 seed 加值服務目錄（26 項）", flush=True)
    except Exception as e:
        print(f"[migrate] ⚠️ seed 加值服務目錄失敗: {e}", flush=True)

    conn.commit()
    conn.close()


init_db()

# ============ 工具函數 ============

def normalize_phone(phone_raw):
    phone = phone_raw.replace(" ", "").replace("-", "")
    if phone.startswith("+886"):
        phone = "0" + phone[4:]
    elif phone.startswith("+81"):
        phone = "0" + phone[3:]
    return phone


def twd_to_jpy(twd_rate):
    """台幣運費 → 日圓運費（四捨五入至整數）"""
    return round(twd_rate * TWD_TO_JPY_RATE)


def jpd_request(operation, data):
    url = f"{JPD_BASE_URL}/api/json.php?Service=SDC&Operation={operation}"
    payload = {
        "login_email": JPD_EMAIL,
        "login_password": JPD_PASSWORD,
        "data": data
    }
    print(f"\n{'='*50}")
    print(f"📤 JPD API 請求: {operation}")
    try:
        response = requests.post(url, json=payload, timeout=30)
        result = response.json()
        return result
    except Exception as e:
        print(f"❌ 錯誤: {e}")
        return {"error": str(e)}


def shopify_graphql(query, variables=None):
    graphql_url = f"https://{SHOPIFY_STORE}/admin/api/2026-01/graphql.json"
    headers = {
        "X-Shopify-Access-Token": SHOPIFY_ACCESS_TOKEN,
        "Content-Type": "application/json"
    }
    payload = {"query": query}
    if variables:
        payload["variables"] = variables
    try:
        response = requests.post(graphql_url, headers=headers, json=payload, timeout=15)
        return response.json()
    except Exception as e:
        print(f"❌ GraphQL 錯誤: {e}")
        return {"error": str(e)}


def shopify_request(endpoint, method="GET", data=None):
    url = f"https://{SHOPIFY_STORE}/admin/api/2026-01/{endpoint}"
    headers = {
        "X-Shopify-Access-Token": SHOPIFY_ACCESS_TOKEN,
        "Content-Type": "application/json"
    }
    try:
        if method == "GET":
            response = requests.get(url, headers=headers, timeout=30)
        elif method == "POST":
            response = requests.post(url, headers=headers, json=data, timeout=30)
        return response.json()
    except Exception as e:
        return {"error": str(e)}


# ============ Shopify 會員快取 ============
# 持久化到磁碟（容器重啟、Zeabur 重新部署都不用重抓）+ stale-while-revalidate

# 快取檔案位置：跟 DB 放同個目錄（Zeabur Volume 持久化）
_db_dir = os.path.dirname(os.path.abspath(DB_PATH))
SHOPIFY_CACHE_FILE = os.environ.get(
    "SHOPIFY_CACHE_FILE",
    os.path.join(_db_dir or ".", "shopify_cache.json")
)
CACHE_TTL = 600  # 10 分鐘

_customers_cache = {"data": None, "time": 0}
_cache_lock = threading.Lock()
_refresh_thread = None  # 背景更新執行緒（同時間只允許一個）


def _load_cache_from_disk():
    """容器啟動時嘗試從磁碟讀取快取，避免每次重啟都要等 Shopify 慢慢回應。"""
    global _customers_cache
    try:
        if os.path.exists(SHOPIFY_CACHE_FILE):
            with open(SHOPIFY_CACHE_FILE, "r", encoding="utf-8") as f:
                data = json.load(f)
            if data.get("data"):
                _customers_cache = {"data": data["data"], "time": data.get("time", 0)}
                age = int(time.time() - _customers_cache["time"])
                print(f"[Shopify] 📂 從磁碟讀取快取：{len(_customers_cache['data'])} 位會員（{age}秒前）", flush=True)
    except Exception as e:
        print(f"[Shopify] ⚠️ 讀取磁碟快取失敗: {e}", flush=True)


def _save_cache_to_disk():
    """以原子方式寫入磁碟（先寫 .tmp 再 rename，避免多 worker 競爭時寫到一半）"""
    try:
        os.makedirs(os.path.dirname(os.path.abspath(SHOPIFY_CACHE_FILE)) or ".", exist_ok=True)
        tmp = SHOPIFY_CACHE_FILE + ".tmp"
        with open(tmp, "w", encoding="utf-8") as f:
            json.dump(_customers_cache, f, ensure_ascii=False)
        os.replace(tmp, SHOPIFY_CACHE_FILE)
    except Exception as e:
        print(f"[Shopify] ⚠️ 寫入磁碟快取失敗: {e}", flush=True)


def _refresh_shopify_async():
    """背景靜默更新（呼叫者立刻回舊資料、不阻塞使用者）"""
    global _customers_cache, _refresh_thread
    try:
        t0 = time.time()
        print("[Shopify] 🔄 背景重新抓取會員…", flush=True)
        customers = _fetch_customers_from_shopify()
        elapsed = time.time() - t0
        if customers:
            with _cache_lock:
                _customers_cache = {"data": customers, "time": time.time()}
            _save_cache_to_disk()
            print(f"[perf] Shopify 背景更新完成: {len(customers)} 位、{elapsed:.2f}s", flush=True)
        else:
            print(f"[Shopify] ⚠️ 背景抓取回空，保留舊快取（{elapsed:.2f}s）", flush=True)
    except Exception as e:
        print(f"[Shopify] ❌ 背景抓取失敗: {e}", flush=True)
    finally:
        with _cache_lock:
            _refresh_thread = None


def get_all_goyoutati_customers(force_refresh=False):
    """
    取得 Shopify 會員清單。Stale-while-revalidate 行為：
      • force_refresh=True：同步等新資料（admin 按「整理」用）
      • 完全無快取（冷啟動、磁碟也沒有）：同步等
      • 有快取但過期：立刻回舊資料，背景靜默更新
      • 有快取且新鮮：直接回（最快路徑，無 print）
    """
    global _customers_cache, _refresh_thread
    now = time.time()
    has_cache = _customers_cache.get("data") is not None
    age = now - _customers_cache.get("time", 0)
    is_stale = age >= CACHE_TTL

    # 情境 1：強制重抓（admin 按「整理」）→ 同步等新資料
    if force_refresh:
        t0 = time.time()
        try:
            customers = _fetch_customers_from_shopify()
            elapsed = time.time() - t0
            if customers:
                with _cache_lock:
                    _customers_cache = {"data": customers, "time": time.time()}
                _save_cache_to_disk()
                print(f"[perf] Shopify force_refresh: {len(customers)} 位、{elapsed:.2f}s", flush=True)
                return customers
            print(f"[Shopify] ⚠️ force_refresh 回空，回傳舊快取（{elapsed:.2f}s）", flush=True)
            return _customers_cache.get("data") or []
        except Exception as e:
            print(f"[Shopify] ❌ force_refresh 失敗: {e}", flush=True)
            return _customers_cache.get("data") or []

    # 情境 2：完全無快取（容器啟動 + 磁碟也沒有）→ 同步等首次抓取
    if not has_cache:
        t0 = time.time()
        try:
            customers = _fetch_customers_from_shopify()
            elapsed = time.time() - t0
            if customers:
                with _cache_lock:
                    _customers_cache = {"data": customers, "time": time.time()}
                _save_cache_to_disk()
                print(f"[perf] Shopify cold-start: {len(customers)} 位、{elapsed:.2f}s", flush=True)
            return customers or []
        except Exception as e:
            print(f"[Shopify] ❌ cold-start 失敗: {e}", flush=True)
            return []

    # 情境 3：有快取但過期 → 啟動背景更新（lock 確保同時間只一個）
    if is_stale:
        with _cache_lock:
            if _refresh_thread is None or not _refresh_thread.is_alive():
                _refresh_thread = threading.Thread(
                    target=_refresh_shopify_async, daemon=True, name="ShopifyRefresh"
                )
                _refresh_thread.start()

    # 情境 3/4：立刻回現有資料（最多就是舊一點點，等背景更新完下次就新的）
    return _customers_cache.get("data") or []


# 啟動時嘗試從磁碟讀取快取
_load_cache_from_disk()


def _fetch_customers_from_shopify():
    customers = []
    cursor = None
    has_next = True
    page = 0

    while has_next and page < 10:  # 最多 10 頁 = 1000 會員
        page += 1
        after_arg = f', after: "{cursor}"' if cursor else ''
        graphql_query = '{metafieldDefinitions(first:1,ownerType:CUSTOMER,namespace:"custom",key:"goyoutati_id"){edges{node{id metafields(first:100' + after_arg + '){edges{node{value owner{...on Customer{id firstName lastName email phone defaultAddress{phone province city zip address1 address2} createdAt shippingRate:metafield(namespace:"custom",key:"shipping_rate"){value}}}} cursor} pageInfo{hasNextPage}}}}}}'

        page_t0 = time.time()
        result = shopify_graphql(graphql_query)
        page_ms = int((time.time() - page_t0) * 1000)
        has_next = False

        if "data" not in result:
            print(f"[Shopify] page {page} error ({page_ms}ms): {result}", flush=True)
            break

        definitions = result["data"].get("metafieldDefinitions", {}).get("edges", [])
        if not definitions:
            print("[Shopify] No metafieldDefinitions found", flush=True)
            break

        metafields_data = definitions[0]["node"].get("metafields", {})
        edges = metafields_data.get("edges", [])
        page_info = metafields_data.get("pageInfo", {})
        has_next = page_info.get("hasNextPage", False)
        print(f"[Shopify] page {page} ({page_ms}ms): got {len(edges)} metafields, hasNextPage={has_next}", flush=True)

        for mf in edges:
            node = mf["node"]
            cursor = mf.get("cursor")
            g_code = node.get("value", "")
            owner = node.get("owner", {})
            if not g_code or not owner:
                continue
            gid = owner.get("id", "")
            customer_id = gid.split("/")[-1] if "/" in gid else gid
            customer_name = f"{owner.get('lastName', '')}{owner.get('firstName', '')}".strip()
            if not customer_name:
                customer_name = owner.get("email", "")
            default_address = owner.get("defaultAddress") or {}
            phone_raw = default_address.get("phone") or owner.get("phone") or ""
            phone = normalize_phone(phone_raw)
            # 用郵遞區號反查補齊缺的縣市/區（Shopify 拆欄常漏縣市區 → 黑貓無法投遞）
            address, _addr_fixed = tw_zip.compose_full_address(
                default_address.get("province", ""),
                default_address.get("city", ""),
                default_address.get("address1", ""),
                default_address.get("address2", ""),
                default_address.get("zip", ""),
            )
            rate_mf = owner.get("shippingRate")
            # shipping_rate 現在儲存台幣值
            shipping_rate_twd = rate_mf["value"] if rate_mf and rate_mf.get("value") else ""
            customers.append({
                "g_code": g_code,
                "customer_id": customer_id,
                "gid": gid,
                "name": customer_name,
                "email": owner.get("email", ""),
                "address": address,
                "phone": phone,
                "phone_raw": phone_raw,
                "shipping_rate": shipping_rate_twd,  # 台幣
                "created_at": owner.get("createdAt", "")
            })
    return customers


# ============ 路由 ============

@app.route("/admin")
def admin_page():
    return render_template("admin.html")


@app.route("/")
def index():
    return render_template("index.html")


@app.route("/api/config")
def get_config():
    """回傳前端所需設定（匯率等）"""
    return jsonify({
        "twd_to_jpy_rate": TWD_TO_JPY_RATE
    })


def get_admin_password():
    """取得管理員密碼：環境變數優先，否則 DB，最後預設"""
    env_pw = os.environ.get("ADMIN_PASSWORD", "")
    if env_pw:
        return env_pw
    conn = get_db()
    row = conn.execute("SELECT value FROM admin_settings WHERE key='admin_password'").fetchone()
    conn.close()
    if row:
        return row["value"]
    return "admin123"


def _ensure_super_admin():
    """確保至少有一個超級管理員"""
    try:
        conn = get_db()
        count = conn.execute("SELECT COUNT(*) as c FROM admin_users").fetchone()["c"]
        if count == 0:
            now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            pwd = get_admin_password()
            conn.execute(
                "INSERT INTO admin_users (username, password, role, created_at) VALUES (?, ?, 'super', ?)",
                ("admin", pwd, now)
            )
            conn.commit()
            print(f"[Admin] ✅ 已建立超級管理員帳號: admin / {pwd}", flush=True)
        else:
            env_pw = os.environ.get("ADMIN_PASSWORD", "")
            if env_pw:
                conn.execute("UPDATE admin_users SET password=? WHERE role='super'", (env_pw,))
                conn.commit()
            print(f"[Admin] ✅ 已有 {count} 個管理員帳號", flush=True)
        conn.close()
    except Exception as e:
        print(f"[Admin] ❌ 初始化失敗: {e}", flush=True)

_ensure_super_admin()
print("[App] ✅ 啟動完成", flush=True)


@app.route("/api/admin/verify", methods=["POST"])
def admin_verify():
    data = request.json
    username = (data.get("username") or "").strip()
    password = data.get("password", "")
    print(f"[Login] 嘗試登入: username='{username}'", flush=True)

    conn = get_db()
    user = None
    user_type = None

    # 1) 先查管理員（admin_users）— 既有行為不變
    if username:
        user = conn.execute(
            "SELECT * FROM admin_users WHERE username=? AND password=?", (username, password)
        ).fetchone()
    else:
        # 相容舊的純密碼登入
        user = conn.execute(
            "SELECT * FROM admin_users WHERE password=?", (password,)
        ).fetchone()
    if user:
        user_type = "admin"

    # 2) admin 找不到 → 再查代理（agents）
    if not user and username:
        user = conn.execute(
            "SELECT * FROM agents WHERE username=? AND password=? AND status='active'",
            (username, password)
        ).fetchone()
        if user:
            user_type = "agent"
    conn.close()

    if user:
        # 寫入 session
        session.permanent = True
        session["user_type"] = user_type
        session["user_id"] = user["id"]
        session["username"] = user["username"]
        if user_type == "admin":
            session["role"] = user["role"]
            session["agent_id"] = 0  # 0 = 主管理員 / 你的員工，看全部
            print(f"[Login] ✅ admin 登入: {user['username']} ({user['role']})", flush=True)
            return jsonify({
                "success": True,
                "user_type": "admin",
                "username": user["username"],
                "role": user["role"]
            })
        else:
            session["role"] = "agent"
            session["agent_id"] = user["id"]
            session["prefix"] = user["prefix"]
            print(f"[Login] ✅ agent 登入: {user['username']} (prefix={user['prefix']}, id={user['id']})", flush=True)
            return jsonify({
                "success": True,
                "user_type": "agent",
                "username": user["username"],
                "name": user["name"],
                "prefix": user["prefix"]
            })

    print(f"[Login] ❌ 登入失敗", flush=True)
    return jsonify({"success": False, "error": "帳號或密碼錯誤"})


# ===== 身份輔助函式 =====
def current_user():
    """回傳當前登入者資訊（從 session）"""
    if "user_type" not in session:
        return None
    return {
        "user_type": session.get("user_type"),
        "user_id": session.get("user_id"),
        "username": session.get("username"),
        "role": session.get("role"),
        "agent_id": session.get("agent_id", 0),
        "prefix": session.get("prefix", "G"),
    }

def is_super_admin():
    """是否為主管理員（admin_users 表的人，看全部）"""
    return session.get("user_type") == "admin"

def get_current_agent_id():
    """當前代理 id（>0 才是代理，0 = 主管理員看全部）"""
    return int(session.get("agent_id", 0))


def _safe_str(v):
    """安全把 None / 數字 / 字串 都轉為 strip 過的字串。

    主要解決 DB 內有些 phone 欄位被存成 float（912345678.0）的問題。
    無論進來是 float、字串 "912345678.0"、還是已是純字串，都正規化為合理形式：
      • None              → ""
      • 912345678.0       → "912345678"   （整數 float 去掉 .0）
      • "912345678.0"     → "912345678"   （SQLite 存進 TEXT 欄位後變字串）
      • "0912345678"      → "0912345678"  （原樣保留）
      • 3.14              → "3.14"        （非整數 float 保留）
    """
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v)).strip()
    s = str(v).strip()
    # SQLite TEXT 欄位存進 float 後變字串 "912345678.0" → 去掉 .0
    if s.endswith(".0"):
        prefix = s[:-2]
        if prefix.lstrip("-").isdigit():
            return prefix
    return s


def _parse_pkg_ids(raw):
    """容錯解析 package_ids 字串 → [int]（保序、不去重，與舊 list comprehension 行為一致）。

    支援這些髒格式：
      • "5,8,12"        → [5, 8, 12]   （正常）
      • "5.0,8.0,12.0"  → [5, 8, 12]   （migration 把整數 floatify；舊版用 .isdigit() 會整批解析成空）
      • " 5 , 8 "       → [5, 8]       （多餘空白）
      • "5，8、12"       → [5, 8, 12]   （全形逗號／頓號／空白分隔）
      • None / ""       → []
    只接受純整數或「整數.000」格式；"5.7" / "abc" / "-3" / "1e3" 一律忽略，
    避免把壞資料硬轉成錯誤 ID。
    """
    if raw is None:
        return []
    out = []
    for tok in re.split(r"[,，、\s]+", str(raw)):
        tok = tok.strip()
        if not tok:
            continue
        # 純整數，或整數後接全 0 的小數（"5"、"5.0"、"5.00"）
        if re.fullmatch(r"\d+(?:\.0+)?", tok):
            n = int(float(tok))
            if n > 0:
                out.append(n)
    return out


@app.route("/api/me", methods=["GET"])
def api_me():
    """前端查當前身份"""
    u = current_user()
    if not u:
        return jsonify({"logged_in": False})
    return jsonify({"logged_in": True, **u})


@app.route("/api/admin/logout", methods=["POST"])
def admin_logout():
    session.clear()
    return jsonify({"success": True})


@app.route("/api/admin/change_password", methods=["POST"])
def admin_change_password():
    data = request.json
    username = (data.get("username") or "admin").strip()
    current = data.get("current", "")
    new_pwd = data.get("new_password", "").strip()
    confirm = data.get("confirm", "").strip()

    conn = get_db()
    user = conn.execute("SELECT * FROM admin_users WHERE username=?", (username,)).fetchone()
    if not user or user["password"] != current:
        conn.close()
        return jsonify({"success": False, "error": "目前密碼錯誤"})
    if not new_pwd or len(new_pwd) < 4:
        conn.close()
        return jsonify({"success": False, "error": "新密碼至少 4 個字元"})
    if new_pwd != confirm:
        conn.close()
        return jsonify({"success": False, "error": "兩次密碼不一致"})

    conn.execute("UPDATE admin_users SET password=? WHERE username=?", (new_pwd, username))
    conn.commit()
    conn.close()
    return jsonify({"success": True, "message": "密碼已更新"})


# ── 管理員帳號管理 ──

@app.route("/api/admin/users", methods=["GET"])
def admin_list_users():
    conn = get_db()
    rows = conn.execute("SELECT id, username, role, created_at FROM admin_users ORDER BY id").fetchall()
    conn.close()
    return jsonify({"success": True, "users": [dict(r) for r in rows]})


# ===== 代理帳號管理（只有主管理員可操作）=====

@app.route("/api/admin/agents", methods=["GET"])
def admin_list_agents():
    if not is_super_admin():
        return jsonify({"success": False, "error": "權限不足"}), 403
    conn = get_db()
    rows = conn.execute(
        "SELECT id, username, prefix, name, min_rate, contact_phone, contact_email, status, note, created_at,"
        " contact_line, insurance_url, insurance_label, insurance_desc, signup_guide, promo_text, promo_price,"
        " owner_name, owner_address,"
        " bank_code, bank_name, bank_branch, bank_account, bank_account_name"
        " FROM agents ORDER BY id"
    ).fetchall()
    # 順便統計每個代理底下的會員數
    counts = {}
    for r in conn.execute("SELECT agent_id, COUNT(*) as c FROM members GROUP BY agent_id").fetchall():
        counts[r["agent_id"]] = r["c"]
    conn.close()
    # 每個代理的分潤摘要（各週明細 + 未撥款總額），讓你不必登入代理帳號就看得到
    result = []
    for r in rows:
        d = dict(r)
        d["member_count"] = counts.get(r["id"], 0)
        weeks = compute_agent_weekly(r["id"])
        d["weeks"] = weeks
        d["unpaid_total"] = round(sum(w["commission"] for w in weeks if not w["paid"]))
        d["paid_total"] = round(sum(w["commission"] for w in weeks if w["paid"]))
        result.append(d)
    return jsonify({"success": True, "agents": result})


# ===== 代理分潤撥款 =====

@app.route("/api/admin/agents/<int:agent_id>/payout", methods=["POST"])
def admin_agent_payout(agent_id):
    """標記某代理某週已撥款（填匯款後五碼 + 時間）。"""
    if not is_super_admin():
        return jsonify({"success": False, "error": "權限不足"}), 403
    data = request.json or {}
    period_key = (data.get("period_key") or "").strip()
    last5 = (data.get("payment_last5") or "").strip()
    note = (data.get("note") or "").strip()
    if not period_key:
        return jsonify({"success": False, "error": "缺少週次"}), 400
    if not last5:
        return jsonify({"success": False, "error": "請填匯款後五碼"}), 400

    # 金額以系統計算為準（避免前端竄改）
    weeks = compute_agent_weekly(agent_id)
    amount = next((w["commission"] for w in weeks if w["period_key"] == period_key), None)
    if amount is None:
        return jsonify({"success": False, "error": "該週無分潤資料"}), 400

    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    paid_at = (data.get("paid_at") or "").strip() or now
    conn = get_db()
    conn.execute(
        "INSERT INTO agent_payouts (agent_id, period_key, amount, payment_last5, paid_at, note, created_at) "
        "VALUES (?, ?, ?, ?, ?, ?, ?) "
        "ON CONFLICT(agent_id, period_key) DO UPDATE SET "
        "amount=excluded.amount, payment_last5=excluded.payment_last5, paid_at=excluded.paid_at, note=excluded.note",
        (agent_id, period_key, amount, last5, paid_at, note, now)
    )
    conn.commit()
    conn.close()
    return jsonify({"success": True, "amount": amount, "paid_at": paid_at})


@app.route("/api/admin/agents/<int:agent_id>/payout", methods=["DELETE"])
def admin_agent_payout_cancel(agent_id):
    """取消某週撥款標記。"""
    if not is_super_admin():
        return jsonify({"success": False, "error": "權限不足"}), 403
    period_key = (request.args.get("period_key") or "").strip()
    if not period_key:
        return jsonify({"success": False, "error": "缺少週次"}), 400
    conn = get_db()
    conn.execute("DELETE FROM agent_payouts WHERE agent_id=? AND period_key=?", (agent_id, period_key))
    conn.commit()
    conn.close()
    return jsonify({"success": True})


@app.route("/api/agent/payouts", methods=["GET"])
def agent_my_payouts():
    """代理端：看自己的每週分潤與撥款狀態。"""
    aid = get_current_agent_id()
    if aid <= 0:
        return jsonify({"success": False, "error": "僅代理帳號可查看"}), 403
    return jsonify({"success": True, "weeks": compute_agent_weekly(aid)})


# ===== 代理品牌資訊（公開：referral URL 使用）=====
def _branding_dict(agent_row=None):
    """組合品牌資料：有 agent → 該代理；無 agent → 你的預設"""
    if agent_row:
        d = dict(agent_row)
        return {
            "is_agent": True,
            "agent_prefix": d.get("prefix", ""),
            "display_name": d.get("name") or "GOYOUTATI",
            "contact_line": d.get("contact_line") or "",
            "insurance_url": d.get("insurance_url") or "",
            "insurance_label": d.get("insurance_label") or "",
            "insurance_desc": d.get("insurance_desc") or "",
            "signup_guide": d.get("signup_guide") or "",
            "promo_text": d.get("promo_text") or "",
            "promo_price": d.get("promo_price") or "",
        }
    # 預設（你的 GOYOUTATI 品牌）
    return {
        "is_agent": False,
        "agent_prefix": "G",
        "display_name": "GOYOUTATI",
        "contact_line": "",  # 你的官方 LINE 由前端寫死的內容處理
        "insurance_url": "https://goyoutati.com/products/goyoutati-%E5%AE%89%E5%BF%83%E8%B3%BC-%E5%AE%89%E5%BF%83go",
        "insurance_label": "立即加購安心GO",
        "insurance_desc": "貨物保險，最高賠償上限 25 萬日圓",
        "signup_guide": "",  # 預設由前端原本的內容處理
        "promo_text": "",   # 預設徽章由前端寫死（限時特價招生 NT$200）
        "promo_price": "",
    }


@app.route("/api/branding", methods=["GET"])
def api_branding():
    """
    公開 API：依 ?a=PREFIX 回傳對應代理的品牌資訊。
    無 a 或找不到 → 回預設（你的 GOYOUTATI 內容）
    """
    prefix = (request.args.get("a") or "").strip().upper()
    if not prefix:
        return jsonify({"success": True, **_branding_dict(None)})
    try:
        conn = get_db()
        row = conn.execute(
            "SELECT * FROM agents WHERE prefix=? AND status='active'", (prefix,)
        ).fetchone()
        conn.close()
        return jsonify({"success": True, **_branding_dict(row)})
    except Exception as e:
        print(f"[api_branding] {e}", flush=True)
        return jsonify({"success": True, **_branding_dict(None)})


@app.route("/api/agent/my_branding", methods=["GET", "PUT"])
def agent_my_branding():
    """代理自助：查看與編輯自己的品牌設定"""
    aid = get_current_agent_id()
    if aid <= 0:
        return jsonify({"success": False, "error": "僅代理可使用此功能"}), 403
    conn = get_db()
    if request.method == "GET":
        row = conn.execute("SELECT * FROM agents WHERE id=?", (aid,)).fetchone()
        conn.close()
        if not row:
            return jsonify({"success": False, "error": "代理資料異常"}), 500
        d = dict(row)
        return jsonify({
            "success": True,
            "agent": {
                "id": d["id"],
                "name": d.get("name", ""),
                "prefix": d.get("prefix", ""),
                "min_rate": d.get("min_rate", 0),
                "contact_phone": d.get("contact_phone", ""),
                "contact_email": d.get("contact_email", ""),
                "contact_line": d.get("contact_line", ""),
                "insurance_url": d.get("insurance_url", ""),
                "insurance_label": d.get("insurance_label", ""),
                "insurance_desc": d.get("insurance_desc", ""),
                "signup_guide": d.get("signup_guide", ""),
                "promo_text": d.get("promo_text", ""),
                "promo_price": d.get("promo_price", ""),
                "owner_name": d.get("owner_name", ""),
                "owner_address": d.get("owner_address", ""),
                "bank_code": d.get("bank_code", ""),
                "bank_name": d.get("bank_name", ""),
                "bank_branch": d.get("bank_branch", ""),
                "bank_account": d.get("bank_account", ""),
                "bank_account_name": d.get("bank_account_name", ""),
            }
        })
    # PUT：更新自己的品牌欄位（不能改帳號、前綴、密碼、狀態、min_rate）
    data = request.json or {}
    fields, values = [], []
    # 允許代理自己改的欄位：聯絡方式 + 品牌 + 特價徽章 + 負責人資訊 + 銀行帳戶
    for col in ["name", "contact_phone", "contact_email", "contact_line",
                "insurance_url", "insurance_label", "insurance_desc", "signup_guide",
                "promo_text", "promo_price",
                "owner_name", "owner_address",
                "bank_code", "bank_name", "bank_branch", "bank_account", "bank_account_name"]:
        if col in data:
            fields.append(f"{col}=?"); values.append((data[col] or "").strip())
    # min_rate 開放代理自設費率
    if "min_rate" in data:
        try:
            mr = float(data["min_rate"])
            fields.append("min_rate=?"); values.append(mr)
        except (ValueError, TypeError):
            conn.close()
            return jsonify({"success": False, "error": "費率必須為數字"})
    if not fields:
        conn.close()
        return jsonify({"success": False, "error": "沒有可更新欄位"})
    values.append(aid)
    conn.execute(f"UPDATE agents SET {', '.join(fields)} WHERE id=?", values)
    conn.commit()
    conn.close()
    return jsonify({"success": True, "message": "已儲存"})


@app.route("/api/admin/agents", methods=["POST"])
def admin_create_agent():
    if not is_super_admin():
        return jsonify({"success": False, "error": "權限不足"}), 403
    data = request.json or {}
    username = (data.get("username") or "").strip()
    password = (data.get("password") or "").strip()
    prefix = (data.get("prefix") or "").strip().upper()
    name = (data.get("name") or "").strip()
    min_rate = float(data.get("min_rate") or 180)

    if not username or not password or not prefix or not name:
        return jsonify({"success": False, "error": "帳號、密碼、前綴、名稱皆為必填"})
    if not re.fullmatch(r"[A-Z]", prefix):
        return jsonify({"success": False, "error": "前綴必須為單一英文字母（A-Z）"})
    if prefix == "G":
        return jsonify({"success": False, "error": "前綴 G 已保留給主管理員"})
    # min_rate 已開放代理自由設定（不再有 180 下限）

    conn = get_db()
    try:
        conn.execute(
            """INSERT INTO agents (username, password, prefix, name, min_rate, contact_phone, contact_email,
                                   status, note, created_at, contact_line, insurance_url, insurance_label, insurance_desc, signup_guide,
                                   promo_text, promo_price, owner_name, owner_address)
               VALUES (?, ?, ?, ?, ?, ?, ?, 'active', ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (username, password, prefix, name, min_rate,
             data.get("contact_phone", ""), data.get("contact_email", ""),
             data.get("note", ""), datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
             data.get("contact_line", ""), data.get("insurance_url", ""),
             data.get("insurance_label", ""), data.get("insurance_desc", ""),
             data.get("signup_guide", ""),
             data.get("promo_text", ""), data.get("promo_price", ""),
             data.get("owner_name", ""), data.get("owner_address", ""))
        )
        conn.commit()
        new_id = conn.execute("SELECT last_insert_rowid() as id").fetchone()["id"]
    except sqlite3.IntegrityError as e:
        conn.close()
        msg = str(e)
        if "agents.username" in msg:
            return jsonify({"success": False, "error": f"帳號「{username}」已被使用"})
        if "agents.prefix" in msg:
            return jsonify({"success": False, "error": f"前綴「{prefix}」已被使用"})
        return jsonify({"success": False, "error": f"資料庫錯誤：{msg}"})
    conn.close()
    return jsonify({"success": True, "id": new_id, "message": f"代理「{name}」已建立（前綴 {prefix}）"})


@app.route("/api/admin/agents/<int:agent_id>", methods=["PUT"])
def admin_update_agent(agent_id):
    if not is_super_admin():
        return jsonify({"success": False, "error": "權限不足"}), 403
    data = request.json or {}
    conn = get_db()
    existing = conn.execute("SELECT * FROM agents WHERE id=?", (agent_id,)).fetchone()
    if not existing:
        conn.close()
        return jsonify({"success": False, "error": "代理不存在"})

    fields = []
    values = []
    # 可改：name, min_rate, contact_phone, contact_email, status, note, password
    if "name" in data:
        fields.append("name=?"); values.append((data["name"] or "").strip())
    if "min_rate" in data:
        try:
            mr = float(data["min_rate"])
            fields.append("min_rate=?"); values.append(mr)
        except (ValueError, TypeError):
            conn.close()
            return jsonify({"success": False, "error": "費率必須為數字"})
    if "contact_phone" in data:
        fields.append("contact_phone=?"); values.append(data["contact_phone"] or "")
    if "contact_email" in data:
        fields.append("contact_email=?"); values.append(data["contact_email"] or "")
    if "status" in data and data["status"] in ("active", "disabled"):
        fields.append("status=?"); values.append(data["status"])
    if "note" in data:
        fields.append("note=?"); values.append(data["note"] or "")
    if data.get("password"):
        fields.append("password=?"); values.append(data["password"])
    # 品牌欄位（referral URL + 登入後內容客製） + 銀行帳戶（撥款用）
    for col in ["contact_line", "insurance_url", "insurance_label", "insurance_desc", "signup_guide",
                "promo_text", "promo_price", "owner_name", "owner_address",
                "bank_code", "bank_name", "bank_branch", "bank_account", "bank_account_name"]:
        if col in data:
            fields.append(f"{col}=?"); values.append((data[col] or "").strip())
    # 前綴與帳號名建立後不可改（避免關聯混亂）

    if not fields:
        conn.close()
        return jsonify({"success": False, "error": "沒有可更新的欄位"})
    values.append(agent_id)
    conn.execute(f"UPDATE agents SET {', '.join(fields)} WHERE id=?", values)
    conn.commit()
    conn.close()
    return jsonify({"success": True, "message": "已更新"})


@app.route("/api/admin/agents/<int:agent_id>", methods=["DELETE"])
def admin_delete_agent(agent_id):
    if not is_super_admin():
        return jsonify({"success": False, "error": "權限不足"}), 403
    transfer = request.args.get("transfer") == "1"
    conn = get_db()
    # 統計關聯資料
    cm = conn.execute("SELECT COUNT(*) as c FROM members WHERE agent_id=?", (agent_id,)).fetchone()["c"]
    cp = conn.execute("SELECT COUNT(*) as c FROM packages WHERE agent_id=?", (agent_id,)).fetchone()["c"]
    cf = conn.execute("SELECT COUNT(*) as c FROM forecasts WHERE agent_id=?", (agent_id,)).fetchone()["c"]
    cs = conn.execute("SELECT COUNT(*) as c FROM shipment_requests WHERE agent_id=?", (agent_id,)).fetchone()["c"]
    has_data = (cm + cp + cf + cs) > 0
    if has_data and not transfer:
        conn.close()
        return jsonify({
            "success": False,
            "needs_transfer": True,
            "stats": {"members": cm, "packages": cp, "forecasts": cf, "shipment_requests": cs},
            "error": f"此代理底下尚有 {cm} 位會員 / {cp} 個包裹 / {cf} 個預報 / {cs} 個出貨紀錄。請改用「離職移交」將資料轉回主管理員。"
        })
    if has_data and transfer:
        # 全部 agent_id 改為 0 （= 主管理員 / 你）
        conn.execute("UPDATE members SET agent_id=0 WHERE agent_id=?", (agent_id,))
        conn.execute("UPDATE packages SET agent_id=0 WHERE agent_id=?", (agent_id,))
        conn.execute("UPDATE forecasts SET agent_id=0 WHERE agent_id=?", (agent_id,))
        conn.execute("UPDATE shipment_requests SET agent_id=0 WHERE agent_id=?", (agent_id,))
        print(f"[transfer] agent_id={agent_id} 移交 {cm} 會員 / {cp} 包裹 / {cf} 預報 / {cs} 出貨 給主管理員", flush=True)
    conn.execute("DELETE FROM agents WHERE id=?", (agent_id,))
    conn.commit()
    conn.close()
    if has_data and transfer:
        return jsonify({
            "success": True,
            "transferred": {"members": cm, "packages": cp, "forecasts": cf, "shipment_requests": cs},
            "message": f"已離職移交：{cm} 位會員、{cp} 個包裹、{cf} 個預報、{cs} 個出貨紀錄已轉回主管理員。會員編號保留不變。"
        })
    return jsonify({"success": True, "message": "代理已刪除"})


# ===== 統一會員查詢（本地 members 優先、找不到回退 Shopify）=====
def get_agent_id_for_g_code(g_code):
    """
    依 g_code 找出歸屬的 agent_id。
    - 在 members 表（代理的客戶）→ 回那個代理的 id
    - 不在 members 表（你 Shopify 來的客戶）→ 回 0（主管理員）
    """
    if not g_code:
        return 0
    try:
        conn = get_db()
        row = conn.execute("SELECT agent_id FROM members WHERE g_code=?", (g_code,)).fetchone()
        conn.close()
        if row:
            return int(row["agent_id"] or 0)
    except Exception as e:
        print(f"[get_agent_id_for_g_code] 失敗: {e}", flush=True)
    return 0


def check_record_ownership(table, record_id):
    """
    檢查當前使用者是否可存取該筆紀錄。
    回傳 (allowed: bool, record_dict_or_none)
    - 主管理員：永遠可以
    - 代理：只有當 record.agent_id == 自己的 agent_id 才可以
    """
    aid = get_current_agent_id()
    conn = get_db()
    row = conn.execute(f"SELECT * FROM {table} WHERE id=?", (record_id,)).fetchone()
    conn.close()
    if not row:
        return False, None
    if aid == 0 or is_super_admin():
        return True, dict(row)
    return (int(row["agent_id"] or 0) == aid), dict(row)


def get_member_unified(g_code):
    """
    回傳 {g_code, name, agent_id, phone, address, source} 或 None
    - source='local'  → 來自代理建的會員（agent_id > 0）
    - source='shopify' → 來自你的 Shopify（agent_id = 0）
    """
    if not g_code:
        return None
    conn = get_db()
    row = conn.execute("SELECT * FROM members WHERE g_code=?", (g_code,)).fetchone()
    conn.close()
    if row:
        d = dict(row)
        d["source"] = "local"
        return d
    # 回退到 Shopify 快取
    try:
        for c in get_all_goyoutati_customers():
            if (c.get("g_code") or "") == g_code:
                return {
                    "g_code": g_code,
                    "name": c.get("name", ""),
                    "phone": c.get("phone", ""),
                    "address": c.get("address", ""),
                    "agent_id": 0,
                    "source": "shopify"
                }
    except Exception as e:
        print(f"[get_member_unified] Shopify 查詢失敗: {e}", flush=True)
    return None


@app.route("/api/admin/users", methods=["POST"])
def admin_create_user():
    if not is_super_admin():
        return jsonify({"success": False, "error": "權限不足"}), 403
    data = request.json
    username = (data.get("username") or "").strip()
    password = (data.get("password") or "").strip()
    role = data.get("role", "admin")
    if not username or not password:
        return jsonify({"success": False, "error": "帳號和密碼為必填"})
    if len(password) < 4:
        return jsonify({"success": False, "error": "密碼至少 4 個字元"})
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    conn = get_db()
    try:
        conn.execute("INSERT INTO admin_users (username, password, role, created_at) VALUES (?, ?, ?, ?)",
                     (username, password, role, now))
        conn.commit()
        conn.close()
        return jsonify({"success": True})
    except:
        conn.close()
        return jsonify({"success": False, "error": "帳號已存在"})

@app.route("/api/admin/users/<int:user_id>", methods=["DELETE"])
def admin_delete_user(user_id):
    if not is_super_admin():
        return jsonify({"success": False, "error": "權限不足"}), 403
    conn = get_db()
    user = conn.execute("SELECT role FROM admin_users WHERE id=?", (user_id,)).fetchone()
    if not user:
        conn.close()
        return jsonify({"success": False, "error": "找不到"})
    if user["role"] == "super":
        conn.close()
        return jsonify({"success": False, "error": "無法刪除超級管理員"})
    conn.execute("DELETE FROM admin_users WHERE id=?", (user_id,))
    conn.commit()
    conn.close()
    return jsonify({"success": True})


def _mark_disabled(members):
    """為 members list 標記停用狀態（依 disabled_members 表）。"""
    try:
        conn = get_db()
        dis = {r["g_code"]: r for r in conn.execute("SELECT g_code, reason, disabled_at FROM disabled_members").fetchall()}
        conn.close()
    except Exception:
        dis = {}
    for m in members:
        d = dis.get(m.get("g_code"))
        m["disabled"] = bool(d)
        m["disabled_reason"] = d["reason"] if d else ""
        m["disabled_at"] = d["disabled_at"] if d else ""
    return members


@app.route("/api/admin/members", methods=["GET"])
def get_all_members():
    try:
        aid = get_current_agent_id()
        # ===== 代理：只看自己本地建的會員 =====
        if aid > 0:
            conn = get_db()
            agent = conn.execute("SELECT prefix, min_rate FROM agents WHERE id=?", (aid,)).fetchone()
            prefix = agent["prefix"] if agent else "X"
            min_rate = float(agent["min_rate"] or 180) if agent else 180.0
            rows = conn.execute(
                "SELECT * FROM members WHERE agent_id=? ORDER BY g_code", (aid,)
            ).fetchall()
            conn.close()
            members = []
            used_numbers = set()
            for r in rows:
                d = dict(r)
                # 會員專屬費率 > 0 → 用該費率；否則 fallback 到代理 min_rate
                member_rate = float(d.get("shipping_rate") or 0)
                effective_rate = member_rate if member_rate > 0 else min_rate
                members.append({
                    "g_code": d.get("g_code", ""),
                    "name": d.get("name", ""),
                    "phone": d.get("phone", ""),
                    "address": d.get("address", ""),
                    "line_id": d.get("line_id", ""),
                    "email": d.get("email", ""),
                    "shipping_rate": effective_rate,
                    "shipping_rate_raw": member_rate,  # 0 表示沿用 min_rate
                    "note": d.get("note", ""),
                    "status": d.get("status", "active"),
                    "source": "local",
                })
                gc = d.get("g_code", "")
                if gc.startswith(prefix):
                    try:
                        used_numbers.add(int(gc[len(prefix):]))
                    except (ValueError, TypeError):
                        pass
            max_number = max(used_numbers) if used_numbers else 0
            next_number = 1
            while next_number in used_numbers:
                next_number += 1
            next_g_code = f"{prefix}{next_number:04d}"
            return jsonify({
                "success": True,
                "members": _mark_disabled(members),
                "total": len(members),
                "max_number": max_number,
                "next_g_code": next_g_code,
                "default_shipping_rate": min_rate,
                "twd_to_jpy_rate": TWD_TO_JPY_RATE,
                "min_rate": min_rate,
                "prefix": prefix,
                "source": "agent_local",
            })

        # ===== 主管理員：Shopify + 全部本地會員（含所有代理底下的）=====
        force = request.args.get("refresh") == "1"
        members = get_all_goyoutati_customers(force_refresh=force)
        # 附加所有本地會員（含代理 agent_id>0 的、以及離職移交 agent_id=0 的）
        try:
            conn0 = get_db()
            local_rows = conn0.execute(
                "SELECT m.*, a.name AS agent_name, a.prefix AS agent_prefix, a.min_rate AS agent_min_rate "
                "FROM members m LEFT JOIN agents a ON a.id = m.agent_id "
                "ORDER BY m.g_code"
            ).fetchall()
            conn0.close()
            for r in local_rows:
                d = dict(r)
                m_rate = float(d.get("shipping_rate") or 0)
                aid_of_member = int(d.get("agent_id") or 0)
                if aid_of_member > 0:
                    # 代理底下的會員：member rate > 0 用會員專屬、否則用代理 min_rate
                    agent_min = float(d.get("agent_min_rate") or DEFAULT_SHIPPING_RATE)
                    effective_rate = m_rate if m_rate > 0 else agent_min
                    source = "agent"
                    agent_name = d.get("agent_name", "") or ""
                else:
                    # 離職移交 / 主管理員直接管
                    effective_rate = m_rate if m_rate > 0 else DEFAULT_SHIPPING_RATE
                    source = "transferred"
                    agent_name = ""
                members.append({
                    "g_code": d.get("g_code", ""),
                    "name": d.get("name", ""),
                    "phone": d.get("phone", ""),
                    "address": d.get("address", ""),
                    "line_id": d.get("line_id", ""),
                    "email": d.get("email", ""),
                    "shipping_rate": effective_rate,
                    "note": d.get("note", ""),
                    "status": d.get("status", "active"),
                    "source": source,
                    "agent_name": agent_name,
                    "agent_id": aid_of_member,
                    "customer_id": "",  # 本地會員無 Shopify ID
                })
        except Exception as e:
            print(f"[admin members] 抓本地會員失敗: {e}", flush=True)
        members.sort(key=lambda x: x["g_code"])
        used_numbers = set()
        for m in members:
            if m["g_code"].startswith("G"):
                try:
                    used_numbers.add(int(m["g_code"][1:]))
                except:
                    pass
        max_number = max(used_numbers) if used_numbers else 0
        next_number = 1
        while next_number in used_numbers:
            next_number += 1
        next_g_code = f"G{next_number:04d}"
        return jsonify({
            "success": True,
            "members": _mark_disabled(members),
            "total": len(members),
            "max_number": max_number,
            "next_g_code": next_g_code,
            "default_shipping_rate": DEFAULT_SHIPPING_RATE,  # 台幣
            "twd_to_jpy_rate": TWD_TO_JPY_RATE
        })
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})


@app.route("/api/admin/search_members", methods=["GET"])
def admin_search_members():
    """
    跨會員搜尋（給新增到貨等場景使用，倉庫人員打字即時搜尋）
    - 主管理員：搜本地 members 表 + Shopify 兩邊（含所有代理客戶）
    - 代理：只搜自己 members 表的客戶
    - 比對欄位：g_code、name、phone（去空白後）
    """
    q = (request.args.get("q") or "").strip()
    try:
        limit = int(request.args.get("limit", 15))
    except (ValueError, TypeError):
        limit = 15
    limit = max(1, min(limit, 50))

    if not q or len(q) < 1:
        return jsonify({"success": True, "results": []})

    q_upper = q.upper()
    q_phone = normalize_phone(q)
    pattern = f"%{q_upper}%"
    pattern_phone = f"%{q_phone}%"
    aid = get_current_agent_id()
    results = []

    conn = get_db()
    # 本地 members（代理建的客戶）
    if aid > 0:
        local = conn.execute("""
            SELECT g_code, name, phone, address, agent_id, status
            FROM members
            WHERE agent_id=? AND status='active'
              AND (UPPER(g_code) LIKE ? OR UPPER(name) LIKE ? OR phone LIKE ?)
            ORDER BY g_code LIMIT ?
        """, (aid, pattern, pattern, pattern_phone, limit)).fetchall()
    else:
        local = conn.execute("""
            SELECT m.g_code, m.name, m.phone, m.address, m.agent_id, m.status,
                   a.name as agent_name, a.prefix as agent_prefix
            FROM members m
            LEFT JOIN agents a ON a.id = m.agent_id
            WHERE m.status='active'
              AND (UPPER(m.g_code) LIKE ? OR UPPER(m.name) LIKE ? OR m.phone LIKE ?)
            ORDER BY m.g_code LIMIT ?
        """, (pattern, pattern, pattern_phone, limit)).fetchall()
    conn.close()

    for r in local:
        d = dict(r)
        results.append({
            "g_code": d.get("g_code"),
            "name": d.get("name") or "",
            "phone": d.get("phone") or "",
            "address": d.get("address") or "",
            "source": "agent",
            "agent_id": d.get("agent_id") or 0,
            "agent_name": d.get("agent_name") or "",
        })

    # Shopify 客戶（僅主管理員視角）
    if aid == 0:
        try:
            customers = get_all_goyoutati_customers()
            for c in customers:
                if len(results) >= limit:
                    break
                gc = (c.get("g_code") or "").upper()
                nm = (c.get("name") or "").upper()
                ph = c.get("phone") or ""
                if q_upper in gc or q_upper in nm or (q_phone and q_phone in ph):
                    results.append({
                        "g_code": c.get("g_code"),
                        "name": c.get("name") or "",
                        "phone": ph,
                        "address": c.get("address") or "",
                        "source": "shopify",
                        "agent_id": 0,
                        "agent_name": "",
                    })
        except Exception as e:
            print(f"[search_members] Shopify 搜尋失敗：{e}", flush=True)

    return jsonify({"success": True, "results": results[:limit]})


# ===== 停用 / 啟用會員（集運系統層級，不動 Shopify）=====

@app.route("/api/admin/members/<g_code>/disable", methods=["POST"])
def admin_disable_member(g_code):
    if not is_super_admin():
        return jsonify({"success": False, "error": "權限不足"}), 403
    g_code = (g_code or "").strip().upper()
    if not g_code:
        return jsonify({"success": False, "error": "缺少會員編號"}), 400
    reason = ((request.json or {}).get("reason") or "").strip()
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    conn = get_db()
    conn.execute(
        "INSERT INTO disabled_members (g_code, reason, disabled_at) VALUES (?, ?, ?) "
        "ON CONFLICT(g_code) DO UPDATE SET reason=excluded.reason, disabled_at=excluded.disabled_at",
        (g_code, reason, now)
    )
    conn.commit()
    conn.close()
    return jsonify({"success": True})


@app.route("/api/admin/members/<g_code>/disable", methods=["DELETE"])
def admin_enable_member(g_code):
    if not is_super_admin():
        return jsonify({"success": False, "error": "權限不足"}), 403
    g_code = (g_code or "").strip().upper()
    conn = get_db()
    conn.execute("DELETE FROM disabled_members WHERE g_code=?", (g_code,))
    conn.commit()
    conn.close()
    return jsonify({"success": True})


@app.route("/api/admin/members", methods=["POST"])
def admin_create_member():
    """代理新增會員（自動補前綴與 agent_id）。主管理員建議直接在 Shopify 操作。"""
    aid = get_current_agent_id()
    if aid <= 0:
        return jsonify({"success": False, "error": "主管理員的會員請在 Shopify 後台建立"}), 400
    data = request.json or {}
    name = (data.get("name") or "").strip()
    phone = normalize_phone((data.get("phone") or "").strip())
    address = (data.get("address") or "").strip()
    if not name:
        return jsonify({"success": False, "error": "會員姓名必填"})
    if not phone:
        return jsonify({"success": False, "error": "電話必填（會員用此登入）"})

    conn = get_db()
    ag = conn.execute("SELECT prefix FROM agents WHERE id=?", (aid,)).fetchone()
    if not ag:
        conn.close()
        return jsonify({"success": False, "error": "代理資料異常"}), 500
    prefix = ag["prefix"]

    # 自動產生下一個 g_code（前綴+四位流水）
    g_code_in = (data.get("g_code") or "").strip().upper()
    if g_code_in:
        # 手動指定的：必須以該代理前綴開頭
        if not g_code_in.startswith(prefix):
            conn.close()
            return jsonify({"success": False, "error": f"會員編號必須以「{prefix}」開頭"})
        # 不可重複
        exists = conn.execute("SELECT 1 FROM members WHERE g_code=?", (g_code_in,)).fetchone()
        if exists:
            conn.close()
            return jsonify({"success": False, "error": f"編號「{g_code_in}」已使用"})
        g_code = g_code_in
    else:
        used = set()
        for r in conn.execute("SELECT g_code FROM members WHERE agent_id=?", (aid,)).fetchall():
            gc = r["g_code"] or ""
            if gc.startswith(prefix):
                try:
                    used.add(int(gc[len(prefix):]))
                except (ValueError, TypeError):
                    pass
        n = 1
        while n in used:
            n += 1
        g_code = f"{prefix}{n:04d}"

    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    # 處理會員專屬費率（>= 代理 min_rate；留空/0 = 沿用 min_rate）
    raw_rate = data.get("shipping_rate")
    rate_val = 0.0
    if raw_rate not in (None, "", 0, "0"):
        try:
            rate_val = float(raw_rate)
        except (ValueError, TypeError):
            conn.close()
            return jsonify({"success": False, "error": "運費必須為數字"})
        # 代理可自由設定費率，無下限
    try:
        conn.execute(
            """INSERT INTO members (g_code, agent_id, name, phone, address, line_id, email, note, status, created_at, shipping_rate)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, 'active', ?, ?)""",
            (g_code, aid, name, phone, address,
             (data.get("line_id") or "").strip(), (data.get("email") or "").strip(),
             (data.get("note") or "").strip(), now, rate_val)
        )
        conn.commit()
    except sqlite3.IntegrityError as e:
        conn.close()
        return jsonify({"success": False, "error": f"資料庫錯誤：{e}"})
    conn.close()
    return jsonify({"success": True, "g_code": g_code, "message": f"已建立「{g_code} {name}」"})


@app.route("/api/admin/members/<g_code>", methods=["PUT"])
def admin_update_member(g_code):
    """代理更新自己會員的資料（姓名、電話、地址等）"""
    g_code = g_code.upper()
    aid = get_current_agent_id()
    conn = get_db()
    row = conn.execute("SELECT * FROM members WHERE g_code=?", (g_code,)).fetchone()
    if not row:
        conn.close()
        return jsonify({"success": False, "error": "找不到會員"})
    if aid > 0 and int(row["agent_id"] or 0) != aid:
        conn.close()
        return jsonify({"success": False, "error": "權限不足"}), 403

    data = request.json or {}
    fields, values = [], []
    for col in ["name", "phone", "address", "line_id", "email", "note", "status"]:
        if col in data:
            v = (data[col] or "").strip()
            if col == "phone":
                v = normalize_phone(v)
            if col == "status" and v not in ("active", "disabled"):
                continue
            fields.append(f"{col}=?"); values.append(v)
    # 會員專屬費率
    if "shipping_rate" in data:
        raw = data["shipping_rate"]
        if raw in (None, "", 0, "0"):
            rate_val = 0.0
        else:
            try:
                rate_val = float(raw)
            except (ValueError, TypeError):
                conn.close()
                return jsonify({"success": False, "error": "運費必須為數字"})
            # 代理可自由設定費率，無下限
        fields.append("shipping_rate=?"); values.append(rate_val)
    if not fields:
        conn.close()
        return jsonify({"success": False, "error": "沒有可更新欄位"})
    values.append(g_code)
    conn.execute(f"UPDATE members SET {', '.join(fields)} WHERE g_code=?", values)
    conn.commit()
    conn.close()
    return jsonify({"success": True})


@app.route("/api/admin/members/<g_code>", methods=["DELETE"])
def admin_delete_member(g_code):
    """代理刪除自己會員（若會員底下有任何包裹/預報/出貨紀錄，擋下，建議停用）"""
    g_code = g_code.upper()
    aid = get_current_agent_id()
    conn = get_db()
    row = conn.execute("SELECT * FROM members WHERE g_code=?", (g_code,)).fetchone()
    if not row:
        conn.close()
        return jsonify({"success": False, "error": "找不到會員"})
    if aid > 0 and int(row["agent_id"] or 0) != aid:
        conn.close()
        return jsonify({"success": False, "error": "權限不足"}), 403
    # 安全檢查：是否有關聯資料
    p = conn.execute("SELECT COUNT(*) as c FROM packages WHERE g_code=?", (g_code,)).fetchone()["c"]
    f = conn.execute("SELECT COUNT(*) as c FROM forecasts WHERE g_code=?", (g_code,)).fetchone()["c"]
    s = conn.execute("SELECT COUNT(*) as c FROM shipment_requests WHERE g_code=?", (g_code,)).fetchone()["c"]
    if p + f + s > 0:
        conn.close()
        return jsonify({
            "success": False,
            "error": f"此會員已有 {p} 個包裹/{f} 個預報/{s} 個出貨紀錄，無法刪除。請改為「停用」狀態。"
        })
    conn.execute("DELETE FROM members WHERE g_code=?", (g_code,))
    conn.commit()
    conn.close()
    return jsonify({"success": True, "message": "會員已刪除"})


@app.route("/api/admin/shipping_rate", methods=["POST"])
def set_shipping_rate():
    data = request.json
    customer_gid = data.get("customer_gid", "")
    shipping_rate = data.get("shipping_rate", "")  # 台幣
    if not customer_gid:
        return jsonify({"success": False, "error": "缺少客戶 ID"})
    if shipping_rate == "" or shipping_rate is None:
        return jsonify({"success": False, "error": "請輸入運費"})
    try:
        rate_val = int(shipping_rate)
        if rate_val < 0:
            return jsonify({"success": False, "error": "運費不能為負數"})
    except ValueError:
        return jsonify({"success": False, "error": "運費必須為整數"})

    mutation = """
    mutation metafieldsSet($metafields: [MetafieldsSetInput!]!) {
        metafieldsSet(metafields: $metafields) {
            metafields { key value }
            userErrors { field message }
        }
    }
    """
    variables = {
        "metafields": [{
            "ownerId": customer_gid,
            "namespace": "custom",
            "key": "shipping_rate",
            "type": "single_line_text_field",
            "value": str(rate_val)  # 儲存台幣值
        }]
    }
    try:
        result = shopify_graphql(mutation, variables)
        if "data" in result:
            mutation_result = result["data"].get("metafieldsSet", {})
            user_errors = mutation_result.get("userErrors", [])
            if user_errors:
                return jsonify({"success": False, "error": "; ".join([e["message"] for e in user_errors])})
            if mutation_result.get("metafields"):
                return jsonify({
                    "success": True,
                    "shipping_rate_twd": rate_val,
                    "shipping_rate_jpy": twd_to_jpy(rate_val)
                })
        if "errors" in result:
            return jsonify({"success": False, "error": str(result["errors"])})
        return jsonify({"success": False, "error": "設定失敗，請重試"})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})


# ============ 管理員：到貨包裹管理 ============

@app.route("/api/admin/packages", methods=["GET"])
def admin_list_packages():
    g_code = request.args.get("g_code", "")
    aid = get_current_agent_id()
    conn = get_db()
    if g_code:
        if aid > 0:
            rows = conn.execute(
                "SELECT * FROM packages WHERE g_code=? AND agent_id=? ORDER BY id DESC",
                (g_code.upper(), aid)
            ).fetchall()
        else:
            rows = conn.execute(
                "SELECT * FROM packages WHERE g_code=? ORDER BY id DESC", (g_code.upper(),)
            ).fetchall()
    else:
        if aid > 0:
            rows = conn.execute(
                "SELECT * FROM packages WHERE agent_id=? ORDER BY id DESC", (aid,)
            ).fetchall()
        else:
            rows = conn.execute("SELECT * FROM packages ORDER BY id DESC").fetchall()
    conn.close()
    return jsonify({"success": True, "packages": [dict(r) for r in rows]})


@app.route("/api/admin/packages", methods=["POST"])
def admin_add_package():
    data = request.json
    g_code      = (data.get("g_code") or "").strip().upper()
    logis_num   = (data.get("logis_num") or "").strip()
    product_name= (data.get("product_name") or "").strip()
    weight      = (data.get("weight") or "").strip()
    note        = (data.get("note") or "").strip()
    status      = data.get("status", "已到貨")
    pkg_type    = data.get("pkg_type", "包裹")
    if pkg_type not in ("包裹", "信件"):
        pkg_type = "包裹"

    if not g_code:
        return jsonify({"success": False, "error": "請輸入客戶編號"})
    # 開頭非字母 → 補上對應前綴（代理用自己的前綴、主管理員預設 G）
    if not g_code[:1].isalpha():
        aid = get_current_agent_id()
        if aid > 0:
            conn0 = get_db()
            ag = conn0.execute("SELECT prefix FROM agents WHERE id=?", (aid,)).fetchone()
            conn0.close()
            g_code = (ag["prefix"] if ag else "G") + g_code
        else:
            g_code = "G" + g_code

    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    today = datetime.now().strftime("%Y-%m-%d")
    pkg_agent_id = get_agent_id_for_g_code(g_code)

    # 代理只能幫自己的客戶建包裹
    aid = get_current_agent_id()
    if aid > 0 and pkg_agent_id != aid:
        return jsonify({"success": False, "error": f"客戶編號「{g_code}」不屬於你的代理帳號"}), 403

    conn = get_db()
    cur = conn.execute(
        """INSERT INTO packages (g_code, logis_num, product_name, weight, status, note, in_date, created_at, agent_id, pkg_type)
           VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
        (g_code, logis_num, product_name, weight, status, note, today, now, pkg_agent_id, pkg_type)
    )
    new_id = cur.lastrowid
    conn.commit()
    conn.close()
    return jsonify({"success": True, "id": new_id})


@app.route("/api/admin/packages/<int:pkg_id>", methods=["PUT"])
def admin_update_package(pkg_id):
    ok, row = check_record_ownership("packages", pkg_id)
    if not row:
        return jsonify({"success": False, "error": "找不到包裹"})
    if not ok:
        return jsonify({"success": False, "error": "權限不足"}), 403
    data = request.json
    fields = []
    values = []
    for key in ["g_code", "logis_num", "product_name", "weight", "status", "note", "in_date", "pkg_type"]:
        if key in data:
            val = data[key]
            if key == "g_code":
                val = val.strip().upper()
                # 改 g_code 時驗證新 g_code 仍屬於同代理
                aid_chk = get_current_agent_id()
                if aid_chk > 0 and get_agent_id_for_g_code(val) != aid_chk:
                    return jsonify({"success": False, "error": f"無法將包裹改至非自己客戶「{val}」"}), 403
            fields.append(f"{key}=?")
            values.append(val)
    if not fields:
        return jsonify({"success": False, "error": "沒有要更新的欄位"})
    values.append(pkg_id)
    conn = get_db()
    conn.execute(f"UPDATE packages SET {', '.join(fields)} WHERE id=?", values)
    conn.commit()
    conn.close()
    return jsonify({"success": True})


@app.route("/api/admin/packages/bulk_ship", methods=["POST"])
def admin_bulk_ship():
    data = request.json
    ids = data.get("ids", [])
    if not ids:
        return jsonify({"success": False, "error": "沒有選取任何包裹"})
    aid = get_current_agent_id()
    conn = get_db()
    if aid > 0:
        # 驗證所有 id 都屬於該代理
        placeholders = ",".join(["?"] * len(ids))
        owned = conn.execute(
            f"SELECT id FROM packages WHERE id IN ({placeholders}) AND agent_id=?",
            ids + [aid]
        ).fetchall()
        owned_ids = [r["id"] for r in owned]
        if len(owned_ids) != len(ids):
            conn.close()
            return jsonify({"success": False, "error": "部分包裹不屬於你的代理帳號"}), 403
        ids = owned_ids
    conn.execute(
        f"UPDATE packages SET status='已出貨' WHERE id IN ({','.join(['?']*len(ids))})",
        ids
    )
    conn.commit()
    conn.close()
    return jsonify({"success": True, "updated": len(ids)})


@app.route("/api/admin/packages/<int:pkg_id>", methods=["DELETE"])
def admin_delete_package(pkg_id):
    ok, row = check_record_ownership("packages", pkg_id)
    if not row:
        return jsonify({"success": False, "error": "找不到包裹"})
    if not ok:
        return jsonify({"success": False, "error": "權限不足"}), 403
    conn = get_db()
    conn.execute("DELETE FROM packages WHERE id=?", (pkg_id,))
    conn.commit()
    conn.close()
    return jsonify({"success": True})


# ============ 客戶端 API ============

@app.route("/api/verify_customer", methods=["POST"])
def verify_customer():
    data = request.json
    g_code = data.get("customer_id", "").strip().upper()
    password = data.get("password", "").strip()
    if not g_code:
        return jsonify({"success": False, "error": "請輸入會員編號"})
    if not password:
        return jsonify({"success": False, "error": "請輸入密碼"})
    # 沒有英文前綴 → 預設加 G（你的客戶）
    if not g_code[:1].isalpha():
        g_code = "G" + g_code
    password_clean = normalize_phone(password)

    # ===== 0) 停用名單檢查（集運系統層級）：被停用者一律擋下，整頁顯示停用訊息 =====
    try:
        _dconn = get_db()
        _drow = _dconn.execute("SELECT 1 FROM disabled_members WHERE g_code=?", (g_code,)).fetchone()
        _dconn.close()
        if _drow:
            return jsonify({"success": False, "disabled": True, "error": "您的帳號已停用，請聯繫客服"})
    except Exception:
        pass

    # ===== 1) 先查本地 members 表（代理建的會員）=====
    try:
        conn = get_db()
        row = conn.execute("SELECT * FROM members WHERE g_code=?", (g_code,)).fetchone()
        if row:
            m = dict(row)
            # 狀態檢查
            if m.get("status") == "disabled":
                conn.close()
                return jsonify({"success": False, "error": "此會員帳號已停用，請聯絡您的代理"})
            # 比對電話（去除空白、橫線、+886/+81 等）
            stored_phone = normalize_phone(m.get("phone") or "")
            if stored_phone != password_clean:
                conn.close()
                return jsonify({"success": False, "error": "密碼錯誤，請輸入您的手機號碼"})
            # 找該代理（含品牌欄位）
            ag = conn.execute("SELECT * FROM agents WHERE id=?", (m["agent_id"],)).fetchone()
            conn.close()
            agent_min = float(ag["min_rate"]) if ag and ag["min_rate"] else float(DEFAULT_SHIPPING_RATE)
            # 會員專屬費率 > 0 → 用該費率；否則用代理 min_rate
            member_rate = float(m.get("shipping_rate") or 0)
            rate_twd = int(member_rate if member_rate > 0 else agent_min)
            branding = _branding_dict(ag) if ag else _branding_dict(None)
            return jsonify({
                "success": True,
                "customer": {
                    "id": g_code,  # 本地會員無 Shopify customer_id，用 g_code
                    "g_code": g_code,
                    "name": m.get("name") or "會員",
                    "email": m.get("email") or "",
                    "phone": stored_phone,
                    "phone_raw": m.get("phone") or "",
                    "address": m.get("address") or "",
                    "shipping_rate_twd": rate_twd,
                    "shipping_rate_jpy": twd_to_jpy(rate_twd) if rate_twd else 0,
                    "source": "agent",
                    "agent_name": ag["name"] if ag else "",
                    "branding": branding,
                }
            })
        conn.close()
    except Exception as e:
        print(f"[verify_customer] 本地查詢失敗：{e}", flush=True)

    # ===== 2) 回退查 Shopify（你的客戶，原有邏輯）=====
    try:
        customers = get_all_goyoutati_customers()
        for c in customers:
            if c["g_code"] == g_code:
                if c["phone"] and c["phone"] == password_clean:
                    try:
                        rate_twd = int(c["shipping_rate"]) if c["shipping_rate"] else DEFAULT_SHIPPING_RATE
                    except (ValueError, TypeError):
                        rate_twd = DEFAULT_SHIPPING_RATE
                    rate_jpy = twd_to_jpy(rate_twd) if rate_twd else 0
                    return jsonify({
                        "success": True,
                        "customer": {
                            "id": c["customer_id"],
                            "g_code": g_code,
                            "name": c["name"] or "會員",
                            "email": c["email"],
                            "phone": c["phone"],
                            "phone_raw": c["phone_raw"],
                            "address": c.get("address", ""),
                            "shipping_rate_twd": rate_twd,
                            "shipping_rate_jpy": rate_jpy,
                            "source": "shopify",
                        }
                    })
                else:
                    return jsonify({"success": False, "error": "密碼錯誤，請輸入您的手機號碼"})
        return jsonify({"success": False, "error": "找不到此會員編號，請確認後重試"})
    except Exception as e:
        return jsonify({"success": False, "error": f"查詢失敗: {str(e)}"})


@app.route("/api/forecast", methods=["POST"])
def create_forecast():
    data = request.json
    customer_id = data.get("customer_id")
    g_code = data.get("g_code", "")
    packages = data.get("packages", [])

    if not customer_id:
        return jsonify({"success": False, "error": "缺少客戶編號"})
    if not packages:
        return jsonify({"success": False, "error": "請至少填寫一個包裹"})

    results = []
    for idx, pkg in enumerate(packages):
        timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
        local_logis_num = f"{g_code}-{timestamp}-{idx+1}"
        declare_list = []
        for item in pkg.get("items", []):
            declare_list.append({
                "product_name": item.get("name", "商品"),
                "product_name_local": item.get("name", "商品"),
                "product_num": int(item.get("quantity", 1)),
                "product_price": int(float(item.get("price", 0))),
                "product_url": item.get("url", "")
            })
        total_num = sum(int(item.get("quantity", 1)) for item in pkg.get("items", []))
        total_price = sum(int(float(item.get("price", 0))) * int(item.get("quantity", 1)) for item in pkg.get("items", []))
        forecast_data = {
            "packages": [{
                "local_logis_num": local_logis_num,
                "client_cid": g_code,
                "client_pid": pkg.get("client_pid") or local_logis_num,
                "client_date": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "warehouse_id": JPD_WAREHOUSE_ID,
                "product_name": declare_list[0]["product_name"] if declare_list else "商品",
                "product_num": total_num,
                "product_price": total_price,
                "declare_list": declare_list
            }]
        }
        result = jpd_request("TForecastPackage", forecast_data)
        if "OperationResult" in result:
            op_result = result["OperationResult"]
            if op_result["Request"]["IsValid"] == "True":
                result_data = op_result.get("Result", {})
                if result_data.get("Result") == "SUCCESS":
                    pkg_data = result_data.get("Data", [{}])[0]
                    results.append({
                        "success": True,
                        "local_logis_num": local_logis_num,
                        "package_id": pkg_data.get("package_id"),
                        "message": pkg_data.get("msg", "預報成功")
                    })
                    continue
        results.append({"success": False, "local_logis_num": local_logis_num, "error": "預報失敗"})

    return jsonify({"success": all(r["success"] for r in results), "results": results})


@app.route("/api/packages", methods=["GET"])
def get_packages():
    g_code = request.args.get("g_code") or request.args.get("customer_id")
    if not g_code:
        return jsonify({"success": False, "error": "缺少會員編號"})

    g_code = g_code.upper()
    conn = get_db()
    rows = conn.execute(
        "SELECT * FROM packages WHERE g_code=? ORDER BY id DESC",
        (g_code,)
    ).fetchall()

    # 找出該會員「進行中」的出貨申請（待處理／處理中），標記其包含的包裹
    # 用途：前端據此隱藏勾選框、顯示「已申請出貨」徽章，避免重複申請
    active_reqs = conn.execute(
        "SELECT id, package_ids FROM shipment_requests "
        "WHERE g_code=? AND status IN ('待處理', '處理中')",
        (g_code,)
    ).fetchall()
    conn.close()

    pkg_to_req = {}
    for r in active_reqs:
        ids_str = r["package_ids"] or ""
        for pid_str in ids_str.split(","):
            try:
                pkg_to_req[int(pid_str.strip())] = r["id"]
            except (ValueError, AttributeError):
                pass

    packages = []
    for row in rows:
        r = dict(row)
        packages.append({
            "id":           r["id"],
            "logis_num":    r["logis_num"] or "-",
            "product_name": r["product_name"] or "-",
            "weight":       r["weight"] or "",
            "status":       r["status"],
            "note":         r["note"] or "",
            "in_date":      r["in_date"] or "",
            "created_at":   r["created_at"],
            "pending_ship_request_id": pkg_to_req.get(r["id"]),  # None 表示未在出貨申請中
        })
    return jsonify({"success": True, "packages": packages})


@app.route("/api/orders", methods=["GET"])
def get_orders():
    g_code = request.args.get("g_code") or request.args.get("customer_id")
    if not g_code:
        return jsonify({"success": False, "error": "缺少會員編號"})
    result = jpd_request("TSearchOrders", {
        "client_cid": g_code,
        "warehouse_id": JPD_WAREHOUSE_ID
    })
    if "OperationResult" in result:
        op_result = result["OperationResult"]
        if op_result["Request"]["IsValid"] == "True":
            orders = op_result.get("Result", {}).get("Data", [])
            formatted = []
            for order in orders:
                formatted.append({
                    "order_id": order.get("order_id"),
                    "customer_order_id": order.get("customer_order_id"),
                    "logis_num": order.get("logis_num"),
                    "status": order.get("status_name"),
                    "recipient": order.get("recipient"),
                    "create_date": order.get("create_date"),
                    "weight": order.get("weight"),
                    "deliv_fee": order.get("deliv_fee")
                })
            return jsonify({"success": True, "orders": formatted})
    return jsonify({"success": False, "error": "查詢失敗"})


# ============ 統計 API ============

def compute_agent_weekly(agent_id, conn=None):
    """算某代理各週分潤（與統計頁同一套算法，保證數字一致）。
    只算：已出貨 + 有金額 + 已收款（有匯款後五碼）。
    分潤 = Σ per_kg × kg，per_kg = max(該單費率 − 180, 20)。
    回傳 [{period_key, period_label, shipments, total_kg, commission, payout:{...}}, ...] 新→舊
    """
    own = False
    if conn is None:
        conn = get_db(); own = True
    rows = conn.execute("""
        SELECT * FROM shipment_requests
        WHERE status='已出貨' AND total_fee > 0 AND agent_id=?
          AND payment_last5 IS NOT NULL AND payment_last5 != ''
    """, (agent_id,)).fetchall()
    payouts = {
        p["period_key"]: dict(p)
        for p in conn.execute("SELECT * FROM agent_payouts WHERE agent_id=?", (agent_id,)).fetchall()
    }
    if own:
        conn.close()

    buckets = {}
    for row in rows:
        r = dict(row)
        date_str = r.get("updated_at") or r.get("created_at") or ""
        key, label = _period_key_from_date(date_str, "week")
        if not key:
            continue
        b = buckets.setdefault(key, {
            "period_key": key, "period_label": label,
            "shipments": 0, "total_kg": 0.0, "commission": 0.0,
        })
        kg = float(r.get("billed_weight") or 0)
        b["shipments"] += 1
        b["total_kg"] += kg
        if kg > 0:
            rate = float(r.get("rate_per_kg") or 0)
            b["commission"] += max(rate - 180, 20) * kg

    result = []
    for key in sorted(buckets.keys(), reverse=True):
        b = buckets[key]
        b["total_kg"] = round(b["total_kg"], 1)
        b["commission"] = round(b["commission"])
        p = payouts.get(key)
        b["paid"] = bool(p and p.get("paid_at"))
        b["payment_last5"] = (p or {}).get("payment_last5", "")
        b["paid_at"] = (p or {}).get("paid_at", "")
        b["payout_note"] = (p or {}).get("note", "")
        result.append(b)
    return result

@app.route("/api/admin/stats/monthly/detail", methods=["GET"])
def admin_monthly_detail():
    """取得指定週/月的出貨明細（month 參數值可為 '2026-06' 或 '2026-W23'）"""
    month = request.args.get("month", "")
    if not month:
        return jsonify({"success": False, "error": "缺少期間"})
    aid = get_current_agent_id()
    try:
        conn = get_db()
        if aid > 0:
            rows = conn.execute("""
                SELECT * FROM shipment_requests
                WHERE status='已出貨' AND total_fee > 0 AND agent_id=?
                ORDER BY updated_at ASC
            """, (aid,)).fetchall()
        else:
            rows = conn.execute("""
                SELECT * FROM shipment_requests
                WHERE status='已出貨' AND total_fee > 0
                ORDER BY updated_at ASC
            """).fetchall()
        conn.close()
        details = []
        for r in rows:
            rd = dict(r)
            date_str = rd.get("updated_at") or rd.get("created_at") or ""
            if not _matches_period(date_str, month):
                continue
            extras = []
            try:
                extras = json.loads(rd.get("extra_services") or "[]")
            except:
                pass
            details.append({
                "date": date_str[:10],
                "g_code": rd.get("g_code", ""),
                "customer_name": rd.get("customer_name", ""),
                "ship_recipient": rd.get("ship_recipient", ""),
                "ship_phone": rd.get("ship_phone", ""),
                "ship_address": rd.get("ship_address", ""),
                "billed_weight": float(rd.get("billed_weight") or 0),
                "rate_per_kg": float(rd.get("rate_per_kg") or 0),
                "shipping_fee": float(rd.get("shipping_fee") or 0),
                "handling_fee": float(rd.get("handling_fee") or 0),
                "consolidation_fee": float(rd.get("consolidation_fee") or 0),
                "letter_fee": float(rd.get("letter_fee") or 0),
                "extra_services": extras,
                "total_fee": float(rd.get("total_fee") or 0),
            })
        # 代理檢視時計算分潤（依合約第六條第 4 項公式）
        # 公式：(該客戶運費 − NT$180/kg) × 包裹重量，最低 NT$20/kg × 包裹重量
        commission = None
        if aid > 0:
            base_cost = 180  # NT$/kg 甲方批發成本
            min_per_kg = 20  # NT$/kg 最低分潤
            total_commission = 0
            total_kg = 0
            for d in details:
                kg = d["billed_weight"]
                per_kg = max(d["rate_per_kg"] - base_cost, min_per_kg)
                d["commission"] = round(per_kg * kg)
                total_commission += d["commission"]
                total_kg += kg
            commission = {
                "total": round(total_commission),
                "total_kg": round(total_kg, 1),
                "min_per_kg": min_per_kg,
                "base_cost_per_kg": base_cost,
            }
        return jsonify({"success": True, "details": details, "is_agent": aid > 0, "commission": commission})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})


@app.route("/api/admin/stats/monthly/excel", methods=["GET"])
def admin_monthly_excel():
    """下載指定月份的出貨明細 Excel"""
    month = request.args.get("month", "")  # e.g. "2026-04"
    if not month:
        return jsonify({"success": False, "error": "缺少月份參數"})
    aid = get_current_agent_id()
    try:
        conn = get_db()
        if aid > 0:
            rows = conn.execute("""
                SELECT * FROM shipment_requests
                WHERE status='已出貨' AND total_fee > 0 AND agent_id=?
                ORDER BY updated_at ASC
            """, (aid,)).fetchall()
        else:
            rows = conn.execute("""
                SELECT * FROM shipment_requests
                WHERE status='已出貨' AND total_fee > 0
                ORDER BY updated_at ASC
            """).fetchall()
        conn.close()

        # 篩選指定週/月
        filtered = []
        for r in rows:
            rd = dict(r)
            date_str = rd.get("updated_at") or rd.get("created_at") or ""
            if _matches_period(date_str, month):
                filtered.append(rd)

        wb = Workbook()
        ws = wb.active
        ws.title = f"{month} 出貨明細"

        # 標題樣式
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )

        headers = ["出貨日期", "客戶編號", "客戶姓名", "寄送地址", "計費重量(kg)",
                    "運費單價", "運費小計", "理貨費", "加值服務明細", "加值服務小計", "信件費", "合計(台幣)"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        total_kg = 0
        total_shipping = 0
        total_handling = 0
        total_extra = 0
        total_letter = 0
        total_all = 0

        for i, r in enumerate(filtered, 2):
            date_str = r["updated_at"] or r["created_at"] or ""
            bw = float(r["billed_weight"] or 0)
            rate = float(r["rate_per_kg"] or 0)
            sf = float(r["shipping_fee"] or 0)
            hf = float(r["handling_fee"] or 0)
            lf = float(r["letter_fee"] or 0)
            tf = float(r["total_fee"] or 0)

            # 加值服務
            extra_desc = ""
            extra_total = 0
            try:
                extras = json.loads(r["extra_services"] or "[]")
                parts = []
                for e in extras:
                    qty = int(e.get("qty", 1))
                    price = int(e.get("price", 0))
                    sub = int(e.get("subtotal", qty * price))
                    parts.append(f"{e.get('name','')} ×{qty} = NT${sub}")
                    extra_total += sub
                extra_desc = " / ".join(parts)
            except:
                pass

            ship_addr = " ".join(filter(None, [str(r.get("ship_recipient") or ""), str(r.get("ship_phone") or ""), str(r.get("ship_address") or "")]))

            ws.cell(row=i, column=1, value=date_str[:10]).border = thin_border
            ws.cell(row=i, column=2, value=r["g_code"]).border = thin_border
            ws.cell(row=i, column=3, value=str(r.get("customer_name") or "")).border = thin_border
            ws.cell(row=i, column=4, value=ship_addr).border = thin_border
            ws.cell(row=i, column=5, value=bw).border = thin_border
            ws.cell(row=i, column=6, value=rate).border = thin_border
            ws.cell(row=i, column=7, value=sf).border = thin_border
            ws.cell(row=i, column=8, value=hf).border = thin_border
            ws.cell(row=i, column=9, value=extra_desc).border = thin_border
            ws.cell(row=i, column=10, value=extra_total).border = thin_border
            ws.cell(row=i, column=11, value=lf).border = thin_border
            ws.cell(row=i, column=12, value=tf).border = thin_border

            total_kg += bw
            total_shipping += sf
            total_handling += hf
            total_extra += extra_total
            total_letter += lf
            total_all += tf

        # 合計列
        sum_row = len(filtered) + 2
        sum_font = Font(bold=True, size=11)
        sum_fill = PatternFill(start_color="F39C12", end_color="F39C12", fill_type="solid")
        ws.cell(row=sum_row, column=1, value="合計").font = sum_font
        ws.cell(row=sum_row, column=1).fill = sum_fill
        ws.cell(row=sum_row, column=1).border = thin_border
        for c in range(2, 13):
            ws.cell(row=sum_row, column=c).border = thin_border
            ws.cell(row=sum_row, column=c).font = sum_font
        ws.cell(row=sum_row, column=2, value=f"{len(filtered)} 筆")
        ws.cell(row=sum_row, column=5, value=total_kg)
        ws.cell(row=sum_row, column=7, value=total_shipping)
        ws.cell(row=sum_row, column=8, value=total_handling)
        ws.cell(row=sum_row, column=10, value=total_extra)
        ws.cell(row=sum_row, column=11, value=total_letter)
        ws.cell(row=sum_row, column=12, value=total_all)

        # 欄寬
        widths = {'A':12, 'B':10, 'C':12, 'D':30, 'E':12, 'F':10, 'G':12, 'H':10, 'I':30, 'J':12, 'K':12}
        for col_letter, w in widths.items():
            ws.column_dimensions[col_letter].width = w

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        filename = f"GOYOUTATI_{month}_出貨明細.xlsx"
        return send_file(buf, as_attachment=True, download_name=filename,
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})


def _period_key_from_date(date_str, period_type):
    """
    把 'YYYY-MM-DD HH:MM:SS' 轉成 (period_key, period_label)
    - month → ('2026-06', '2026-06')
    - week  → ('2026-W23', '6/1 - 6/7')  (ISO 週、週一開始)
    """
    if period_type == "week":
        from datetime import datetime as _dt, timedelta as _td
        try:
            d = _dt.strptime(date_str[:10], "%Y-%m-%d")
        except Exception:
            return None, None
        iso_year, iso_week, iso_weekday = d.isocalendar()
        monday = d - _td(days=iso_weekday - 1)
        sunday = monday + _td(days=6)
        key = f"{iso_year}-W{iso_week:02d}"
        if monday.year == sunday.year and monday.month == sunday.month:
            label = f"{monday.month}/{monday.day} - {sunday.day}"
        elif monday.year == sunday.year:
            label = f"{monday.month}/{monday.day} - {sunday.month}/{sunday.day}"
        else:
            label = f"{monday.year}/{monday.month}/{monday.day} - {sunday.year}/{sunday.month}/{sunday.day}"
        return key, label
    else:
        # month
        if len(date_str) < 7:
            return None, None
        key = date_str[:7]
        return key, key


def _matches_period(date_str, period_key):
    """date_str 是否屬於 period_key（自動偵測週/月）"""
    if not date_str or not period_key:
        return False
    if "W" in period_key:
        k, _ = _period_key_from_date(date_str, "week")
        return k == period_key
    else:
        return date_str[:7] == period_key


@app.route("/api/admin/stats/monthly", methods=["GET"])
def admin_monthly_stats():
    """月/週統計：代理→按週、主帳號→可選月/週（?period=month|week，預設 month）"""
    aid = get_current_agent_id()
    if aid > 0:
        period_type = "week"
    else:
        period_type = request.args.get("period", "month")
        if period_type not in ("month", "week"):
            period_type = "month"
    try:
        conn = get_db()
        if aid > 0:
            rows = conn.execute("""
                SELECT * FROM shipment_requests
                WHERE status='已出貨' AND total_fee > 0 AND agent_id=?
                  AND payment_last5 IS NOT NULL AND payment_last5 != ''
                ORDER BY updated_at DESC
            """, (aid,)).fetchall()
        else:
            rows = conn.execute("""
                SELECT * FROM shipment_requests
                WHERE status='已出貨' AND total_fee > 0
                  AND payment_last5 IS NOT NULL AND payment_last5 != ''
                ORDER BY updated_at DESC
            """).fetchall()
        conn.close()

        buckets = {}
        for row in rows:
            r = dict(row)
            date_str = r.get("updated_at") or r.get("created_at") or ""
            if not date_str:
                continue
            key, label = _period_key_from_date(date_str, period_type)
            if not key:
                continue

            if key not in buckets:
                buckets[key] = {
                    "month": key,            # 保留欄位名稱以維持向後相容（前端、明細 API、Excel 都用此）
                    "period_label": label,
                    "shipments": 0,
                    "total_kg": 0,
                    "shipping_fee": 0,
                    "handling_fee": 0,
                    "consolidation_fee": 0,
                    "letter_fee": 0,
                    "extra_fee": 0,
                    "total_revenue": 0,
                    "commission": 0,        # 代理分潤累計（主帳號為 0）
                    "customers": set()
                }

            m = buckets[key]
            m["shipments"] += 1
            kg = float(r["billed_weight"] or 0)
            m["total_kg"] += kg
            m["shipping_fee"] += float(r["shipping_fee"] or 0)
            m["handling_fee"] += float(r["handling_fee"] or 0)
            m["consolidation_fee"] += float(r.get("consolidation_fee") or 0)
            m["letter_fee"] += float(r.get("letter_fee") or 0)
            m["total_revenue"] += float(r["total_fee"] or 0)
            m["customers"].add(r["g_code"])

            # 代理分潤累加：(rate - 180) × kg，最低 20 × kg
            if aid > 0 and kg > 0:
                rate = float(r["rate_per_kg"] or 0)
                per_kg = max(rate - 180, 20)
                m["commission"] += per_kg * kg

            try:
                extras = json.loads(r["extra_services"] or "[]")
                for e in extras:
                    m["extra_fee"] += int(e.get("subtotal") or e.get("qty", 1) * e.get("price", 0) or 0)
            except:
                pass

        result = []
        for key in sorted(buckets.keys(), reverse=True):
            m = buckets[key]
            m["customer_count"] = len(m["customers"])
            m["commission"] = round(m["commission"])
            del m["customers"]
            result.append(m)

        # 代理端：附上每週撥款狀態（讓代理在統計頁看得到「已撥款/後五碼」）
        if aid > 0 and result:
            conn2 = get_db()
            payouts = {
                p["period_key"]: dict(p)
                for p in conn2.execute("SELECT * FROM agent_payouts WHERE agent_id=?", (aid,)).fetchall()
            }
            conn2.close()
            for m in result:
                p = payouts.get(m["month"])
                m["paid"] = bool(p and p.get("paid_at"))
                m["payment_last5"] = (p or {}).get("payment_last5", "")
                m["paid_at"] = (p or {}).get("paid_at", "")

        return jsonify({
            "success": True,
            "period_type": period_type,
            "is_agent": aid > 0,
            "monthly": result
        })
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})


# ============ JPD 自動建單 ============

@app.route("/api/admin/shipment_requests/<int:req_id>/jpd_create", methods=["POST"])
def admin_create_jpd_order(req_id):
    """從出貨申請自動在 JPD 建立運單"""
    ok, _row = check_record_ownership("shipment_requests", req_id)
    if not _row:
        return jsonify({"success": False, "error": "找不到出貨申請"})
    if not ok:
        return jsonify({"success": False, "error": "權限不足"}), 403
    conn = get_db()
    req = conn.execute("SELECT * FROM shipment_requests WHERE id=?", (req_id,)).fetchone()
    req = dict(req)
    g_code = req.get("g_code", "")
    
    # 收件人資訊
    recipient = str(req.get("ship_recipient") or "")
    phone = str(req.get("ship_phone") or "")
    address = str(req.get("ship_address") or "")
    note = str(req.get("note") or "")
    
    if not recipient or not phone or not address:
        conn.close()
        return jsonify({"success": False, "error": "缺少寄送地址資訊"})
    
    # 從預報取得申報品項
    forecasts = conn.execute(
        "SELECT * FROM forecasts WHERE g_code=? AND status='待處理' ORDER BY id", (g_code,)
    ).fetchall()
    
    declare_list = []
    for fc in forecasts:
        fc = dict(fc)
        try:
            items = json.loads(fc.get("items_json") or "[]")
            for item in items:
                declare_list.append({
                    "product_name": item.get("name", "雑貨"),
                    "product_name_local": item.get("name", "雑貨"),
                    "product_num": int(item.get("quantity", 1)),
                    "product_price": int(float(item.get("price", 0))),
                    "product_url": item.get("url", "")
                })
        except Exception as e:
            print(f"[JPD] declare_list 解析失敗: {e}", flush=True)

    # 如果沒有預報品項，給一個預設品項
    if not declare_list:
        declare_list = [{
            "product_name": "雑貨",
            "product_name_local": "雑貨",
            "product_num": 1,
            "product_price": 0,
            "product_url": ""
        }]
    
    # 入庫包裹 ID 由管理員手動輸入（多個用逗號/頓號/空白分隔）
    body = request.get_json(silent=True) or {}
    raw_ids = body.get("package_ids", "")
    if isinstance(raw_ids, list):
        id_tokens = [str(x).strip() for x in raw_ids]
    else:
        id_tokens = re.split(r"[,，、\s]+", str(raw_ids))
    jpd_package_ids = []
    for tok in id_tokens:
        tok = tok.strip()
        if not tok:
            continue
        try:
            jpd_package_ids.append(int(tok))
        except (ValueError, TypeError):
            conn.close()
            return jsonify({"success": False, "error": f"包裹 ID「{tok}」格式錯誤，應為數字"})

    if not jpd_package_ids:
        conn.close()
        return jsonify({"success": False, "error": "請填入 JPD 入庫包裹 ID（可在 JPD 雲倉的入庫列表查到，多個用逗號分隔）"})

    # 客戶運單號前綴：客編 + 日期
    today_str = datetime.now().strftime("%m%d")
    order_prefix = f"{g_code}-{today_str}"

    # 方案 A：一個包裹建一張運單，運單號一律加流水號（-1, -2, -3...）
    created_orders = []
    failed_orders = []

    for idx, pid in enumerate(jpd_package_ids):
        customer_order_id = f"{order_prefix}-{idx + 1}"
        order_data = {
            "customer_order_id": customer_order_id,
            "deliv_id": JPD_DELIV_ID,
            "recipient": recipient,
            "id_issure": "",
            "area": 3,
            "addr1": address,
            "addr2": "",
            "addr3": "",
            "addr4": "",
            "tel": phone,
            "memo": note,
            "create_order_pdf": "n",
            "warehouse_id": JPD_WAREHOUSE_ID,
            "create_package": "n",
            "create_sender": "y",
            "packages": [{"package_id": pid, "declare_list": declare_list}]
        }
        print(f"[JPD] 建立運單: {customer_order_id}, 收件人: {recipient}, 包裹 id={pid}", flush=True)
        result = jpd_request("TCreateOrder", order_data)

        ok = False
        if "OperationResult" in result:
            op = result["OperationResult"]
            if op.get("Request", {}).get("IsValid") == "True":
                res_data = op.get("Result", {})
                if res_data.get("Result") == "SUCCESS":
                    data = res_data.get("Data", {})
                    jpd_order_id = str(data.get("order_id", ""))
                    jpd_logis_num = str(data.get("logis_num", ""))
                    print(f"[JPD] ✅ 建單成功: {customer_order_id} order_id={jpd_order_id}, logis_num={jpd_logis_num}", flush=True)
                    created_orders.append({
                        "customer_order_id": customer_order_id,
                        "jpd_order_id": jpd_order_id,
                        "jpd_logis_num": jpd_logis_num,
                        "package_id": pid
                    })
                    ok = True
                else:
                    err = res_data.get("ErrorMsg", str(res_data))
                    print(f"[JPD] ❌ 建單失敗 {customer_order_id}: {err}", flush=True)
                    failed_orders.append({"customer_order_id": customer_order_id, "error": str(err)})
            else:
                errs = op.get("Request", {}).get("Errors", {})
                print(f"[JPD] ❌ 請求無效 {customer_order_id}: {errs}", flush=True)
                failed_orders.append({"customer_order_id": customer_order_id, "error": str(errs)})
        if not ok and not failed_orders:
            failed_orders.append({"customer_order_id": customer_order_id, "error": "JPD API 無回應"})

    conn.close()

    if created_orders and not failed_orders:
        nums = ', '.join([o["jpd_logis_num"] or o["customer_order_id"] for o in created_orders])
        return jsonify({
            "success": True,
            "created": created_orders,
            "message": f"已建立 {len(created_orders)} 張 JPD 運單：{nums}"
        })
    elif created_orders and failed_orders:
        return jsonify({
            "success": True,
            "created": created_orders,
            "failed": failed_orders,
            "message": f"成功 {len(created_orders)} 張、失敗 {len(failed_orders)} 張。失敗：" +
                       '、'.join([f["customer_order_id"] + '(' + f["error"] + ')' for f in failed_orders])
        })
    else:
        first_err = failed_orders[0]["error"] if failed_orders else "未知錯誤"
        return jsonify({"success": False, "error": f"JPD 建單全部失敗：{first_err}"})


# ============ 公告 API ============

@app.route("/api/announcements", methods=["GET"])
def get_announcements():
    """取得啟用中的公告"""
    conn = get_db()
    rows = conn.execute(
        "SELECT * FROM announcements WHERE is_active=1 ORDER BY id DESC"
    ).fetchall()
    conn.close()
    return jsonify({"success": True, "announcements": [dict(r) for r in rows]})


@app.route("/api/admin/announcements", methods=["GET"])
def admin_get_announcements():
    """管理員取得所有公告"""
    conn = get_db()
    rows = conn.execute("SELECT * FROM announcements ORDER BY id DESC").fetchall()
    conn.close()
    return jsonify({"success": True, "announcements": [dict(r) for r in rows]})


@app.route("/api/admin/announcements", methods=["POST"])
def admin_create_announcement():
    """管理員新增公告（僅主管理員）"""
    if not is_super_admin():
        return jsonify({"success": False, "error": "權限不足"}), 403
    data = request.json
    title = (data.get("title") or "").strip()
    content = (data.get("content") or "").strip()
    if not title or not content:
        return jsonify({"success": False, "error": "標題和內容為必填"})
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    conn = get_db()
    cur = conn.execute(
        "INSERT INTO announcements (title, content, is_active, created_at) VALUES (?, ?, 1, ?)",
        (title, content, now)
    )
    conn.commit()
    conn.close()
    return jsonify({"success": True, "id": cur.lastrowid})


@app.route("/api/admin/announcements/<int:ann_id>", methods=["PUT"])
def admin_update_announcement(ann_id):
    """管理員更新公告（僅主管理員）"""
    if not is_super_admin():
        return jsonify({"success": False, "error": "權限不足"}), 403
    data = request.json
    conn = get_db()
    fields = {}
    for key in ["title", "content"]:
        if key in data:
            fields[key] = (data[key] or "").strip()
    if "is_active" in data:
        fields["is_active"] = 1 if data["is_active"] else 0
    if fields:
        sets = ", ".join(f"{k}=?" for k in fields)
        vals = list(fields.values()) + [ann_id]
        conn.execute(f"UPDATE announcements SET {sets} WHERE id=?", vals)
    conn.commit()
    conn.close()
    return jsonify({"success": True})


@app.route("/api/admin/announcements/<int:ann_id>", methods=["DELETE"])
def admin_delete_announcement(ann_id):
    """管理員刪除公告（僅主管理員）"""
    if not is_super_admin():
        return jsonify({"success": False, "error": "權限不足"}), 403
    conn = get_db()
    conn.execute("DELETE FROM announcements WHERE id=?", (ann_id,))
    conn.commit()
    conn.close()
    return jsonify({"success": True})


# ============ 地址簿 API ============

@app.route("/api/addresses", methods=["GET"])
def get_addresses():
    """取得客戶地址簿"""
    g_code = request.args.get("g_code", "").upper()
    if not g_code:
        return jsonify({"success": False, "error": "缺少會員編號"})
    conn = get_db()
    rows = conn.execute(
        "SELECT * FROM addresses WHERE g_code=? ORDER BY is_default DESC, id DESC", (g_code,)
    ).fetchall()
    conn.close()
    return jsonify({"success": True, "addresses": [dict(r) for r in rows]})


@app.route("/api/addresses", methods=["POST"])
def add_address():
    """新增地址"""
    data = request.json
    g_code = (data.get("g_code") or "").strip().upper()
    recipient = (data.get("recipient") or "").strip()
    phone = (data.get("phone") or "").strip()
    address = (data.get("address") or "").strip()
    label = (data.get("label") or "").strip()
    zipcode = (data.get("zipcode") or "").strip()
    is_default = 1 if data.get("is_default") else 0

    if not g_code or not recipient or not phone or not address:
        return jsonify({"success": False, "error": "收件人、電話、地址為必填"})

    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    conn = get_db()
    # 如果設為預設，先清除其他預設
    if is_default:
        conn.execute("UPDATE addresses SET is_default=0 WHERE g_code=?", (g_code,))
    # 如果是第一筆，自動設為預設
    count = conn.execute("SELECT COUNT(*) as c FROM addresses WHERE g_code=?", (g_code,)).fetchone()["c"]
    if count == 0:
        is_default = 1

    conn.execute(
        """INSERT INTO addresses (g_code, label, recipient, phone, zipcode, address, is_default, created_at)
           VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
        (g_code, label, recipient, phone, zipcode, address, is_default, now)
    )
    conn.commit()
    conn.close()
    return jsonify({"success": True, "message": "地址已新增"})


@app.route("/api/addresses/<int:addr_id>", methods=["PUT"])
def update_address(addr_id):
    """更新地址"""
    data = request.json
    g_code = (data.get("g_code") or "").strip().upper()
    conn = get_db()
    # 驗證是本人的
    row = conn.execute("SELECT g_code FROM addresses WHERE id=?", (addr_id,)).fetchone()
    if not row or row["g_code"] != g_code:
        conn.close()
        return jsonify({"success": False, "error": "找不到該地址"})

    fields = {}
    for key in ["label", "recipient", "phone", "zipcode", "address"]:
        if key in data:
            fields[key] = (data[key] or "").strip()
    if "is_default" in data and data["is_default"]:
        conn.execute("UPDATE addresses SET is_default=0 WHERE g_code=?", (g_code,))
        fields["is_default"] = 1

    if fields:
        sets = ", ".join(f"{k}=?" for k in fields)
        vals = list(fields.values()) + [addr_id]
        conn.execute(f"UPDATE addresses SET {sets} WHERE id=?", vals)
    conn.commit()
    conn.close()
    return jsonify({"success": True})


@app.route("/api/addresses/<int:addr_id>", methods=["DELETE"])
def delete_address(addr_id):
    """刪除地址"""
    data = request.json or {}
    g_code = (data.get("g_code") or request.args.get("g_code", "")).strip().upper()
    conn = get_db()
    row = conn.execute("SELECT g_code, is_default FROM addresses WHERE id=?", (addr_id,)).fetchone()
    if not row or row["g_code"] != g_code:
        conn.close()
        return jsonify({"success": False, "error": "找不到該地址"})
    conn.execute("DELETE FROM addresses WHERE id=?", (addr_id,))
    # 如果刪的是預設，把第一筆設為預設
    if row["is_default"]:
        first = conn.execute("SELECT id FROM addresses WHERE g_code=? ORDER BY id LIMIT 1", (g_code,)).fetchone()
        if first:
            conn.execute("UPDATE addresses SET is_default=1 WHERE id=?", (first["id"],))
    conn.commit()
    conn.close()
    return jsonify({"success": True})


@app.route("/api/addresses/<int:addr_id>/default", methods=["POST"])
def set_default_address(addr_id):
    """設為預設地址"""
    data = request.json
    g_code = (data.get("g_code") or "").strip().upper()
    conn = get_db()
    row = conn.execute("SELECT g_code FROM addresses WHERE id=?", (addr_id,)).fetchone()
    if not row or row["g_code"] != g_code:
        conn.close()
        return jsonify({"success": False, "error": "找不到該地址"})
    conn.execute("UPDATE addresses SET is_default=0 WHERE g_code=?", (g_code,))
    conn.execute("UPDATE addresses SET is_default=1 WHERE id=?", (addr_id,))
    conn.commit()
    conn.close()
    return jsonify({"success": True})


# ============ 加值服務目錄 API ============

@app.route("/api/extra_services/catalog", methods=["GET"])
def public_extra_service_catalog():
    """客戶端用：
      • 預設（出貨申請可勾清單）→ 只回 sel=True 固定價項目
      • ?full=1（費用價目表）→ 回全部項目（含 sel 旗標）
    無需登入。"""
    cat = get_extra_service_catalog()
    full = request.args.get("full") in ("1", "true", "yes")
    src = cat if full else [c for c in cat if c.get("sel")]
    items = [
        {"id": c.get("id"), "name": c.get("name", ""), "cat": c.get("cat", ""),
         "desc": c.get("desc", ""), "price": int(c.get("price") or 0), "sel": bool(c.get("sel"))}
        for c in src
    ]
    return jsonify({"success": True, "services": items})


@app.route("/api/admin/extra_services/catalog", methods=["GET"])
def admin_get_extra_service_catalog():
    """後台管理用：回傳完整目錄（含 sel 旗標與變動價項目）。"""
    if not is_super_admin():
        return jsonify({"success": False, "error": "僅主管理員可管理加值服務目錄"}), 403
    return jsonify({"success": True, "services": get_extra_service_catalog()})


@app.route("/api/admin/extra_services/catalog", methods=["POST"])
def admin_save_extra_service_catalog():
    """後台管理用：整批覆寫目錄。"""
    if not is_super_admin():
        return jsonify({"success": False, "error": "僅主管理員可管理加值服務目錄"}), 403
    data = request.json or {}
    raw = data.get("services", [])
    if not isinstance(raw, list):
        return jsonify({"success": False, "error": "資料格式錯誤"}), 400
    cleaned = []
    # 先蒐集所有明確 id（避免自動配號撞到後面才出現的明確 id）
    explicit_ids = set()
    for c in raw:
        if isinstance(c, dict):
            cid = (c.get("id") or "").strip()
            if cid:
                explicit_ids.add(cid)
    used_ids = set()
    _counter = [0]
    def _fresh_id():
        while True:
            _counter[0] += 1
            cand = f"es{_counter[0]:02d}"
            if cand not in explicit_ids and cand not in used_ids:
                return cand
    for c in raw:
        if not isinstance(c, dict):
            continue
        name = (c.get("name") or "").strip()
        if not name:
            continue
        try:
            price = int(float(c.get("price") or 0))
        except (ValueError, TypeError):
            price = 0
        cid = (c.get("id") or "").strip()
        if not cid or cid in used_ids:   # 空的或重複 → 配一個不衝突的新 id
            cid = _fresh_id()
        used_ids.add(cid)
        cleaned.append({
            "id": cid,
            "name": name,
            "cat": (c.get("cat") or "").strip(),
            "desc": (c.get("desc") or "").strip(),
            "price": max(price, 0),
            "sel": bool(c.get("sel")),
        })
    conn = get_db()
    conn.execute(
        "INSERT INTO admin_settings (key, value) VALUES ('extra_service_catalog', ?) "
        "ON CONFLICT(key) DO UPDATE SET value=excluded.value",
        (json.dumps(cleaned, ensure_ascii=False),)
    )
    conn.commit()
    conn.close()
    return jsonify({"success": True, "count": len(cleaned)})


# ============ 台灣配送貨況 API ============

@app.route("/api/admin/tracking/status", methods=["GET"])
def admin_tracking_status():
    if not is_super_admin():
        return jsonify({"success": False, "error": "僅主管理員可操作"}), 403
    conn = get_db()
    total = conn.execute("SELECT COUNT(*) AS c FROM delivery_tracking").fetchone()["c"]
    conn.close()
    return jsonify({
        "success": True,
        "last_sync": _get_setting("tracking_last_sync", ""),
        "sheet_url": _get_setting("tracking_sheet_url", DEFAULT_TRACKING_SHEET_URL),
        "total": total,
    })


@app.route("/api/admin/tracking/sync", methods=["POST"])
def admin_tracking_sync():
    if not is_super_admin():
        return jsonify({"success": False, "error": "僅主管理員可操作"}), 403
    data = request.json or {}
    new_url = (data.get("sheet_url") or "").strip()
    if new_url:
        _set_setting("tracking_sheet_url", new_url)
    try:
        count = sync_delivery_tracking()
        return jsonify({"success": True, "count": count, "last_sync": _get_setting("tracking_last_sync", "")})
    except Exception as e:
        return jsonify({"success": False, "error": f"同步失敗：{e}"}), 500


# ============ 出貨申請 API ============

@app.route("/api/shipment_request", methods=["POST"])
def create_shipment_request():
    """客戶申請出貨"""
    data = request.json
    g_code = (data.get("g_code") or "").strip().upper()
    customer_name = data.get("customer_name", "")
    package_ids = data.get("package_ids", [])
    note = (data.get("note") or "").strip()
    # 收件地址
    ship_recipient = (data.get("ship_recipient") or "").strip()
    ship_phone = (data.get("ship_phone") or "").strip()
    ship_address = (data.get("ship_address") or "").strip()
    # 地址完整性守門：缺縣市/區的地址黑貓無法投遞。不完整且客戶未確認 → 擋下請補
    if ship_address and not tw_zip.is_address_complete(ship_address) and not data.get("address_confirmed"):
        return jsonify({
            "success": False,
            "need_address_confirm": True,
            "error": "收件地址似乎缺少縣市或區（例：嘉義市西區），台灣宅配可能無法投遞。請確認或補齊地址。"
        }), 400
    # 客戶勾選的加值服務（[{id, qty}]）→ 以伺服端目錄價驗證後存入（防前端竄改價格）
    sel_services = data.get("extra_services", []) or []

    if not g_code:
        return jsonify({"success": False, "error": "缺少會員編號"})
    if not package_ids:
        return jsonify({"success": False, "error": "請選擇要出貨的包裹"})
    if not ship_recipient or not ship_phone or not ship_address:
        return jsonify({"success": False, "error": "請選擇寄送地址"})

    # 組合包裹摘要
    conn = get_db()

    # 重複申請檢查：阻擋同一包裹同時出現在多筆「進行中」的出貨申請
    # 進行中 = 待處理 / 處理中（已出貨會把 packages.status 更新為 已出貨，不會被選到；已拒絕視為釋出）
    requested_set = set()
    for p in package_ids:
        try:
            requested_set.add(int(p))
        except (ValueError, TypeError):
            pass

    active_reqs = conn.execute(
        "SELECT id, package_ids FROM shipment_requests "
        "WHERE g_code=? AND status IN ('待處理', '處理中')",
        (g_code,)
    ).fetchall()
    already_pending = set()
    for r in active_reqs:
        ids_str = r["package_ids"] or ""
        for pid_str in ids_str.split(","):
            try:
                already_pending.add(int(pid_str.strip()))
            except (ValueError, AttributeError):
                pass

    duplicates = requested_set & already_pending
    if duplicates:
        # 把重複的包裹 id 對應到 product_name 顯示得更友善
        dup_rows = conn.execute(
            f"SELECT id, product_name, logis_num FROM packages "
            f"WHERE id IN ({','.join(['?']*len(duplicates))})",
            list(duplicates)
        ).fetchall()
        dup_names = []
        for d in dup_rows:
            name = d["product_name"] or "未命名"
            logis = d["logis_num"] or ""
            dup_names.append(f"{name}（末四碼 {logis}）" if logis and logis != "-" else name)
        conn.close()
        return jsonify({
            "success": False,
            "error": f"以下包裹已在進行中的出貨申請中，無法重複申請：\n• " + "\n• ".join(dup_names) +
                     "\n\n請等管理員處理完成後再申請新出貨，或聯繫客服取消舊申請。"
        })

    placeholders = ",".join(["?"] * len(package_ids))
    rows = conn.execute(
        f"SELECT id, logis_num, product_name, weight FROM packages WHERE id IN ({placeholders}) AND g_code=?",
        package_ids + [g_code]
    ).fetchall()

    if not rows:
        conn.close()
        return jsonify({"success": False, "error": "找不到對應的包裹"})

    summary_parts = []
    total_weight = 0
    for idx, r in enumerate(rows, 1):
        r = dict(r)
        name = r["product_name"] or "商品"
        logis = r["logis_num"] or ""
        w = r["weight"] or ""
        line = f"{idx}. {name}"
        if w:
            line += f" / {w} kg"
        if logis and logis != "-":
            line += f" / {logis}"
        summary_parts.append(line)
        try:
            total_weight += float(w) if w else 0
        except:
            pass
    summary = "\n".join(summary_parts)
    if total_weight > 0:
        summary += f"\n合計約 {total_weight:.1f} kg"

    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ids_str = ",".join(str(i) for i in package_ids)
    sr_agent_id = get_agent_id_for_g_code(g_code)

    # 依伺服端目錄驗證客戶勾選：只收 sel=True 的項目、價格一律用目錄價、數量下限 1
    catalog = {c["id"]: c for c in get_extra_service_catalog() if c.get("sel")}
    customer_extras = []
    for s in sel_services:
        if not isinstance(s, dict):
            continue
        c = catalog.get(s.get("id"))
        if not c:
            continue
        try:
            qty = int(s.get("qty", 1))
        except (ValueError, TypeError):
            qty = 1
        if qty < 1:
            qty = 1
        price = int(c.get("price") or 0)
        # 「合箱」是意願選項：實際合箱費由系統依箱數計算（consolidation_fee），
        # 客戶勾選一律 0 元，避免與系統合箱費重複收費
        if c["name"] == "合箱":
            price = 0
        customer_extras.append({
            "id": c["id"], "name": c["name"], "qty": qty,
            "price": price, "subtotal": price * qty,
            "src": "customer",  # 客戶申請（管理員請款時可增刪，最終以帳單為準）
        })
    extra_services_json = json.dumps(customer_extras, ensure_ascii=False)

    conn.execute(
        """INSERT INTO shipment_requests (g_code, customer_name, package_ids, package_summary, status, note, ship_recipient, ship_phone, ship_address, extra_services, created_at, agent_id)
           VALUES (?, ?, ?, ?, '待處理', ?, ?, ?, ?, ?, ?, ?)""",
        (g_code, customer_name, ids_str, summary, note, ship_recipient, ship_phone, ship_address, extra_services_json, now, sr_agent_id)
    )
    conn.commit()
    conn.close()

    return jsonify({"success": True, "message": "出貨申請已送出，管理員會盡快處理！"})


@app.route("/api/shipment_requests", methods=["GET"])
def get_my_shipment_requests():
    """客戶查看自己的出貨申請"""
    g_code = request.args.get("g_code", "").upper()
    if not g_code:
        return jsonify({"success": False, "error": "缺少會員編號"})
    conn = get_db()
    rows = conn.execute(
        "SELECT * FROM shipment_requests WHERE g_code=? ORDER BY id DESC", (g_code,)
    ).fetchall()

    # 台灣配送貨況：每筆出貨單用「存的 export_code ＋ 現算 {g_code}-{MMDD}」多候選比對，
    # 讓 export_code 上線前的舊單也能對到（MMDD 取 updated_at＝標記已出貨那天，其次 created_at）
    def _mmdd(v):
        try:
            return datetime.strptime(str(v)[:10], "%Y-%m-%d").strftime("%m%d")
        except (ValueError, TypeError):
            return None

    req_candidates = {}
    all_codes = set()
    for r in rows:
        cands, seen = [], set()
        for c in ([r["export_code"]] +
                  [f"{r['g_code']}-{mm}" for mm in (_mmdd(r["updated_at"]), _mmdd(r["created_at"])) if mm]):
            if c and c not in seen:
                seen.add(c); cands.append(c)
        req_candidates[r["id"]] = cands
        all_codes.update(cands)

    tmap = {}
    if all_codes:
        codes_list = list(all_codes)
        ph = ",".join(["?"] * len(codes_list))
        for t in conn.execute(
            f"SELECT customer_code, carrier, tracking_num FROM delivery_tracking WHERE customer_code IN ({ph})",
            codes_list
        ).fetchall():
            tmap[t["customer_code"]] = t
    conn.close()

    result = []
    for r in rows:
        d = dict(r)
        for c in req_candidates.get(r["id"], []):
            t = tmap.get(c)
            if t and t["tracking_num"]:
                d["delivery_carrier"] = t["carrier"]
                d["delivery_tracking"] = t["tracking_num"]
                d["delivery_url"] = delivery_tracking_url(t["carrier"], t["tracking_num"])
                break
        result.append(d)
    return jsonify({"success": True, "requests": result})


@app.route("/api/shipment_requests/<int:req_id>/payment", methods=["POST"])
def submit_payment_info(req_id):
    """客戶回報匯款後五碼"""
    data = request.json
    last5 = (data.get("last5") or "").strip()
    g_code = (data.get("g_code") or "").strip().upper()

    if not last5 or len(last5) != 5:
        return jsonify({"success": False, "error": "請輸入帳號後五碼（5位數字）"})
    if not last5.isdigit():
        return jsonify({"success": False, "error": "請輸入數字"})

    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    conn = get_db()
    # 確認是該客戶的申請
    row = conn.execute("SELECT g_code FROM shipment_requests WHERE id=?", (req_id,)).fetchone()
    if not row or row["g_code"] != g_code:
        conn.close()
        return jsonify({"success": False, "error": "找不到該申請"})

    conn.execute(
        "UPDATE shipment_requests SET payment_last5=?, payment_at=? WHERE id=?",
        (last5, now, req_id)
    )
    conn.commit()
    conn.close()
    return jsonify({"success": True, "message": "匯款回報成功！"})


@app.route("/api/admin/old_packages", methods=["GET"])
def admin_old_packages():
    """查詢倉庫滯留超過 N 天的未出貨包裹（預設 30 天）"""
    try:
        days = int(request.args.get("days", 30))
    except (ValueError, TypeError):
        days = 30
    cutoff_date = (datetime.now() - timedelta(days=days)).strftime("%Y-%m-%d")
    aid = get_current_agent_id()
    conn = get_db()
    if aid > 0:
        rows = conn.execute(
            """SELECT * FROM packages
               WHERE status != '已出貨'
                 AND agent_id = ?
                 AND COALESCE(NULLIF(in_date, ''), substr(created_at, 1, 10)) <= ?
               ORDER BY COALESCE(NULLIF(in_date, ''), substr(created_at, 1, 10)) ASC""",
            (aid, cutoff_date)
        ).fetchall()
    else:
        rows = conn.execute(
            """SELECT * FROM packages
               WHERE status != '已出貨'
                 AND COALESCE(NULLIF(in_date, ''), substr(created_at, 1, 10)) <= ?
               ORDER BY COALESCE(NULLIF(in_date, ''), substr(created_at, 1, 10)) ASC""",
            (cutoff_date,)
        ).fetchall()
    conn.close()
    today = datetime.now().date()
    result = []
    for r in rows:
        d = dict(r)
        ref_date_str = d.get("in_date") or (d.get("created_at") or "")[:10]
        try:
            ref_date = datetime.strptime(ref_date_str, "%Y-%m-%d").date()
            age_days = (today - ref_date).days
        except (ValueError, TypeError):
            age_days = 0
        d["age_days"] = age_days
        d["ref_date"] = ref_date_str
        result.append(d)
    return jsonify({
        "success": True,
        "count": len(result),
        "days": days,
        "packages": result
    })


@app.route("/api/admin/customer_unpaid/<g_code>", methods=["GET"])
def admin_customer_unpaid(g_code):
    """查詢客戶未付款的已出貨筆數和金額（用於出貨警告）"""
    # 代理只能查自己的客戶
    aid = get_current_agent_id()
    if aid > 0 and get_agent_id_for_g_code(g_code.upper()) != aid:
        return jsonify({"success": True, "count": 0, "total": 0, "latest": "", "ids": []})
    conn = get_db()
    rows = conn.execute(
        "SELECT id, total_fee, updated_at FROM shipment_requests "
        "WHERE g_code=? AND status='已出貨' AND total_fee > 0 "
        "AND (payment_last5 IS NULL OR payment_last5='') "
        "ORDER BY updated_at DESC",
        (g_code.upper(),)
    ).fetchall()
    conn.close()
    items = [dict(r) for r in rows]
    return jsonify({
        "success": True,
        "count": len(items),
        "total": sum(int(r.get("total_fee") or 0) for r in items),
        "latest": items[0]["updated_at"] if items else "",
        "ids": [r["id"] for r in items]
    })


@app.route("/api/admin/shipment_requests/<int:req_id>/confirm_payment", methods=["POST"])
def admin_confirm_payment(req_id):
    """管理員確認匯款已收到（可填後五碼、LINE Pay、現金等任意備註）"""
    ok, _row = check_record_ownership("shipment_requests", req_id)
    if not _row:
        return jsonify({"success": False, "error": "找不到該申請"})
    if not ok:
        return jsonify({"success": False, "error": "權限不足"}), 403
    data = request.json or {}
    note = (data.get("last5") or "").strip()

    if len(note) > 20:
        return jsonify({"success": False, "error": "備註請勿超過 20 字"})
    if not note:
        note = "管確認"

    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    conn = get_db()
    conn.execute(
        "UPDATE shipment_requests SET payment_last5=?, payment_at=? WHERE id=?",
        (note, now, req_id)
    )
    conn.commit()
    conn.close()
    return jsonify({"success": True, "message": "已確認匯款"})


@app.route("/api/admin/shipment_requests/<int:req_id>/unconfirm_payment", methods=["POST"])
def admin_unconfirm_payment(req_id):
    """管理員取消已確認的匯款（誤按時用）"""
    ok, _row = check_record_ownership("shipment_requests", req_id)
    if not _row:
        return jsonify({"success": False, "error": "找不到該申請"})
    if not ok:
        return jsonify({"success": False, "error": "權限不足"}), 403
    conn = get_db()
    conn.execute(
        "UPDATE shipment_requests SET payment_last5='', payment_at='' WHERE id=?",
        (req_id,)
    )
    conn.commit()
    conn.close()
    return jsonify({"success": True, "message": "已取消匯款確認"})


# ============ 出檔案給廠商（Nigel / JpD ） ============

@app.route("/api/admin/vendors", methods=["GET"])
def admin_list_vendors():
    """前端 UI 廠商下拉選單用"""
    if not is_super_admin():
        return jsonify({"success": False, "error": "權限不足"}), 403
    try:
        return jsonify({"success": True, "vendors": vendor_templates.list_vendors()})
    except Exception as e:
        import traceback
        print(f"[vendors] 💥 例外:\n{traceback.format_exc()}", flush=True)
        return jsonify({"success": False, "error": f"{type(e).__name__}: {e}"}), 500


@app.route("/api/admin/customer_vendor_codes", methods=["GET"])
def admin_get_vendor_codes():
    """查詢一批客戶的廠商編號（FWT0001 等）"""
    if not is_super_admin():
        return jsonify({"success": False, "error": "權限不足"}), 403
    g_codes_str = request.args.get("g_codes", "")
    g_codes = [s.strip().upper() for s in g_codes_str.split(",") if s.strip()]
    conn = get_db()
    if g_codes:
        placeholders = ",".join(["?"] * len(g_codes))
        rows = conn.execute(
            f"SELECT g_code, vendor, code FROM customer_vendor_codes WHERE g_code IN ({placeholders})",
            g_codes
        ).fetchall()
    else:
        rows = conn.execute(
            "SELECT g_code, vendor, code FROM customer_vendor_codes ORDER BY g_code, vendor"
        ).fetchall()
    conn.close()
    # 回傳 nested dict: {g_code: {vendor: code}}
    result = {}
    for r in rows:
        result.setdefault(r["g_code"], {})[r["vendor"]] = r["code"]
    return jsonify({"success": True, "codes": result})


@app.route("/api/admin/customer_vendor_codes", methods=["POST"])
def admin_set_vendor_code():
    """設定／更新某客戶的廠商編號"""
    if not is_super_admin():
        return jsonify({"success": False, "error": "權限不足"}), 403
    data = request.json or {}
    g_code = (data.get("g_code") or "").strip().upper()
    vendor = (data.get("vendor") or "").strip().lower()
    code = (data.get("code") or "").strip()
    if not g_code or not vendor:
        return jsonify({"success": False, "error": "缺少 g_code 或 vendor"})
    if vendor not in vendor_templates.VENDORS:
        return jsonify({"success": False, "error": f"未知廠商：{vendor}"})

    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    conn = get_db()
    if code:
        # INSERT OR REPLACE
        conn.execute(
            "INSERT INTO customer_vendor_codes (g_code, vendor, code, updated_at) VALUES (?,?,?,?) "
            "ON CONFLICT(g_code, vendor) DO UPDATE SET code=excluded.code, updated_at=excluded.updated_at",
            (g_code, vendor, code, now)
        )
    else:
        # 空字串 = 刪除
        conn.execute(
            "DELETE FROM customer_vendor_codes WHERE g_code=? AND vendor=?",
            (g_code, vendor)
        )
    conn.commit()
    conn.close()
    return jsonify({"success": True, "message": "已儲存"})


@app.route("/api/admin/exports/pending", methods=["GET"])
def admin_exports_pending():
    """列出已付款但未匯出給廠商的出貨單"""
    if not is_super_admin():
        return jsonify({"success": False, "error": "權限不足"}), 403
    try:
        return _admin_exports_pending_impl()
    except Exception as e:
        import traceback
        tb = traceback.format_exc()
        print(f"[exports/pending] 💥 例外:\n{tb}", flush=True)
        # 回 JSON 而不是 HTML 500，讓前端能解析
        return jsonify({
            "success": False,
            "error": f"後端錯誤: {type(e).__name__}: {e}",
            "traceback_excerpt": tb.splitlines()[-1] if tb else "",
        }), 500


def _admin_exports_pending_impl():
    conn = get_db()
    rows = conn.execute("""
        SELECT * FROM shipment_requests
        WHERE status='已出貨'
          AND payment_last5 IS NOT NULL AND payment_last5 != ''
          AND (exported_at IS NULL OR exported_at = '')
        ORDER BY payment_at ASC, id ASC
    """).fetchall()

    # 撈所有相關 packages
    pkg_ids = set()
    for r in rows:
        pkg_ids.update(_parse_pkg_ids(r["package_ids"]))

    pkg_map = {}
    if pkg_ids:
        placeholders = ",".join(["?"] * len(pkg_ids))
        pkg_rows = conn.execute(
            f"SELECT id, g_code, logis_num, product_name, weight FROM packages WHERE id IN ({placeholders})",
            list(pkg_ids)
        ).fetchall()
        for p in pkg_rows:
            pkg_map[p["id"]] = dict(p)

    # 撈 vendor codes
    g_codes = list({r["g_code"] for r in rows})
    codes_map = {}
    if g_codes:
        placeholders = ",".join(["?"] * len(g_codes))
        code_rows = conn.execute(
            f"SELECT g_code, vendor, code FROM customer_vendor_codes WHERE g_code IN ({placeholders})",
            g_codes
        ).fetchall()
        for c in code_rows:
            codes_map.setdefault(c["g_code"], {})[c["vendor"]] = c["code"]

    # Fallback 資料來源（與 /generate 一致）：members 表 + Shopify cache
    members_map = {}
    if g_codes:
        placeholders = ",".join(["?"] * len(g_codes))
        for m in conn.execute(
            f"SELECT g_code, name, phone, address FROM members WHERE g_code IN ({placeholders})",
            g_codes
        ).fetchall():
            members_map[m["g_code"]] = {"name": m["name"], "phone": m["phone"], "address": m["address"]}
    # ★ 重要：用「目前已快取」的 Shopify 客戶，不觸發新抓取（避免冷啟動阻塞）
    # 如果快取為空 → shopify_map 空 → fallback 只用 members 表 + customer_name
    shopify_map = {}
    try:
        cached = _customers_cache.get("data") or []
        for c in cached:
            if c.get("g_code") in g_codes:
                shopify_map[c["g_code"]] = {"name": c.get("name", ""), "phone": c.get("phone", ""), "address": c.get("address", "")}
        # 順手在背景觸發更新（如果過期、不會阻塞請求）
        if cached and (time.time() - _customers_cache.get("time", 0)) >= CACHE_TTL:
            with _cache_lock:
                global _refresh_thread
                if _refresh_thread is None or not _refresh_thread.is_alive():
                    _refresh_thread = threading.Thread(
                        target=_refresh_shopify_async, daemon=True, name="ShopifyRefresh"
                    )
                    _refresh_thread.start()
    except Exception as e:
        print(f"[export-pending] Shopify cache 讀取失敗（不致命）: {e}", flush=True)
    conn.close()

    items = []
    for r in rows:
        rd = dict(r)
        pids = _parse_pkg_ids(rd.get("package_ids"))
        # ship_* 為空時用 fallback 在 UI 也能看到正確資料
        ship_recipient = _safe_str(rd.get("ship_recipient"))
        ship_phone     = _safe_str(rd.get("ship_phone"))
        ship_address   = _safe_str(rd.get("ship_address"))
        if not (ship_recipient and ship_phone and ship_address):
            fb = members_map.get(rd["g_code"]) or shopify_map.get(rd["g_code"]) or {}
            if not ship_recipient: ship_recipient = _safe_str(fb.get("name")) or _safe_str(rd.get("customer_name"))
            if not ship_phone:     ship_phone     = _safe_str(fb.get("phone"))
            if not ship_address:   ship_address   = _safe_str(fb.get("address"))

        items.append({
            "id":               rd["id"],
            "g_code":           rd["g_code"],
            "customer_name":    rd.get("customer_name") or "",
            "ship_recipient":   ship_recipient,
            "ship_phone":       ship_phone,
            "ship_address":     ship_address,
            "billed_weight":    rd.get("billed_weight") or 0,
            "total_fee":        rd.get("total_fee") or 0,
            "payment_at":       rd.get("payment_at") or "",
            "payment_last5":    rd.get("payment_last5") or "",
            "updated_at":       rd.get("updated_at") or "",
            "package_count":    len(pids),
            "packages":         [pkg_map[i] for i in pids if i in pkg_map],
            "vendor_codes":     codes_map.get(rd["g_code"], {}),
        })
    return jsonify({"success": True, "items": items})


@app.route("/api/admin/exports/generate", methods=["POST"])
def admin_exports_generate():
    """匯出選定的出貨單為廠商 Excel，並標記 exported_*"""
    if not is_super_admin():
        return jsonify({"success": False, "error": "權限不足"}), 403
    try:
        return _admin_exports_generate_impl()
    except Exception as e:
        import traceback
        tb = traceback.format_exc()
        print(f"[exports/generate] 💥 例外:\n{tb}", flush=True)
        return jsonify({
            "success": False,
            "error": f"後端錯誤: {type(e).__name__}: {e}",
            "traceback_excerpt": tb.splitlines()[-1] if tb else "",
        }), 500


def _admin_exports_generate_impl():
    data = request.json or {}
    vendor_id = (data.get("vendor") or "").strip().lower()
    ids = data.get("ids") or []

    vendor = vendor_templates.get_vendor(vendor_id)
    if not vendor:
        return jsonify({"success": False, "error": f"未知廠商：{vendor_id}"}), 400
    if not ids:
        return jsonify({"success": False, "error": "請至少選一筆"}), 400
    ids = [int(x) for x in ids if str(x).isdigit()]
    if not ids:
        return jsonify({"success": False, "error": "無有效 ID"}), 400

    conn = get_db()
    placeholders = ",".join(["?"] * len(ids))
    rows = conn.execute(
        f"""SELECT * FROM shipment_requests
            WHERE id IN ({placeholders})
              AND status='已出貨'
              AND payment_last5 IS NOT NULL AND payment_last5 != ''
              AND (exported_at IS NULL OR exported_at = '')
        """,
        ids
    ).fetchall()
    if not rows:
        conn.close()
        return jsonify({"success": False, "error": "選定的單都已匯出或狀態不符（可能被別人剛剛搶先匯出了）"}), 400

    # 撈所有相關 packages
    all_pkg_ids = set()
    for r in rows:
        all_pkg_ids.update(_parse_pkg_ids(r["package_ids"]))
    pkg_map = {}
    if all_pkg_ids:
        ph = ",".join(["?"] * len(all_pkg_ids))
        for p in conn.execute(
            f"SELECT id, g_code, logis_num, product_name, weight FROM packages WHERE id IN ({ph})",
            list(all_pkg_ids)
        ).fetchall():
            pkg_map[p["id"]] = dict(p)

    # 撈會員資料做 fallback（舊出貨單 ship_* 欄位可能空白）
    g_codes_needed = list({r["g_code"] for r in rows})
    members_map = {}  # g_code → {name, phone, address}
    if g_codes_needed:
        ph = ",".join(["?"] * len(g_codes_needed))
        for m in conn.execute(
            f"SELECT g_code, name, phone, address FROM members WHERE g_code IN ({ph})",
            g_codes_needed
        ).fetchall():
            members_map[m["g_code"]] = {"name": m["name"], "phone": m["phone"], "address": m["address"]}
    # Shopify 客戶 fallback：從「目前已快取」撈，不觸發新抓取（避免冷啟動阻塞）
    shopify_map = {}
    try:
        cached = _customers_cache.get("data") or []
        for c in cached:
            if c.get("g_code") in g_codes_needed:
                shopify_map[c["g_code"]] = {"name": c.get("name", ""), "phone": c.get("phone", ""), "address": c.get("address", "")}
    except Exception as ex:
        print(f"[export] Shopify fallback 失敗（不致命）: {ex}", flush=True)

    # 組 shipments list
    shipments = []
    fallback_updates = []  # (id, ship_recipient, ship_phone, ship_address) - 把 fallback 後的值寫回 DB
    missing_pkg_count = 0
    for r in rows:
        rd = dict(r)
        pids = _parse_pkg_ids(rd.get("package_ids"))
        # 包裹資料：找到的用真實值、找不到的用 stub
        # vendor 範本實際上只用 package_id 當隨機種子，不用 logis_num/weight 等具體欄位
        # 所以孤立資料（migration 後 packages 表沒對應）也能撐過 Excel 產出
        pkgs = []
        for i in pids:
            if i in pkg_map:
                pkgs.append(pkg_map[i])
            else:
                pkgs.append({"id": i, "g_code": rd["g_code"], "logis_num": "", "product_name": "", "weight": 0})
                missing_pkg_count += 1
        # 若 package_ids 字串完全解析不出任何整數 → 真的沒辦法產，跳過
        # （容錯解析已支援 "5.0" 型髒資料；走到這代表原始值真的空/全壞）
        if not pkgs:
            print(f"[export] ⚠️ shipment id={rd.get('id')} g_code={rd.get('g_code')} 的 "
                  f"package_ids 解析後為空，跳過（原始值={rd.get('package_ids')!r}）", flush=True)
            continue

        # Fallback 順序：shipment_requests.ship_* → members 表 → Shopify cache → customer_name
        ship_recipient = _safe_str(rd.get("ship_recipient"))
        ship_phone     = _safe_str(rd.get("ship_phone"))
        ship_address   = _safe_str(rd.get("ship_address"))
        if not (ship_recipient and ship_phone and ship_address):
            g_code = rd["g_code"]
            fallback_src = members_map.get(g_code) or shopify_map.get(g_code) or {}
            if not ship_recipient: ship_recipient = _safe_str(fallback_src.get("name")) or _safe_str(rd.get("customer_name"))
            if not ship_phone:     ship_phone     = _safe_str(fallback_src.get("phone"))
            if not ship_address:   ship_address   = _safe_str(fallback_src.get("address"))
            # 如果填到任何值，順手寫回 DB（下次匯出不用再 fallback）
            if ship_recipient or ship_phone or ship_address:
                fallback_updates.append((ship_recipient, ship_phone, ship_address, rd["id"]))

        shipments.append({
            "id":                   rd["id"],
            "g_code":               rd["g_code"],
            "ship_recipient":       ship_recipient,
            "ship_phone":           ship_phone,
            "ship_address":         ship_address,
            "billed_weight":        rd.get("billed_weight") or 0,
            "total_fee":            rd.get("total_fee") or 0,
            # 出貨追蹤號碼（多箱換行）→ Nigel 填「清關號碼」/ JpD 填「JpD包裹ID」
            "tracking_num":         _safe_str(rd.get("tracking_num")),
            # 打包日期來源：admin 標記出貨時的 updated_at（fallback 到客戶申請的 created_at）
            "updated_at":           rd.get("updated_at") or "",
            "created_at":           rd.get("created_at") or "",
            "packages":             pkgs,
        })

    if not shipments:
        conn.close()
        return jsonify({"success": False, "error": "選定的單沒有包裹資料"}), 400

    # 產生 Excel
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    headers, table_rows = vendor_templates.build_rows(vendor_id, shipments)

    wb = Workbook()
    ws = wb.active
    ws.title = vendor["display_name"]

    # 標頭樣式
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    hdr_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    hdr_align = Alignment(horizontal="center", vertical="center")
    thin = Border(left=Side(style="thin"), right=Side(style="thin"),
                  top=Side(style="thin"), bottom=Side(style="thin"))

    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = hdr_font; cell.fill = hdr_fill
        cell.alignment = hdr_align; cell.border = thin

    for row_idx, row_data in enumerate(table_rows, start=2):
        for col_idx, val in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin
            # 多箱追蹤號碼是換行字串 → 開自動換行才不會擠成一行
            if isinstance(val, str) and "\n" in val:
                cell.alignment = Alignment(wrap_text=True, vertical="center")

    # 欄寬自動
    for col_idx, h in enumerate(headers, start=1):
        max_len = len(str(h))
        for row_data in table_rows:
            v = row_data[col_idx - 1] if col_idx - 1 < len(row_data) else ""
            if v is not None:
                max_len = max(max_len, len(str(v)))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 4, 40)

    # 標記 exported
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    batch_id = f"{vendor_id}-{datetime.now().strftime('%Y%m%d%H%M%S')}"
    shipment_ids_actually_used = [s["id"] for s in shipments]
    ph = ",".join(["?"] * len(shipment_ids_actually_used))
    conn.execute(
        f"""UPDATE shipment_requests
            SET exported_at=?, exported_vendor=?, exported_batch_id=?
            WHERE id IN ({ph})
        """,
        [now, vendor_id, batch_id] + shipment_ids_actually_used
    )
    # 存 export_code（{g_code}-{MMDD}，與 Excel 客戶編號一致）供台灣配送貨況比對
    conn.executemany(
        "UPDATE shipment_requests SET export_code=? WHERE id=?",
        [(vendor_templates.export_code_for(s), s["id"]) for s in shipments]
    )
    # 順手把 fallback 出來的 ship_* 值寫回（下次匯出不用再算）
    if fallback_updates:
        conn.executemany(
            "UPDATE shipment_requests SET ship_recipient=?, ship_phone=?, ship_address=? WHERE id=?",
            fallback_updates
        )
        print(f"[export] fallback 補回 ship_* 欄位 {len(fallback_updates)} 筆", flush=True)
    conn.commit()
    conn.close()

    # 輸出檔案
    filename = vendor_templates.filename_for(vendor_id)
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    from urllib.parse import quote
    response = make_response(bio.read())
    response.headers["Content-Type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    response.headers["Content-Disposition"] = f'attachment; filename="{quote(filename)}"; filename*=UTF-8\'\'{quote(filename)}'
    response.headers["X-Batch-Id"] = batch_id
    response.headers["X-Shipments-Exported"] = str(len(shipments))
    if missing_pkg_count:
        response.headers["X-Missing-Packages"] = str(missing_pkg_count)
        print(f"[export] ⚠️ 本批有 {missing_pkg_count} 個包裹資料缺失（用 stub 撐過），shipments={len(shipments)} 筆", flush=True)
    return response


@app.route("/api/admin/exports/history", methods=["GET"])
def admin_exports_history():
    """檢視歷史批次（最近 50 批）"""
    if not is_super_admin():
        return jsonify({"success": False, "error": "權限不足"}), 403
    conn = get_db()
    rows = conn.execute("""
        SELECT exported_batch_id, exported_vendor, MIN(exported_at) AS exported_at,
               COUNT(*) AS shipment_count,
               SUM(billed_weight) AS total_kg,
               SUM(total_fee) AS total_fee
        FROM shipment_requests
        WHERE exported_at IS NOT NULL AND exported_at != ''
        GROUP BY exported_batch_id, exported_vendor
        ORDER BY MIN(exported_at) DESC
        LIMIT 50
    """).fetchall()
    conn.close()
    return jsonify({
        "success": True,
        "batches": [dict(r) for r in rows]
    })


@app.route("/api/admin/shipment_requests", methods=["GET"])
def admin_get_shipment_requests():
    """管理員查看所有出貨申請（含對應客戶的待處理預報資料）"""
    maybe_auto_sync()  # 後台有人活動時，距上次同步>24h 就背景同步台灣配送貨況
    status = request.args.get("status", "")
    aid = get_current_agent_id()
    # 代理過濾片段
    af = " AND agent_id=?" if aid > 0 else ""
    aparams = (aid,) if aid > 0 else ()
    try:
        conn = get_db()
        if status == "已付款":
            sql = "SELECT * FROM shipment_requests WHERE status='已出貨' AND payment_last5 != '' AND payment_last5 IS NOT NULL" + af + " ORDER BY payment_at DESC, id DESC"
            rows = conn.execute(sql, aparams).fetchall()
        elif status == "recent":
            if aid > 0:
                rows = conn.execute(
                    "SELECT * FROM shipment_requests WHERE agent_id=? ORDER BY id DESC LIMIT 50", (aid,)
                ).fetchall()
            else:
                rows = conn.execute("SELECT * FROM shipment_requests ORDER BY id DESC LIMIT 50").fetchall()
        elif status:
            sql = "SELECT * FROM shipment_requests WHERE status=?" + af + " ORDER BY id DESC"
            rows = conn.execute(sql, (status,) + aparams).fetchall()
        else:
            if aid > 0:
                rows = conn.execute(
                    "SELECT * FROM shipment_requests WHERE agent_id=? ORDER BY id DESC LIMIT 50", (aid,)
                ).fetchall()
            else:
                rows = conn.execute("SELECT * FROM shipment_requests ORDER BY id DESC LIMIT 50").fetchall()

        # 一次撈出涉及到的客戶的待處理預報（避免 N+1 查詢）
        g_codes = list({r["g_code"] for r in rows if r["g_code"]})
        forecast_map = {}
        if g_codes:
            placeholders = ",".join(["?"] * len(g_codes))
            fc_rows = conn.execute(
                f"SELECT * FROM forecasts WHERE g_code IN ({placeholders}) AND status='待處理' ORDER BY id",
                g_codes
            ).fetchall()
            for fc in fc_rows:
                try:
                    items = json.loads(fc["items_json"] or "[]")
                except (TypeError, ValueError, json.JSONDecodeError):
                    items = []
                forecast_map.setdefault(fc["g_code"], []).append({
                    "id": fc["id"],
                    "note": fc["note"] or "",
                    "created_at": fc["created_at"] or "",
                    "items": items
                })

        # 每筆出貨單的「信件」件數 → 帳單自動帶入信件費（件數 × NT$20）
        req_pids = {}
        all_pids = set()
        for r in rows:
            pids = _parse_pkg_ids(r["package_ids"])
            req_pids[r["id"]] = pids
            all_pids.update(pids)
        letter_ids = set()
        if all_pids:
            ph = ",".join(["?"] * len(all_pids))
            trows = conn.execute(
                f"SELECT id FROM packages WHERE id IN ({ph}) AND pkg_type='信件'",
                list(all_pids)
            ).fetchall()
            letter_ids = {t["id"] for t in trows}

        conn.close()

        result = []
        for r in rows:
            d = dict(r)
            d["pending_forecasts"] = forecast_map.get(r["g_code"], [])
            d["letter_count"] = sum(1 for pid in req_pids.get(r["id"], []) if pid in letter_ids)
            result.append(d)
        return jsonify({"success": True, "requests": result})
    except Exception as e:
        return jsonify({"success": False, "error": str(e), "requests": []})


@app.route("/api/admin/shipment_requests/<int:req_id>", methods=["PUT"])
def admin_update_shipment_request(req_id):
    """管理員更新出貨申請狀態（含帳單資訊）"""
    ok, _row = check_record_ownership("shipment_requests", req_id)
    if not _row:
        return jsonify({"success": False, "error": "找不到該申請"})
    if not ok:
        return jsonify({"success": False, "error": "權限不足"}), 403
    data = request.json
    status = data.get("status", "")
    admin_note = data.get("admin_note", "")
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    conn = get_db()

    # 帳單欄位（出貨時填寫）
    billed_weight = data.get("billed_weight", 0)
    rate_per_kg = data.get("rate_per_kg", 0)
    shipping_fee = data.get("shipping_fee", 0)
    handling_fee = data.get("handling_fee", 0)
    consolidation_fee = data.get("consolidation_fee", 0)
    letter_fee = data.get("letter_fee", 0)
    total_fee = data.get("total_fee", 0)
    tracking_num = data.get("tracking_num", "")
    extra_services = json.dumps(data.get("extra_services", []), ensure_ascii=False)

    # 代理可自由設定費率，無下限（min_rate 僅為 UI 預設值參考）

    if status == "已出貨" and billed_weight:
        conn.execute(
            """UPDATE shipment_requests 
               SET status=?, admin_note=?, updated_at=?,
                   billed_weight=?, rate_per_kg=?, shipping_fee=?, handling_fee=?, consolidation_fee=?, letter_fee=?, total_fee=?,
                   tracking_num=?, extra_services=?
               WHERE id=?""",
            (status, admin_note, now, billed_weight, rate_per_kg, shipping_fee, handling_fee, consolidation_fee, letter_fee, total_fee, tracking_num, extra_services, req_id)
        )
    else:
        conn.execute(
            "UPDATE shipment_requests SET status=?, admin_note=?, updated_at=? WHERE id=?",
            (status, admin_note, now, req_id)
        )

    # 如果管理員標記為「已出貨」，同步更新包裹狀態 + 預報標為已處理
    g_code_val = ""
    customer_name_val = ""
    if status == "已出貨":
        req = conn.execute("SELECT package_ids, g_code, customer_name FROM shipment_requests WHERE id=?", (req_id,)).fetchone()
        if req:
            g_code_val = req["g_code"]
            customer_name_val = req["customer_name"] or ""
            pkg_ids = [int(x.strip()) for x in req["package_ids"].split(",") if x.strip()]
            if pkg_ids:
                placeholders = ",".join(["?"] * len(pkg_ids))
                conn.execute(
                    f"UPDATE packages SET status='已出貨' WHERE id IN ({placeholders})", pkg_ids
                )
            # 自動把該客戶的待處理預報標為已處理
            conn.execute(
                "UPDATE forecasts SET status='已處理' WHERE g_code=? AND status='待處理'",
                (g_code_val,)
            )
    else:
        req = conn.execute("SELECT g_code, customer_name FROM shipment_requests WHERE id=?", (req_id,)).fetchone()
        if req:
            g_code_val = req["g_code"]
            customer_name_val = req["customer_name"] or ""

    conn.commit()
    conn.close()
    return jsonify({"success": True, "g_code": g_code_val, "customer_name": customer_name_val})


@app.route("/api/admin/shipment_requests/<int:req_id>/revert", methods=["POST"])
def admin_revert_shipment_request(req_id):
    """還原出貨申請：狀態回到待處理，包裹回到已到貨，清空帳單"""
    ok, _row = check_record_ownership("shipment_requests", req_id)
    if not _row:
        return jsonify({"success": False, "error": "找不到該申請"})
    if not ok:
        return jsonify({"success": False, "error": "權限不足"}), 403
    conn = get_db()
    req = conn.execute("SELECT * FROM shipment_requests WHERE id=?", (req_id,)).fetchone()

    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    conn.execute(
        """UPDATE shipment_requests 
           SET status='待處理', updated_at=?,
               billed_weight=0, rate_per_kg=0, shipping_fee=0, handling_fee=0,
               consolidation_fee=0, letter_fee=0, total_fee=0,
               tracking_num='', payment_last5='', payment_at='', extra_services=''
           WHERE id=?""",
        (now, req_id)
    )

    # 包裹狀態還原為「已到貨」
    pkg_ids_str = req["package_ids"]
    if pkg_ids_str:
        pkg_ids = [int(x.strip()) for x in pkg_ids_str.split(",") if x.strip()]
        if pkg_ids:
            placeholders = ",".join(["?"] * len(pkg_ids))
            conn.execute(
                f"UPDATE packages SET status='已到貨' WHERE id IN ({placeholders})", pkg_ids
            )

    conn.commit()
    conn.close()
    return jsonify({"success": True})


# ============ 預報包裹 API（本地存檔，不連 JPD）============

@app.route("/api/forecast_simple", methods=["POST"])
def create_forecast_simple():
    """客戶提交預報（存到本地 DB）"""
    data = request.json
    g_code = (data.get("g_code") or "").strip().upper()
    customer_name = data.get("customer_name", "")
    items = data.get("items", [])
    note = (data.get("note") or "").strip()

    if not g_code:
        return jsonify({"success": False, "error": "缺少會員編號"})
    if not items:
        return jsonify({"success": False, "error": "請至少填寫一個商品"})

    # 過濾空的
    valid_items = [i for i in items if (i.get("name") or "").strip()]
    if not valid_items:
        return jsonify({"success": False, "error": "請至少填寫一個商品名稱"})

    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    fc_agent_id = get_agent_id_for_g_code(g_code)
    conn = get_db()
    conn.execute(
        """INSERT INTO forecasts (g_code, customer_name, items_json, status, note, created_at, agent_id)
           VALUES (?, ?, ?, '待處理', ?, ?, ?)""",
        (g_code, customer_name, json.dumps(valid_items, ensure_ascii=False), note, now, fc_agent_id)
    )
    conn.commit()
    conn.close()
    return jsonify({"success": True, "message": "預報已送出！我們收到後會盡快處理。"})


@app.route("/api/my_forecasts", methods=["GET"])
def get_my_forecasts():
    """客戶查看自己的預報"""
    g_code = request.args.get("g_code", "").upper()
    if not g_code:
        return jsonify({"success": False, "error": "缺少會員編號"})
    conn = get_db()
    rows = conn.execute(
        "SELECT * FROM forecasts WHERE g_code=? ORDER BY id DESC LIMIT 20", (g_code,)
    ).fetchall()
    conn.close()
    results = []
    for r in rows:
        row = dict(r)
        try:
            row["items"] = json.loads(row.get("items_json") or "[]")
        except:
            row["items"] = []
        results.append(row)
    return jsonify({"success": True, "forecasts": results})


@app.route("/api/admin/forecasts", methods=["GET"])
def admin_get_forecasts():
    """管理員查看所有預報"""
    status = request.args.get("status", "")
    g_code = request.args.get("g_code", "").upper()
    aid = get_current_agent_id()
    af = " AND agent_id=?" if aid > 0 else ""
    aparams = (aid,) if aid > 0 else ()
    conn = get_db()
    if g_code and status:
        rows = conn.execute(
            "SELECT * FROM forecasts WHERE g_code=? AND status=?" + af + " ORDER BY id DESC",
            (g_code, status) + aparams
        ).fetchall()
    elif g_code:
        rows = conn.execute(
            "SELECT * FROM forecasts WHERE g_code=?" + af + " ORDER BY id DESC", (g_code,) + aparams
        ).fetchall()
    elif status:
        rows = conn.execute(
            "SELECT * FROM forecasts WHERE status=?" + af + " ORDER BY id DESC", (status,) + aparams
        ).fetchall()
    else:
        if aid > 0:
            rows = conn.execute("SELECT * FROM forecasts WHERE agent_id=? ORDER BY id DESC", (aid,)).fetchall()
        else:
            rows = conn.execute("SELECT * FROM forecasts ORDER BY id DESC").fetchall()
    conn.close()
    results = []
    for r in rows:
        row = dict(r)
        try:
            row["items"] = json.loads(row.get("items_json") or "[]")
        except:
            row["items"] = []
        results.append(row)
    return jsonify({"success": True, "forecasts": results})


@app.route("/api/admin/forecasts/<int:fc_id>", methods=["PUT"])
def admin_update_forecast(fc_id):
    """管理員更新預報狀態"""
    ok, _row = check_record_ownership("forecasts", fc_id)
    if not _row:
        return jsonify({"success": False, "error": "找不到該預報"})
    if not ok:
        return jsonify({"success": False, "error": "權限不足"}), 403
    data = request.json
    status = data.get("status", "")
    conn = get_db()
    conn.execute("UPDATE forecasts SET status=? WHERE id=?", (status, fc_id))
    conn.commit()
    conn.close()
    return jsonify({"success": True})


@app.route("/api/admin/forecasts/<int:fc_id>/excel")
def admin_download_forecast_excel(fc_id):
    """下載單筆預報的 JPD Excel"""
    ok, _row = check_record_ownership("forecasts", fc_id)
    if not _row:
        return "Not found", 404
    if not ok:
        return "Forbidden", 403
    conn = get_db()
    row = conn.execute("SELECT * FROM forecasts WHERE id=?", (fc_id,)).fetchone()
    conn.close()
    if not row:
        return "Not found", 404
    row = dict(row)
    try:
        items = json.loads(row.get("items_json") or "[]")
    except:
        items = []

    g_code = row["g_code"]
    today_str = datetime.now().strftime("%m%d")
    customer_order_id = f"{g_code}-{today_str}"

    wb = Workbook()
    ws = wb.active
    ws.title = "預報資料"

    # 標頭
    headers = [
        "客戶運單號", "JpD包裹ID", "運單ID", "包裹特殊服務",
        "收件人", "收件人身份證ID", "收件人詳細地址", "收件人电话号码",
        "備註", "特殊服务", "渠道ID",
        "申報人", "申報人身份證ID", "申報人詳細地址", "申報人电话号码",
        "品名", "数量", "金额", "材質", "產地", "URL/JanCode"
    ]
    hfill = PatternFill("solid", fgColor="1F4E79")
    hfont = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = hfill
        cell.font = hfont
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin

    # 資料
    for row_idx, item in enumerate(items, 2):
        data_row = [
            customer_order_id, "", "", "",
            "", "", "", "",
            row.get("note", ""), "", "40",
            "", "", "", "",
            item.get("name", ""),
            item.get("quantity", 1),
            item.get("price", 0),
            "", "Japan",
            item.get("url", "")
        ]
        for col_idx, val in enumerate(data_row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = Font(name="Arial", size=10)
            cell.border = thin
            cell.alignment = Alignment(vertical="center")

    # 欄寬
    col_widths = {1:16, 2:14, 5:12, 7:20, 8:16, 9:12, 11:8, 16:20, 17:8, 18:10, 21:30}
    for col, w in col_widths.items():
        ws.column_dimensions[chr(64+col) if col<=26 else 'A'].width = w
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"{g_code}_{today_str}_forecast.xlsx"
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5001))
    debug = os.environ.get("FLASK_DEBUG", "false").lower() == "true"
    print(f"""
    ╔═══════════════════════════════════════════════════════════╗
    ║       客人集運預報系統                                      ║
    ║       御用達 × JPD 雲倉                                     ║
    ╚═══════════════════════════════════════════════════════════╝
    🌐 服務啟動於 Port: {port}
    💱 TWD → JPY 匯率: {TWD_TO_JPY_RATE}
    """)
    app.run(host="0.0.0.0", port=port, debug=debug)
